from fastapi import FastAPI, Request
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles
from fastapi.responses import StreamingResponse, RedirectResponse
from dotenv import load_dotenv
from routers import jira
import os
import io
import csv
import json
import re
import threading
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

load_dotenv()

from routers.database import init_db
init_db()

from routers.citrix_apps import init_clone_jobs_table
init_clone_jobs_table()

from routers.vdi_cost import get_vdi_cost_report

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

app = FastAPI(title="Zinnia Infrastructure Portal")

app.mount("/static",             StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
app.mount("/infraportal/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static_prefixed")
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))

class _SafeTemplateCache:
    """Replaces Jinja2's LRUCache to handle unhashable cache keys in Python 3.14.
    Newer Jinja2 includes the globals dict in the cache key tuple, making it unhashable."""
    def __init__(self, capacity: int = 400):
        self._store: dict = {}
        self.capacity = capacity

    @staticmethod
    def _k(key):
        try:
            hash(key)
            return key
        except TypeError:
            return repr(key)

    def get(self, key, default=None):
        return self._store.get(self._k(key), default)

    def __setitem__(self, key, value):
        k = self._k(key)
        if k not in self._store and len(self._store) >= self.capacity:
            self._store.pop(next(iter(self._store)), None)
        self._store[k] = value

    def __getitem__(self, key):
        return self._store[self._k(key)]

    def __contains__(self, key):
        return self._k(key) in self._store

    def __len__(self):
        return len(self._store)

    def keys(self):   return self._store.keys()
    def values(self): return self._store.values()
    def items(self):  return self._store.items()
    def clear(self):  self._store.clear()

    # Some Jinja2 internals access ._mapping directly
    @property
    def _mapping(self): return self._store

templates.env.cache = _SafeTemplateCache()

from routers.scheduler import start_scheduler, scheduler
start_scheduler()

@app.on_event("shutdown")
def _stop_scheduler():
    # Without this, tearing down the process (uvicorn --reload picking up a file
    # change, or an IIS app pool recycle) lets Python's atexit machinery shut down
    # the scheduler's thread pool out from under it mid-run, logging a scary but
    # harmless "cannot schedule new futures after shutdown" traceback per job that
    # was about to fire. Stopping the scheduler first avoids that race.
    if scheduler.running:
        scheduler.shutdown(wait=False)

from routers.auth import (
    login_handler, callback_handler, logout_handler, unauthorized_handler,
    require_auth_check, get_current_user
)

# ── Auth routes ───────────────────────────────────────────────────────────────

@app.get("/infraportal/auth/login")
async def auth_login(request: Request):
    return await login_handler(request)

@app.get("/infraportal/auth/callback")
async def auth_callback(request: Request):
    return await callback_handler(request)

@app.get("/infraportal/auth/logout")
async def auth_logout(request: Request):
    return await logout_handler(request)

@app.get("/infraportal/auth/unauthorized")
async def auth_unauthorized(request: Request):
    return await unauthorized_handler(request, templates)

# ── Page routes ───────────────────────────────────────────────────────────────

@app.get("/infraportal")
async def infraportal_redirect():
    return RedirectResponse(url="/infraportal/")

def _ctx(request, user, active_page):
    """Build common template context including user groups."""
    from routers.auth import get_user_groups
    groups = list(get_user_groups(request))
    return {"request": request, "user": user, "user_groups": groups, "active_page": active_page}

@app.get("/")
@app.get("/infraportal/")
async def dashboard(request: Request):
    user, redirect = require_auth_check(request)
    if redirect: return redirect

    from routers.database import load_findings, get_findings_age
    from routers.findings import compute_system_health
    from routers.auth import GROUPS
    from datetime import datetime, timedelta

    findings = load_findings()
    system_health = compute_system_health(findings)

    # Jira summary stats for exec panels
    jira_stats = {
        "itsd": {"open": 0, "sev1": 0, "sev2": 0, "open_24h_plus": 0},
        "tasi": {"changes_48h": 0},
        "ito":  {"open_epics": 0, "overdue": 0, "stalled": 0},
    }
    try:
        from routers.database import get_conn
        conn = get_conn()
        now = datetime.now()
        rows = conn.execute(
            "SELECT severity, created FROM jira_tickets WHERE project='ITSD' AND status_category != 'Done'"
        ).fetchall()
        jira_stats["itsd"]["open"] = len(rows)
        for r in rows:
            if r["severity"] == "Sev-1": jira_stats["itsd"]["sev1"] += 1
            if r["severity"] == "Sev-2": jira_stats["itsd"]["sev2"] += 1
            if r["created"]:
                age_h = (now - datetime.fromisoformat(r["created"][:19])).total_seconds() / 3600
                if age_h > 24: jira_stats["itsd"]["open_24h_plus"] += 1
        cutoff = (now - timedelta(hours=48)).strftime("%Y-%m-%dT%H:%M:%S")
        cnt = conn.execute(
            "SELECT COUNT(*) FROM jira_tickets WHERE project='TASI' AND created >= ?", (cutoff,)
        ).fetchone()[0]
        jira_stats["tasi"]["changes_48h"] = cnt
        rows = conn.execute(
            "SELECT due_date, updated FROM jira_tickets WHERE project='ITO' AND issue_type='Epic' AND status_category != 'Done'"
        ).fetchall()
        jira_stats["ito"]["open_epics"] = len(rows)
        today = now.date()
        for r in rows:
            try:
                if r["due_date"] and datetime.fromisoformat(r["due_date"][:10]).date() < today:
                    jira_stats["ito"]["overdue"] += 1
            except Exception:
                pass
            try:
                if r["updated"] and (now - datetime.fromisoformat(r["updated"][:19])).days > 30:
                    jira_stats["ito"]["stalled"] += 1
            except Exception:
                pass
        conn.close()
    except Exception:
        pass

    groups = list(_ctx(request, user, "dashboard").get("user_groups", []))
    is_admin = GROUPS["admin"] in groups

    findings_ts = get_findings_age()
    findings_age_str = ""
    if findings_ts:
        age_min = int((datetime.now().timestamp() - findings_ts) / 60)
        findings_age_str = f"{age_min}m ago" if age_min < 60 else f"{age_min // 60}h ago"

    # Live metrics for exec strip
    metrics = {}
    try:
        from routers.cache import cache as _cache
        vm_data, _    = _cache.get("all_vms")
        host_data, _  = _cache.get("all_hosts")
        citrix, _     = _cache.get("citrix_summary")
        ad_sum, _     = _cache.get("ad_summary")
        opm_alarms, _ = _cache.get("opm_alarms")
        cert_sum, _   = _cache.get("ca_summary")
        ls_sum, _     = _cache.get("ls_asset_summary")

        if vm_data:
            vms = vm_data.get("vms") or (vm_data if isinstance(vm_data, list) else [])
            metrics["vm_count"]      = len(vms)
            metrics["vm_powered_on"] = sum(1 for v in vms if (v.get("power_state") or "").upper() == "POWEREDON")
        if host_data:
            hosts = host_data.get("hosts") or (host_data if isinstance(host_data, list) else [])
            metrics["host_count"] = len(hosts)
        if citrix:
            metrics["citrix_sessions"]     = citrix.get("active_sessions") or citrix.get("total_sessions") or 0
            metrics["citrix_machines"]     = citrix.get("total_machines") or 0
            metrics["citrix_unregistered"] = citrix.get("unregistered") or 0
            metrics["citrix_faults"]       = citrix.get("with_faults") or 0
            metrics["citrix_maintenance"]  = citrix.get("in_maintenance") or 0
        if ad_sum:
            metrics["ad_users"]       = ad_sum.get("active_users") or ad_sum.get("total_users") or 0
            metrics["ad_computers"]   = ad_sum.get("total_computers") or 0
            metrics["ad_stale_users"] = ad_sum.get("stale_users") or 0
            metrics["ad_pwd_never"]   = ad_sum.get("pwd_never_expires") or 0
        if opm_alarms:
            alarms = opm_alarms if isinstance(opm_alarms, list) else []
            metrics["opm_sev1"]    = sum(1 for a in alarms if (a.get("severity_num") or 5) == 1)
            metrics["opm_sev2"]    = sum(1 for a in alarms if (a.get("severity_num") or 5) == 2)
            metrics["opm_minor"]   = sum(1 for a in alarms if (a.get("severity_num") or 5) == 3)
            metrics["opm_total"]   = len(alarms)
            metrics["opm_critical"] = sum(1 for a in alarms if (a.get("severity_num") or 5) <= 1 and not a.get("acknowledged"))
        if cert_sum:
            metrics["certs_expiring_30"] = cert_sum.get("expiring_30") or cert_sum.get("expiring30") or 0
            metrics["certs_expired"]     = cert_sum.get("expired") or 0
        ca_exp, _ = _cache.get("ca_expiring")
        if ca_exp:
            exp_list = ca_exp if isinstance(ca_exp, list) else ca_exp.get("certs", [])
            metrics["certs_expiring_14"] = sum(
                1 for c in exp_list
                if isinstance(c.get("days_remaining"), (int, float)) and 0 < c.get("days_remaining", 999) <= 14
            )
        if ls_sum:
            metrics["asset_count"]       = ls_sum.get("total") or 0
            metrics["eol_count"]         = ls_sum.get("eol_count") or 0
            metrics["asset_not_seen_30d"] = ls_sum.get("not_seen_30d") or 0
    except Exception:
        pass

    ctx = _ctx(request, user, "dashboard")
    ctx.update({
        "findings":          findings,
        "system_health":     system_health,
        "jira_stats":        jira_stats,
        "is_admin":          is_admin,
        "findings_age_str":  findings_age_str,
        "metrics":           metrics,
    })
    return templates.TemplateResponse(request, "dashboard.html", ctx)


@app.get("/infraportal/api/dashboard/findings")
async def api_dashboard_findings(request: Request):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import load_findings
        findings = load_findings()
        return {"status": "ok", "findings": findings}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.get("/infraportal/api/jira/itsd-stats")
async def api_jira_itsd_stats(request: Request):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import get_conn
        from datetime import datetime, timedelta
        conn = get_conn()
        now = datetime.now()
        rows = conn.execute(
            "SELECT severity, created FROM jira_tickets WHERE project='ITSD' AND status_category != 'Done'"
        ).fetchall()
        conn.close()
        open_total = len(rows)
        sev1 = sum(1 for r in rows if (r["severity"] or "").startswith("Sev-1"))
        sev2 = sum(1 for r in rows if (r["severity"] or "").startswith("Sev-2"))
        cutoff_24h = (now - timedelta(hours=24)).strftime("%Y-%m-%dT%H:%M:%S")
        overdue = sum(1 for r in rows if r["created"] and r["created"] < cutoff_24h)
        return {"status": "ok", "open": open_total, "sev1": sev1, "sev2": sev2, "overdue": overdue}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/infraportal/api/dashboard/refresh-findings")
async def api_refresh_findings(request: Request):
    user, redirect = require_auth_check(request)
    if redirect: return {"status": "unauthorized"}
    try:
        from routers.findings import refresh_findings
        import threading
        threading.Thread(target=refresh_findings, daemon=True).start()
        return {"status": "started"}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/infraportal/api/dashboard/finding-ai")
async def api_finding_ai(request: Request):
    user, redirect = require_auth_check(request)
    if redirect: return {"status": "unauthorized"}
    try:
        body = await request.json()
        finding_id = body.get("finding_id")
        from routers.database import load_findings, get_conn
        findings = load_findings()
        finding = next((f for f in findings if f["id"] == finding_id), None)
        if not finding:
            return {"status": "error", "error": "Finding not found"}

        meta = {}
        try:
            meta = json.loads(finding.get("meta_json") or "{}")
        except Exception:
            pass

        # Pull related ITSD tickets for context
        related = []
        try:
            conn = get_conn()
            device_hint = meta.get("vm_name") or meta.get("device") or meta.get("subject") or ""
            if device_hint:
                kw = device_hint.split(".")[0].upper()
                rows = conn.execute("""
                    SELECT key, summary, severity, status, created FROM jira_tickets
                    WHERE project='ITSD' AND (
                        UPPER(server_name) LIKE ? OR UPPER(summary) LIKE ?
                    ) ORDER BY created DESC LIMIT 5
                """, (f"%{kw}%", f"%{kw}%")).fetchall()
                related = [dict(r) for r in rows]
            conn.close()
        except Exception:
            pass

        import httpx
        from anthropic import Anthropic
        client = Anthropic(
            api_key=os.getenv("CLAUDE_API_KEY"),
            http_client=httpx.Client(verify=False, timeout=60.0),
        )
        prompt = (
            f"You are an infrastructure analyst. Provide a concise diagnostic assessment "
            f"for this finding in 3-5 bullet points. Be specific about likely causes, "
            f"immediate actions, and whether this needs a ticket.\n\n"
            f"Finding: {finding['title']}\n"
            f"System: {finding['system']} / {finding['category']}\n"
            f"Detail: {finding['detail']}\n"
            f"Data: {json.dumps(meta, indent=2)}\n"
        )
        if related:
            prompt += f"\nRelated ITSD tickets:\n" + "\n".join(
                f"- {r['key']}: {r['summary']} ({r['severity']} / {r['status']})" for r in related
            )
        resp = client.messages.create(
            model="claude-sonnet-4-6",
            max_tokens=500,
            messages=[{"role": "user", "content": prompt}],
        )
        return {"status": "ok", "analysis": resp.content[0].text}
    except Exception as e:
        return {"status": "error", "error": str(e)}

@app.get("/vmware")
@app.get("/infraportal/vmware")
async def vmware_page(request: Request):
    user, redirect = require_auth_check(request, "vmware")
    if redirect: return redirect
    return templates.TemplateResponse(request, "vmware.html", _ctx(request, user, "vmware"))

@app.get("/analysis")
@app.get("/infraportal/analysis")
async def analysis_page(request: Request):
    user, redirect = require_auth_check(request)
    if redirect: return redirect
    return templates.TemplateResponse(request, "analysis.html", _ctx(request, user, "analysis"))

@app.get("/alerts")
@app.get("/infraportal/alerts")
async def alerts_page(request: Request):
    user, redirect = require_auth_check(request)
    if redirect: return redirect
    return templates.TemplateResponse(request, "alerts.html", _ctx(request, user, "alerts"))

@app.get("/active-directory")
@app.get("/infraportal/active-directory")
async def ad_page(request: Request):
    user, redirect = require_auth_check(request)
    if redirect: return redirect
    return templates.TemplateResponse(request, "active_directory.html", _ctx(request, user, "active-directory"))

@app.get("/entra")
@app.get("/infraportal/entra")
async def entra_page(request: Request):
    user, redirect = require_auth_check(request)
    if redirect: return redirect
    try:
        from routers.cache import cache
        from datetime import datetime
        licenses_raw,    _ = cache.get("entra_licenses")
        users_raw,       _ = cache.get("entra_users")
        groups_raw,      _ = cache.get("entra_groups")
        apps_raw,        _ = cache.get("entra_apps")
        mailbox_raw,     _ = cache.get("entra_mailbox")
        risky_raw,       _ = cache.get("entra_risky")
        ca_raw,          _ = cache.get("entra_ca_policies")
        fetched_at = None
        if users_raw and users_raw.get("fetched_at"):
            ts = users_raw["fetched_at"]
            age_min = int((datetime.now().timestamp() - ts) / 60)
            fetched_at = f"{age_min}m ago" if age_min < 60 else f"{age_min // 60}h ago"
        ctx = _ctx(request, user, "entra")
        ctx.update({
            "licenses":     licenses_raw or [],
            "users_data":   users_raw    or {},
            "groups_data":  groups_raw   or {},
            "apps_data":    apps_raw     or {},
            "mailbox_data": mailbox_raw  or {},
            "risky_users":  risky_raw    or [],
            "ca_policies":  ca_raw       or [],
            "fetched_at":   fetched_at,
        })
    except Exception:
        ctx = _ctx(request, user, "entra")
        ctx.update({"licenses":[],"users_data":{},"groups_data":{},"apps_data":{},
                    "mailbox_data":{},"risky_users":[],"ca_policies":[],"fetched_at":None})
    return templates.TemplateResponse(request, "entra.html", ctx)

@app.post("/infraportal/api/entra/refresh")
async def api_entra_refresh(request: Request):
    user, redirect = require_auth_check(request)
    if redirect: return {"status": "unauthorized"}
    try:
        from routers.entra import refresh_all
        import threading
        threading.Thread(target=refresh_all, daemon=True).start()
        return {"status": "started"}
    except Exception as e:
        return {"status": "error", "error": str(e)}

@app.get("/certificates")
@app.get("/infraportal/certificates")
async def certificates_page(request: Request):
    user, redirect = require_auth_check(request)
    if redirect: return redirect
    return templates.TemplateResponse(request, "certificates.html", _ctx(request, user, "certificates"))

@app.get("/my-dashboard")
@app.get("/infraportal/my-dashboard")
async def my_dashboard_page(request: Request):
    user, redirect = require_auth_check(request)
    if redirect: return redirect
    widget_ids = None
    try:
        from routers.database import get_user_widgets
        widget_ids = get_user_widgets((user or {}).get("email", ""))
    except Exception:
        pass
    ctx = _ctx(request, user, "my-dashboard")
    ctx["saved_widget_ids"] = json.dumps(widget_ids) if widget_ids else "null"
    return templates.TemplateResponse(request, "my_dashboard.html", ctx)

@app.get("/jira-intelligence")
@app.get("/infraportal/jira-intelligence")
async def jira_intelligence_page(request: Request):
    user, redirect = require_auth_check(request)
    if redirect: return redirect
    return templates.TemplateResponse(request, "jira_intelligence.html", _ctx(request, user, "jira-intelligence"))

@app.get("/soc-reports")
@app.get("/infraportal/soc-reports")
async def soc_reports_page(request: Request):
    user, redirect = require_auth_check(request)
    if redirect: return redirect
    return templates.TemplateResponse(request, "soc_reports.html", _ctx(request, user, "soc-reports"))

@app.post("/infraportal/api/jira-intelligence/run")
async def api_jira_intelligence_run(request: Request):
    user, redirect = require_auth_check(request)
    if redirect: return {"status": "unauthorized"}
    try:
        from routers.jira_intelligence import trigger_background, is_running
        if is_running():
            return {"status": "running"}
        trigger_background()
        return {"status": "started"}
    except Exception as e:
        return {"status": "error", "error": str(e)}

@app.get("/infraportal/api/jira-intelligence/status")
async def api_jira_intelligence_status(request: Request, days: int = 90):
    user, _ = require_auth_check(request)
    try:
        from routers.jira_intelligence import is_running, compute_stats_for_period
        from routers.database import load_latest_jira_intelligence
        running = is_running()

        if days != 90:
            stats = compute_stats_for_period(days)
            if not stats:
                return {"status": "never_run"}
            latest = load_latest_jira_intelligence()
            narrative = ""
            if latest and latest.get("analysis_json"):
                try:
                    narrative = json.loads(latest["analysis_json"]).get("narrative") or ""
                except Exception:
                    pass
            return {
                "status": "ready",
                "data": {
                    "analysis_json": {
                        "stats": stats,
                        "narrative": narrative,
                        "generated_at": stats["generated_at"],
                    },
                    "period_days": days,
                },
            }

        latest = load_latest_jira_intelligence()
        if latest and latest.get("analysis_json"):
            latest["analysis_json"] = json.loads(latest["analysis_json"])
        return {
            "status": "running" if running else ("ready" if latest else "never_run"),
            "data":   latest,
        }
    except Exception as e:
        return {"status": "error", "error": str(e)}

@app.get("/infraportal/api/jira-intelligence/itsd-team-trend")
async def api_jira_itsd_team_trend(request: Request, weeks: int = 8):
    user, _ = require_auth_check(request)
    try:
        from routers.jira_intelligence import compute_itsd_team_trend
        return {"status": "ready", "data": compute_itsd_team_trend(weeks=weeks)}
    except Exception as e:
        return {"status": "error", "error": str(e)}

@app.get("/infraportal/api/jira-intelligence/ito-closed-trend")
async def api_jira_ito_closed_trend(request: Request, bucket_weeks: int = 2, num_buckets: int = 4):
    user, _ = require_auth_check(request)
    try:
        from routers.jira_intelligence import compute_ito_closed_trend
        return {"status": "ready", "data": compute_ito_closed_trend(bucket_weeks=bucket_weeks, num_buckets=num_buckets)}
    except Exception as e:
        return {"status": "error", "error": str(e)}

# ── SOC Audit Report API ──────────────────────────────────────────────────────

@app.get("/infraportal/api/soc/project-statuses")
@app.get("/api/soc/project-statuses")
async def api_soc_project_statuses(request: Request, project: str = "TASI"):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.soc_report import fetch_project_statuses
        proj = project.strip().upper()
        statuses = fetch_project_statuses(proj)
        if not statuses:
            return {"status": "error", "project": proj,
                    "error": f"No statuses found for project '{proj}'. "
                             "Check that the project key is correct and the API token has access."}
        return {"status": "ok", "project": proj, "statuses": statuses}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.get("/infraportal/api/soc/preview")
async def api_soc_preview(request: Request, start: str = "", end: str = "",
                           statuses: str = "Validated,Pending Production Validation",
                           projects: str = "TASI"):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.soc_report import preview_tickets
        if not start:
            from datetime import datetime
            start = f"{datetime.now().year}-01-01"
        status_list  = [s.strip() for s in statuses.split(",") if s.strip()]
        project_list = [p.strip().upper() for p in projects.split(",") if p.strip()]
        rows = preview_tickets(start, end, status_list, project_list)
        return {"status": "ok", "tickets": rows, "count": len(rows)}
    except Exception as e:
        import traceback, logging
        tb = traceback.format_exc()
        err = str(e) or repr(e) or tb.splitlines()[-1]
        logging.getLogger(__name__).error("SOC preview error: %s", tb)
        return {"status": "error", "error": err}


@app.post("/infraportal/api/soc/generate")
async def api_soc_generate(request: Request):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        body = await request.json()
        keys            = body.get("ticket_keys", [])
        include_images  = body.get("include_images", True)
        date_range      = body.get("date_range", "")
        title           = body.get("title", "SOC Audit Report")
        if not keys:
            return {"status": "error", "error": "No ticket keys provided"}
        from routers.soc_report import start_report_generation
        started = start_report_generation(keys, include_images, date_range, title)
        return {"status": "started" if started else "already_running"}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.get("/infraportal/api/soc/status")
async def api_soc_status(request: Request):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.soc_report import get_report_state
        state = get_report_state()
        return {
            "running":        state["running"],
            "progress":       state["progress"],
            "total":          state["total"],
            "current_ticket": state["current_ticket"],
            "error":          state["error"],
            "has_result":     state["result_html"] is not None,
            "started_at":     state["started_at"],
            "completed_at":   state["completed_at"],
            "report_meta":    state["report_meta"],
        }
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.get("/infraportal/api/soc/download")
async def api_soc_download(request: Request):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.soc_report import get_report_state
        from fastapi.responses import Response
        state = get_report_state()
        html = state.get("result_html")
        if not html:
            return {"status": "error", "error": "No report available yet"}
        meta = state.get("report_meta", {})
        safe_title = re.sub(r'[^a-zA-Z0-9_-]', '_', meta.get("title", "SOC_Report"))
        filename = f"{safe_title}.html"
        return Response(
            content=html.encode("utf-8"),
            media_type="text/html",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.get("/infraportal/api/soc/pdf")
@app.get("/api/soc/pdf")
async def api_soc_pdf(request: Request):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.soc_report import get_report_state, generate_pdf
        from fastapi.responses import Response
        state = get_report_state()
        html = state.get("result_html")
        if not html:
            return {"status": "error", "error": "No report available yet"}
        pdf_bytes = generate_pdf(html)
        meta = state.get("report_meta", {})
        safe_title = re.sub(r'[^a-zA-Z0-9_-]', '_', meta.get("title", "SOC_Report"))
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{safe_title}.pdf"'}
        )
    except ImportError as e:
        return {"status": "error", "error": str(e), "install_hint": "pip install xhtml2pdf"}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.get("/infraportal/jira-config")
async def jira_config_page(request: Request):
    user, redirect = require_auth_check(request, "admin")
    if redirect: return redirect
    return RedirectResponse(url="/infraportal/settings#jira")


@app.get("/infraportal/settings")
async def settings_page(request: Request):
    user, redirect = require_auth_check(request, "admin")
    if redirect: return redirect
    ctx = _ctx(request, user, "settings")
    ctx.update(_build_scheduler_ctx(request, user))
    return templates.TemplateResponse(request, "settings.html", ctx)


@app.get("/infraportal/api/settings/lookup/{list_key}")
async def api_lookup_list(list_key: str, request: Request):
    """Return values for a lookup list (e.g. owner_teams). Any authenticated user may read."""
    user, redirect = require_auth_check(request)
    if redirect: return {"status": "unauthorized"}
    try:
        from routers.database import list_lookup_values
        return {"status": "ok", "values": list_lookup_values(list_key)}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.post("/infraportal/api/settings/lookup/{list_key}")
async def api_lookup_add(list_key: str, request: Request):
    """Add a value to a lookup list. Body: { value }"""
    user, _ = require_auth_check(request, "admin")
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import add_lookup_value
        body = await request.json()
        value = (body.get("value") or "").strip()
        if not value:
            return {"status": "error", "message": "value is required"}
        ok = add_lookup_value(list_key, value)
        return {"status": "ok" if ok else "error"}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.post("/infraportal/api/settings/lookup/{list_key}/{value_id}/delete")
async def api_lookup_delete(list_key: str, value_id: int, request: Request):
    user, _ = require_auth_check(request, "admin")
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import delete_lookup_value
        ok = delete_lookup_value(value_id)
        return {"status": "ok" if ok else "error"}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.post("/infraportal/api/settings/lookup/{list_key}/reorder")
async def api_lookup_reorder(list_key: str, request: Request):
    """Body: { ids: [...] }"""
    user, _ = require_auth_check(request, "admin")
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import reorder_lookup_values
        body = await request.json()
        ok = reorder_lookup_values(list_key, body.get("ids", []))
        return {"status": "ok" if ok else "error"}
    except Exception as e:
        return {"status": "error", "message": str(e)}


# ── .env editor ────────────────────────────────────────────────────────────────

@app.get("/infraportal/api/settings/env")
async def api_settings_env_get(request: Request):
    user, _ = require_auth_check(request, "admin")
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.settings import read_env_entries
        return {"status": "ok", "entries": read_env_entries()}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.post("/infraportal/api/settings/env")
async def api_settings_env_save(request: Request):
    """Body: { changes: {KEY: value, ...}, recycle: bool }
    Only keys with a non-empty value are written — blank means "leave unchanged",
    which lets the front end never round-trip masked secret values."""
    user, _ = require_auth_check(request, "admin")
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.settings import write_env_changes, recycle_app_pool
        body = await request.json()
        changes = {k: v for k, v in (body.get("changes") or {}).items() if v not in (None, "")}
        if not changes:
            return {"status": "error", "message": "No changes to save"}
        backup_path = write_env_changes(changes)
        result = {"status": "ok", "backup": backup_path, "keys_updated": list(changes.keys())}
        if body.get("recycle"):
            ok, msg = recycle_app_pool()
            result["recycled"] = ok
            result["recycle_message"] = msg
        return result
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.get("/criticality")
@app.get("/infraportal/criticality")
async def criticality_page(request: Request):
    user, redirect = require_auth_check(request, "admin")
    if redirect: return redirect
    return RedirectResponse(url="/infraportal/settings#criticality")

@app.get("/citrix")
@app.get("/infraportal/citrix")
async def citrix_page(request: Request):
    user, redirect = require_auth_check(request, "citrix")
    if redirect: return redirect
    return templates.TemplateResponse(request, "citrix.html", _ctx(request, user, "citrix"))

@app.get("/citrix/app-manager")
@app.get("/infraportal/citrix/app-manager")
async def citrix_app_manager_page(request: Request):
    user, redirect = require_auth_check(request, "citrix")
    if redirect: return redirect
    return templates.TemplateResponse(request, "citrix_app_manager.html", _ctx(request, user, "app-manager"))

@app.get("/assets")
@app.get("/infraportal/assets")
async def assets_page(request: Request):
    user, redirect = require_auth_check(request)
    if redirect: return redirect
    return templates.TemplateResponse(request, "assets.html", _ctx(request, user, "assets"))

@app.get("/network")
@app.get("/infraportal/network")
async def network_page(request: Request):
    user, redirect = require_auth_check(request, "network")
    if redirect: return redirect
    return templates.TemplateResponse(request, "network.html", _ctx(request, user, "network"))

@app.get("/vdi-cost")
@app.get("/infraportal/vdi-cost")
async def vdi_cost_page(request: Request):
    user, redirect = require_auth_check(request)
    if redirect: return redirect
    return templates.TemplateResponse(request, "vdi_cost.html", _ctx(request, user, "vdi-cost"))

# ── VMware API ────────────────────────────────────────────────────────────────

@app.get("/api/vms")
@app.get("/infraportal/api/vms")
async def api_get_vms(refresh: bool = False):
    try:
        from routers.vmware import get_all_vms
        from routers.cache import cache
        vms, errors, timestamp = get_all_vms(force_refresh=refresh)
        age = cache.age_string(timestamp)
        return {"status": "ok", "count": len(vms), "vms": vms, "errors": errors, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/hosts")
@app.get("/infraportal/api/hosts")
async def api_get_hosts(refresh: bool = False):
    try:
        from routers.vmware import get_all_hosts
        from routers.cache import cache
        hosts, errors, timestamp = get_all_hosts(force_refresh=refresh)
        age = cache.age_string(timestamp)
        return {"status": "ok", "count": len(hosts), "hosts": hosts, "errors": errors, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/detailed-vms")
@app.get("/infraportal/api/detailed-vms")
async def api_get_detailed_vms(refresh: bool = False):
    try:
        from routers.vmware import stream_detailed_vms
        return StreamingResponse(stream_detailed_vms(force_refresh=refresh), media_type="text/event-stream")
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/detailed-hosts")
@app.get("/infraportal/api/detailed-hosts")
async def api_get_detailed_hosts(refresh: bool = False):
    try:
        from routers.vmware import stream_detailed_hosts
        return StreamingResponse(stream_detailed_hosts(force_refresh=refresh), media_type="text/event-stream")
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/untagged-vms")
@app.get("/infraportal/api/untagged-vms")
async def api_get_untagged_vms(refresh: bool = False):
    try:
        from routers.vmware import stream_untagged_vms
        return StreamingResponse(stream_untagged_vms(force_refresh=refresh), media_type="text/event-stream")
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/vm-storage")
@app.get("/infraportal/api/vm-storage")
async def api_get_vm_storage(refresh: bool = False):
    try:
        from routers.vmware import stream_vm_storage
        return StreamingResponse(stream_vm_storage(force_refresh=refresh), media_type="text/event-stream")
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/vm-storage-summary")
@app.get("/infraportal/api/vm-storage-summary")
async def api_get_vm_storage_summary():
    try:
        from routers.cache import cache
        data, timestamp = cache.get("vm_storage")
        if data is None:
            return {"status": "ok", "low_disk_count": 0, "total_count": 0, "available": False}
        low_disk = sum(1 for v in data.get("vms", []) if v.get("has_low_disk"))
        return {"status": "ok", "low_disk_count": low_disk, "total_count": data.get("count", 0), "available": True}
    except Exception as e:
        return {"status": "error", "message": str(e)}

# ── OpManager API ─────────────────────────────────────────────────────────────

@app.get("/api/alerts")
@app.get("/infraportal/api/alerts")
async def api_get_alerts(refresh: bool = False):
    try:
        from routers.opmanager import get_alarms
        from routers.cache import cache
        from routers.database import save_alert_history, get_criticality_map
        # Bug #5 fix: serve from cache on normal page loads — scheduler refreshes every 5 min.
        # Only hit OpManager when the user explicitly clicks Refresh or cache is empty.
        if not refresh:
            cached, ts = cache.get("opm_alarms")
            if cached is not None:
                age = cache.age_string(ts)
                enriched = _enrich_alarms_with_criticality(cached)
                return {"status": "ok", "count": len(enriched), "alarms": enriched, "timestamp": age}
        alarms, timestamp = get_alarms(force_refresh=refresh)
        save_alert_history(alarms)
        age = cache.age_string(timestamp)
        enriched = _enrich_alarms_with_criticality(alarms)
        return {"status": "ok", "count": len(enriched), "alarms": enriched, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}


def _enrich_alarms_with_criticality(alarms: list) -> list:
    """Cross-reference alarms against the criticality registry and attach tier metadata."""
    try:
        from routers.database import get_criticality_map
        crit_map = get_criticality_map()
    except Exception:
        return alarms
    if not crit_map:
        return alarms

    enriched = []
    for alarm in alarms:
        a = dict(alarm)
        device = (a.get("device_name") or "").strip().upper()
        entry = crit_map.get(device)
        if entry:
            a["criticality_tier"]        = entry.get("tier")
            a["criticality_desc"]        = entry.get("service_description")
            a["criticality_blast"]       = entry.get("blast_radius")
            a["criticality_team"]        = entry.get("owner_team")
            a["criticality_slack"]       = entry.get("escalation_slack")
            a["criticality_email"]       = entry.get("escalation_email")
            a["criticality_singleton"]   = bool(entry.get("is_singleton"))
        else:
            a["criticality_tier"] = None
        enriched.append(a)

    # Re-sort so P1 alarms float to the very top regardless of OpManager severity order
    tier_order = {"P1": 0, "P2": 1, "P3": 2, "INFO": 3, None: 4}
    enriched.sort(key=lambda x: (tier_order.get(x.get("criticality_tier"), 4), x.get("severity_order", 99)))
    return enriched

@app.get("/api/devices")
@app.get("/infraportal/api/devices")
async def api_get_devices(refresh: bool = False):
    try:
        from routers.opmanager import get_devices
        from routers.cache import cache
        devices, timestamp = get_devices(force_refresh=refresh)
        age = cache.age_string(timestamp)
        return {"status": "ok", "count": len(devices), "devices": devices, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

# ── AI Analysis API ───────────────────────────────────────────────────────────

# Background analysis state — shared across requests
_analysis_state: dict = {"status": "idle", "result": None, "error": None}

def _run_analysis_background():
    """Run the full analysis in a background thread. Updates _analysis_state when done."""
    global _analysis_state
    try:
        from routers.cache import cache
        from routers.analysis import analyze_infrastructure, generate_grade_deepdives
        from routers.vmware import get_all_vms, get_all_hosts
        from routers.opmanager import get_alarms
        from routers.database import get_recurring_alerts, get_disk_trends, get_powered_off_vms
        from routers.active_directory import get_ad_analysis_for_ai
        from routers.ca_analysis import get_ca_analysis_for_ai
        from routers.citrix import get_citrix_analysis_for_ai

        vms, _, _   = get_all_vms()
        hosts, _, _ = get_all_hosts()

        untagged_data, _ = cache.get("untagged_vms")
        untagged = untagged_data["vms"] if untagged_data else []

        storage_data, _ = cache.get("vm_storage")
        storage = storage_data["vms"] if storage_data else []

        detailed_data, _ = cache.get("detailed_vms")
        detailed_vms = detailed_data["vms"] if detailed_data else vms

        alarms, _   = get_alarms()
        recurring   = get_recurring_alerts(min_occurrences=3)
        trends      = get_disk_trends()
        powered_off = get_powered_off_vms(days=30)
        ad_data     = get_ad_analysis_for_ai()
        cert_data   = get_ca_analysis_for_ai()
        citrix_data = get_citrix_analysis_for_ai()

        # Pull focused lens analyses from cache — None if a lens hasn't run yet
        alerts_ai, _ = cache.get("alerts_ai_analysis")
        vmware_ai, _ = cache.get("vmware_ai_analysis")
        citrix_ai, _ = cache.get("citrix_ai_analysis")
        ad_ai, _     = cache.get("ad_ai_analysis")
        assets_ai, _ = cache.get("assets_ai_analysis")

        analysis = analyze_infrastructure(
            detailed_vms, hosts, untagged, storage,
            alarms=alarms, recurring_alerts=recurring,
            disk_trends=trends, powered_off_vms=powered_off,
            ad_data=ad_data, cert_data=cert_data, citrix_data=citrix_data,
            alerts_ai_analysis=alerts_ai,
            vmware_ai_analysis=vmware_ai,
            citrix_ai_analysis=citrix_ai,
            ad_ai_analysis=ad_ai,
            assets_ai_analysis=assets_ai,
        )

        grades = {}
        grades_section = re.search(r'\[GRADES\](.*?)(\[|$)', analysis, re.DOTALL)
        if grades_section:
            for line in grades_section.group(1).split('\n'):
                m = re.match(r'^([A-Z_]+):\s*([A-F])\s*[-–]?\s*(.*)', line.strip())
                if m:
                    grades[m.group(1)] = {"letter": m.group(2), "note": m.group(3).strip()}

        if grades:
            deepdives = generate_grade_deepdives(analysis, grades)
            cache.set("last_analysis_deepdives", deepdives)

        cache.set("last_analysis", {"text": analysis})
        from routers.database import save_analysis
        save_analysis(analysis)

        _, ts = cache.get("last_analysis")
        age = cache.age_string(ts)
        _analysis_state["result"] = {"analysis": analysis, "timestamp": age}
        _analysis_state["status"] = "done"

    except Exception as e:
        _analysis_state["status"] = "error"
        _analysis_state["error"] = str(e)


@app.get("/api/analyze")
@app.get("/infraportal/api/analyze")
async def api_analyze(refresh: bool = False):
    global _analysis_state
    from routers.cache import cache

    # Always serve cached result when not explicitly refreshing
    if not refresh:
        cached, ts = cache.get("last_analysis")
        if cached:
            age = cache.age_string(ts)
            return {"status": "ok", "analysis": cached["text"], "timestamp": age, "cached": True}

    # Fresh analysis — fire off background thread and return immediately
    # so IIS doesn't time out the connection while Claude is thinking
    if _analysis_state["status"] != "running":
        _analysis_state["status"] = "running"
        _analysis_state["result"] = None
        _analysis_state["error"]  = None
        t = threading.Thread(target=_run_analysis_background, daemon=True)
        t.start()

    return {"status": "running"}


@app.get("/api/analyze/status")
@app.get("/infraportal/api/analyze/status")
async def api_analyze_status():
    """Poll this endpoint after triggering a fresh analysis."""
    global _analysis_state
    if _analysis_state["status"] == "running":
        return {"status": "running"}
    if _analysis_state["status"] == "error":
        return {"status": "error", "message": _analysis_state["error"]}
    if _analysis_state["status"] == "done" and _analysis_state["result"]:
        r = _analysis_state["result"]
        _analysis_state["status"] = "idle"  # reset so next run can start
        return {"status": "ok", "analysis": r["analysis"], "timestamp": r["timestamp"], "cached": False}
    # idle or no result yet
    from routers.cache import cache
    cached, ts = cache.get("last_analysis")
    if cached:
        age = cache.age_string(ts)
        return {"status": "ok", "analysis": cached["text"], "timestamp": age, "cached": True}
    return {"status": "running"}  # still starting up

@app.get("/api/analyze/deepdive/{grade}")
@app.get("/infraportal/api/analyze/deepdive/{grade}")
async def api_deepdive(grade: str):
    from routers.cache import cache
    cached, _ = cache.get("last_analysis_deepdives")
    if not cached:
        return {"status": "error", "message": "No deep dives available. Run analysis first."}
    key = grade.upper()
    if key not in cached:
        return {"status": "error", "message": f"No deep dive found for grade: {key}"}
    return {"status": "ok", "deepdive": cached[key]}

@app.get("/api/scheduler-status")
@app.get("/infraportal/api/scheduler-status")
async def scheduler_status():
    try:
        from routers.scheduler import scheduler, _job_history, JOB_LABELS, get_trigger_info
        jobs = []
        for job in scheduler.get_jobs():
            history = _job_history.get(job.id, {})
            ti = get_trigger_info(job)
            jobs.append({
                "id":          job.id,
                "label":       JOB_LABELS.get(job.id, job.id),
                "next_run":    str(job.next_run_time) if job.next_run_time else None,
                "last_run":    history.get("last_run"),
                "last_status": history.get("last_status"),
                "last_error":  history.get("last_error"),
                "run_count":   history.get("run_count", 0),
                "trigger":     ti,
                "paused":      job.next_run_time is None,
            })
        jobs.sort(key=lambda j: (0 if j["last_status"] == "error" else 1, j["label"]))
        return {"status": "ok", "jobs": jobs, "total": len(jobs)}
    except Exception as e:
        return {"status": "error", "message": str(e)}


# ── Scheduler management page ─────────────────────────────────────────────────

@app.get("/infraportal/scheduler")
async def scheduler_page(request: Request):
    user, redirect = require_auth_check(request, "admin")
    if redirect: return redirect
    return RedirectResponse(url="/infraportal/settings#scheduler")


def _build_scheduler_ctx(request: Request, user: dict) -> dict:
    """Shared job-grouping logic for the Settings page's Scheduler tab."""
    try:
        from collections import defaultdict
        from routers.scheduler import scheduler, _job_history, JOB_LABELS, JOB_GROUPS, JOB_DEFAULTS, GROUP_ORDER, get_trigger_info
        from routers.database import get_scheduler_config

        db_cfg = get_scheduler_config()
        jobs_by_group = defaultdict(list)
        for job in scheduler.get_jobs():
            history = _job_history.get(job.id, {})
            ti = get_trigger_info(job)
            dc = db_cfg.get(job.id, {})
            paused = job.next_run_time is None
            jobs_by_group[JOB_GROUPS.get(job.id, "Other")].append({
                "id":              job.id,
                "label":           JOB_LABELS.get(job.id, job.id).split(" — ", 1)[-1],
                "next_run":        str(job.next_run_time)[:19] if job.next_run_time else None,
                "last_run":        history.get("last_run"),
                "last_status":     history.get("last_status"),
                "last_error":      history.get("last_error"),
                "run_count":       history.get("run_count", 0),
                "trigger":         ti,
                "paused":          paused,
                "db_interval":     dc.get("interval_minutes"),
                "default_minutes": JOB_DEFAULTS.get(job.id, {}).get("minutes"),
            })

        grouped = {g: jobs_by_group[g] for g in GROUP_ORDER if g in jobs_by_group}
        for g, jobs in jobs_by_group.items():
            if g not in grouped:
                grouped[g] = jobs

        total   = sum(len(v) for v in grouped.values())
        errors  = sum(1 for v in grouped.values() for j in v if j["last_status"] == "error")
        paused  = sum(1 for v in grouped.values() for j in v if j["paused"])
        never   = sum(1 for v in grouped.values() for j in v if not j["last_run"])

        return {"grouped_jobs": grouped, "total": total, "errors": errors,
                "paused": paused, "never_run": never}
    except Exception as e:
        return {"grouped_jobs": {}, "total": 0, "errors": 0, "paused": 0,
                "never_run": 0, "load_error": str(e)}


@app.post("/infraportal/api/scheduler/run/{job_id}")
async def api_scheduler_run(job_id: str, request: Request):
    user, _ = require_auth_check(request, "admin")
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.scheduler import run_job_now
        ok = run_job_now(job_id)
        return {"status": "started" if ok else "not_found"}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/infraportal/api/scheduler/config")
async def api_scheduler_config(request: Request):
    user, _ = require_auth_check(request, "admin")
    if not user:
        return {"status": "unauthorized"}
    try:
        body = await request.json()
        job_id   = body.get("job_id")
        action   = body.get("action")  # "pause", "resume", "set_interval"
        minutes  = body.get("minutes")
        if not job_id:
            return {"status": "error", "error": "job_id required"}

        from routers.scheduler import set_job_paused, set_job_interval
        from routers.database import upsert_scheduler_config

        if action == "pause":
            ok = set_job_paused(job_id, True)
            upsert_scheduler_config(job_id, enabled=False)
            return {"status": "ok" if ok else "error"}
        elif action == "resume":
            ok = set_job_paused(job_id, False)
            upsert_scheduler_config(job_id, enabled=True)
            return {"status": "ok" if ok else "error"}
        elif action == "set_interval" and minutes:
            ok = set_job_interval(job_id, int(minutes))
            upsert_scheduler_config(job_id, interval_minutes=int(minutes))
            return {"status": "ok" if ok else "error"}
        elif action == "reset_interval":
            from routers.scheduler import JOB_DEFAULTS, scheduler
            from apscheduler.triggers.interval import IntervalTrigger as IT
            default = JOB_DEFAULTS.get(job_id, {})
            if default.get("type") == "interval":
                scheduler.reschedule_job(job_id, trigger=IT(minutes=default["minutes"]))
            upsert_scheduler_config(job_id, interval_minutes=None)
            return {"status": "ok"}
        return {"status": "error", "error": "unknown action"}
    except Exception as e:
        return {"status": "error", "error": str(e)}


# ── My Dashboard widget config API ────────────────────────────────────────────

@app.get("/infraportal/api/my-dashboard/config")
async def api_my_dashboard_config(request: Request):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import get_user_widgets
        widget_ids = get_user_widgets((user or {}).get("email", ""))
        return {"status": "ok", "widget_ids": widget_ids}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/infraportal/api/my-dashboard/config")
async def api_my_dashboard_config_save(request: Request):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        body = await request.json()
        widget_ids = body.get("widget_ids", [])
        from routers.database import save_user_widgets
        save_user_widgets((user or {}).get("email", ""), widget_ids)
        return {"status": "ok"}
    except Exception as e:
        return {"status": "error", "error": str(e)}


# ── Export API ────────────────────────────────────────────────────────────────

@app.post("/api/export/csv")
@app.post("/infraportal/api/export/csv")
async def export_csv(request: Request):
    try:
        body = await request.json()
        columns  = body.get("columns", [])
        rows     = body.get("rows", [])
        filename = body.get("filename", "export")
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(columns)
        for row in rows:
            writer.writerow(row)
        output.seek(0)
        return StreamingResponse(
            io.BytesIO(output.getvalue().encode()),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}.csv"}
        )
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.post("/api/export/excel")
@app.post("/infraportal/api/export/excel")
async def export_excel(request: Request):
    try:
        body = await request.json()
        columns  = body.get("columns", [])
        rows     = body.get("rows", [])
        filename = body.get("filename", "export")
        title    = body.get("title", "Export")

        wb = openpyxl.Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = title[:31]

        header_font      = Font(bold=True, color="FFFFFF")
        header_fill      = PatternFill(start_color="0F9B8E", end_color="0F9B8E", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_alignment

        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        for col in ws.columns:
            first_cell = col[0]
            if hasattr(first_cell, "column_letter"):
                max_length = max(len(str(cell.value or "")) for cell in col if hasattr(cell, "value"))
                ws.column_dimensions[first_cell.column_letter].width = min(max_length + 4, 50)  # type: ignore[union-attr]

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}.xlsx"}
        )
    except Exception as e:
        return {"status": "error", "message": str(e)}

# ── Active Directory API ──────────────────────────────────────────────────────

@app.get("/api/ad/summary")
@app.get("/infraportal/api/ad/summary")
async def api_ad_summary(refresh: bool = False):
    try:
        from routers.active_directory import get_ad_summary
        from routers.cache import cache
        data, timestamp = get_ad_summary(force_refresh=refresh)
        age = cache.age_string(timestamp)
        return {"status": "ok", "data": data, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/ad/stale-users")
@app.get("/infraportal/api/ad/stale-users")
async def api_ad_stale_users(refresh: bool = False):
    try:
        from routers.active_directory import get_stale_users
        from routers.cache import cache
        data, timestamp = get_stale_users(force_refresh=refresh)
        age = cache.age_string(timestamp)
        return {"status": "ok", "count": len(data), "users": data, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/ad/pwd-never-expires")
@app.get("/infraportal/api/ad/pwd-never-expires")
async def api_ad_pwd_never_expires(refresh: bool = False):
    try:
        from routers.active_directory import get_pwd_never_expires
        from routers.cache import cache
        data, timestamp = get_pwd_never_expires(force_refresh=refresh)
        age = cache.age_string(timestamp)
        return {"status": "ok", "count": len(data), "users": data, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/ad/domain-admins")
@app.get("/infraportal/api/ad/domain-admins")
async def api_ad_domain_admins(refresh: bool = False):
    try:
        from routers.active_directory import get_domain_admins
        from routers.cache import cache
        data, timestamp = get_domain_admins(force_refresh=refresh)
        age = cache.age_string(timestamp)
        return {"status": "ok", "count": len(data), "admins": data, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/ad/empty-groups")
@app.get("/infraportal/api/ad/empty-groups")
async def api_ad_empty_groups(refresh: bool = False):
    try:
        from routers.active_directory import get_empty_groups
        from routers.cache import cache
        data, timestamp = get_empty_groups(force_refresh=refresh)
        age = cache.age_string(timestamp)
        return {"status": "ok", "count": len(data), "groups": data, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/ad/stale-computers")
@app.get("/infraportal/api/ad/stale-computers")
async def api_ad_stale_computers(refresh: bool = False):
    try:
        from routers.active_directory import get_stale_computers
        from routers.cache import cache
        data, timestamp = get_stale_computers(force_refresh=refresh)
        age = cache.age_string(timestamp)
        return {"status": "ok", "count": len(data), "computers": data, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/ad/gpo-analysis")
@app.get("/infraportal/api/ad/gpo-analysis")
async def api_ad_gpo_analysis(refresh: bool = False):
    try:
        from routers.active_directory import get_gpo_analysis
        from routers.cache import cache
        data, timestamp = get_gpo_analysis(force_refresh=refresh)
        age = cache.age_string(timestamp)
        return {"status": "ok", "data": data, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

# ── Certificates / CA API ─────────────────────────────────────────────────────

@app.get("/api/ca/summary")
@app.get("/infraportal/api/ca/summary")
async def api_ca_summary(refresh: bool = False):
    try:
        import asyncio
        from routers.ca_analysis import get_cert_summary
        from routers.cache import cache
        loop = asyncio.get_event_loop()
        data, timestamp = await loop.run_in_executor(None, lambda: get_cert_summary(force_refresh=refresh))
        age = cache.age_string(timestamp)
        return {"status": "ok", "data": data, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/ca/expiring")
@app.get("/infraportal/api/ca/expiring")
async def api_ca_expiring(refresh: bool = False):
    try:
        import asyncio
        from routers.ca_analysis import get_expiring_certs
        from routers.cache import cache
        loop = asyncio.get_event_loop()
        data, timestamp = await loop.run_in_executor(None, lambda: get_expiring_certs(force_refresh=refresh))
        age = cache.age_string(timestamp)
        return {"status": "ok", "count": len(data), "certs": data, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/ca/manual")
@app.get("/infraportal/api/ca/manual")
async def api_ca_manual(refresh: bool = False):
    try:
        import asyncio
        from routers.ca_analysis import get_manual_certs
        from routers.cache import cache
        loop = asyncio.get_event_loop()
        data, timestamp = await loop.run_in_executor(None, lambda: get_manual_certs(force_refresh=refresh))
        age = cache.age_string(timestamp)
        return {"status": "ok", "count": len(data), "certs": data, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/ca/dc-certs")
@app.get("/infraportal/api/ca/dc-certs")
async def api_ca_dc_certs(refresh: bool = False):
    try:
        import asyncio
        from routers.ca_analysis import get_dc_kerberos_certs
        from routers.cache import cache
        loop = asyncio.get_event_loop()
        data, timestamp = await loop.run_in_executor(None, lambda: get_dc_kerberos_certs(force_refresh=refresh))
        age = cache.age_string(timestamp)
        return {"status": "ok", "count": len(data), "certs": data, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/ca/all")
@app.get("/infraportal/api/ca/all")
async def api_ca_all(refresh: bool = False):
    try:
        import asyncio
        from routers.ca_analysis import get_all_issued
        from routers.cache import cache
        loop = asyncio.get_event_loop()
        data, timestamp = await loop.run_in_executor(None, lambda: get_all_issued(force_refresh=refresh))
        age = cache.age_string(timestamp)
        return {"status": "ok", "count": len(data), "certs": data, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

# ── Citrix API ────────────────────────────────────────────────────────────────

@app.get("/api/citrix/summary")
@app.get("/infraportal/api/citrix/summary")
async def api_citrix_summary(refresh: bool = False):
    try:
        from routers.citrix import get_citrix_summary
        from routers.cache import cache
        data, timestamp = get_citrix_summary(force_refresh=refresh)
        age = cache.age_string(timestamp)
        return {"status": "ok", "data": data, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/citrix/machines")
@app.get("/infraportal/api/citrix/machines")
async def api_citrix_machines(refresh: bool = False):
    try:
        from routers.citrix import get_citrix_machines
        from routers.cache import cache
        data, timestamp = get_citrix_machines(force_refresh=refresh)
        age = cache.age_string(timestamp)
        return {"status": "ok", "count": len(data), "machines": data, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/citrix/sessions")
@app.get("/infraportal/api/citrix/sessions")
async def api_citrix_sessions(refresh: bool = False):
    try:
        from routers.citrix import get_citrix_sessions
        from routers.cache import cache
        data, timestamp = get_citrix_sessions(force_refresh=refresh)
        age = cache.age_string(timestamp)
        return {"status": "ok", "count": len(data), "sessions": data, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/citrix/delivery-groups")
@app.get("/infraportal/api/citrix/delivery-groups")
async def api_citrix_delivery_groups(refresh: bool = False):
    try:
        from routers.citrix import get_citrix_delivery_groups
        from routers.cache import cache
        data, timestamp = get_citrix_delivery_groups(force_refresh=refresh)
        age = cache.age_string(timestamp)
        return {"status": "ok", "count": len(data), "delivery_groups": data, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/citrix/catalogs")
@app.get("/infraportal/api/citrix/catalogs")
async def api_citrix_catalogs(refresh: bool = False):
    try:
        from routers.citrix import get_citrix_catalogs
        from routers.cache import cache
        data, timestamp = get_citrix_catalogs(force_refresh=refresh)
        age = cache.age_string(timestamp)
        return {"status": "ok", "count": len(data), "catalogs": data, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.post("/api/citrix/sessions/{session_id}/logoff")
@app.post("/infraportal/api/citrix/sessions/{session_id}/logoff")
async def api_citrix_logoff(session_id: str):
    try:
        from routers.citrix import logoff_session
        status, text = logoff_session(session_id)
        if status in (200, 204):
            return {"status": "ok", "message": "Session logged off successfully"}
        return {"status": "error", "message": f"API returned {status}: {text}"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/citrix/sessions/{session_id}/shadow")
@app.get("/infraportal/api/citrix/sessions/{session_id}/shadow")
async def api_citrix_shadow(session_id: str):
    try:
        from routers.citrix import get_shadow_url
        status, content = get_shadow_url(session_id)
        if status == 200:
            from fastapi.responses import Response
            return Response(
                content=content,
                media_type="application/x-ica",
                headers={"Content-Disposition": f"attachment; filename=shadow_{session_id}.ica"}
            )
        return {"status": "error", "message": f"Shadow not available: {status}"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/citrix/master-images")
@app.get("/infraportal/api/citrix/master-images")
async def api_citrix_master_images():
    try:
        from routers.citrix import get_master_images
        images = get_master_images()
        return {"status": "ok", "count": len(images), "images": images}
    except Exception as e:
        return {"status": "error", "message": str(e)}


# ── Citrix App Manager API ────────────────────────────────────────────────────

@app.get("/api/citrix/applications/raw-sample")
@app.get("/infraportal/api/citrix/applications/raw-sample")
async def api_citrix_applications_raw(dg_id: str):
    try:
        from routers.citrix import get_session, API_BASE
        session = get_session()
        sub_r = session.get(f"{API_BASE}/DeliveryGroups/{dg_id}/Applications", params={"limit": 3}, timeout=30)
        if sub_r.ok:
            items = sub_r.json().get("Items", [])
            return {"status": "ok", "source": "DeliveryGroups/{id}/Applications", "raw_items": items, "total_returned": len(items)}
        r = session.get(f"{API_BASE}/Applications", params={"limit": 50}, timeout=30)
        r.raise_for_status()
        all_items = r.json().get("Items", [])
        matched = [a for a in all_items if dg_id.lower() in [str(x).lower() for x in (a.get("AssociatedDeliveryGroupUuids") or [])]]
        sample = matched[:3] if matched else all_items[:3]
        return {"status": "ok", "source": "Applications (full fetch)", "dg_id_searched": dg_id,
                "total_fetched": len(all_items), "matched_count": len(matched), "raw_items": sample,
                "dg_associations_sample": [{"id": a.get("Id"), "name": a.get("Name"),
                    "AssociatedDeliveryGroupUuids": a.get("AssociatedDeliveryGroupUuids")} for a in all_items[:20]]}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.get("/api/citrix/applications")
@app.get("/infraportal/api/citrix/applications")
async def api_citrix_applications(dg_id: str):
    try:
        from routers.citrix_apps import get_applications_for_dg
        apps = get_applications_for_dg(dg_id)
        return {"status": "ok", "apps": apps, "count": len(apps)}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.get("/api/citrix/entra-search")
@app.get("/infraportal/api/citrix/entra-search")
async def api_entra_search(q: str = ""):
    try:
        from routers.citrix_apps import search_entra_principals
        results = search_entra_principals(q)
        return {"status": "ok", "results": results}
    except RuntimeError as e:
        return {"status": "error", "message": str(e)}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.post("/api/citrix/applications/clone")
@app.post("/infraportal/api/citrix/applications/clone")
async def api_clone_apps(request: Request):
    try:
        from routers.citrix_apps import clone_apps
        body   = await request.json()
        result = clone_apps(
            source_dg_id   = body["source_dg_id"],
            source_dg_name = body.get("source_dg_name", ""),
            target_dg_id   = body["target_dg_id"],
            target_dg_name = body.get("target_dg_name", ""),
            app_ids        = body["app_ids"],
            attr_flags     = body.get("attr_flags",     {}),
            folder_map     = body.get("folder_map",     {}),
            path_rules     = body.get("path_rules",     []),
            test_assignees = body.get("test_assignees", []),
        )
        return {"status": "ok", **result}
    except KeyError as e:
        return {"status": "error", "message": f"Missing required field: {e}"}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.get("/api/citrix/clone-jobs")
@app.get("/infraportal/api/citrix/clone-jobs")
async def api_get_clone_jobs():
    try:
        from routers.citrix_apps import get_clone_jobs
        return {"status": "ok", "jobs": get_clone_jobs()}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.get("/api/citrix/clone-jobs/{job_id}")
@app.get("/infraportal/api/citrix/clone-jobs/{job_id}")
async def api_get_clone_job(job_id: str):
    try:
        from routers.citrix_apps import get_clone_job
        job = get_clone_job(job_id)
        if not job:
            return {"status": "error", "message": "Job not found"}
        return {"status": "ok", "job": job}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.post("/api/citrix/clone-jobs/{job_id}/cutover")
@app.post("/infraportal/api/citrix/clone-jobs/{job_id}/cutover")
async def api_cutover(job_id: str):
    try:
        from routers.citrix_apps import run_security_cutover
        result = run_security_cutover(job_id)
        return {"status": "ok", **result}
    except ValueError as e:
        return {"status": "error", "message": str(e)}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.delete("/api/citrix/clone-jobs/{job_id}")
@app.delete("/infraportal/api/citrix/clone-jobs/{job_id}")
async def api_delete_clone_job(job_id: str):
    try:
        from routers.citrix_apps import delete_clone_job
        deleted = delete_clone_job(job_id)
        if not deleted:
            return {"status": "error", "message": "Job not found"}
        return {"status": "ok"}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.delete("/api/citrix/clone-jobs")
@app.delete("/infraportal/api/citrix/clone-jobs")
async def api_delete_clone_jobs(status: str | None = None):
    try:
        from routers.citrix_apps import delete_clone_jobs
        count = delete_clone_jobs(status=status)
        return {"status": "ok", "deleted": count}
    except Exception as e:
        return {"status": "error", "message": str(e)}


# ── Meraki API ────────────────────────────────────────────────────────────────

@app.get("/api/meraki/summary")
@app.get("/infraportal/api/meraki/summary")
async def api_meraki_summary():
    try:
        from routers.meraki import get_summary
        return get_summary()
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/meraki/networks")
@app.get("/infraportal/api/meraki/networks")
async def api_meraki_networks():
    try:
        from routers.meraki import get_networks
        return get_networks()
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/meraki/devices")
@app.get("/infraportal/api/meraki/devices")
async def api_meraki_devices(org: str = None, network_id: str = None, device_type: str = None):  # type: ignore[arg-type]
    try:
        from routers.meraki import get_devices
        return get_devices(org=org, network_id=network_id, device_type=device_type)
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/meraki/uplinks")
@app.get("/infraportal/api/meraki/uplinks")
async def api_meraki_uplinks():
    try:
        from routers.meraki import get_uplinks
        return get_uplinks()
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/meraki/clients")
@app.get("/infraportal/api/meraki/clients")
async def api_meraki_clients(network_id: str = None, limit: int = 200):  # type: ignore[arg-type]
    try:
        from routers.meraki import get_clients
        return get_clients(network_id=network_id, limit=limit)
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/meraki/lookup")
@app.get("/infraportal/api/meraki/lookup")
async def api_meraki_lookup(q: str = ""):
    try:
        from routers.meraki import lookup_endpoint
        return lookup_endpoint(q)
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.post("/api/meraki/refresh")
@app.post("/infraportal/api/meraki/refresh")
async def api_meraki_refresh():
    try:
        from routers.meraki import job_meraki_refresh
        job_meraki_refresh()
        return {"status": "ok"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

# ── Lansweeper API ────────────────────────────────────────────────────────────

@app.get("/api/lansweeper/summary")
@app.get("/infraportal/api/lansweeper/summary")
async def api_lansweeper_summary(refresh: bool = False):
    try:
        from routers.lansweeper import get_asset_summary
        from routers.cache import cache
        data, timestamp = get_asset_summary(force_refresh=refresh)
        age = cache.age_string(timestamp)
        return {"status": "ok", "data": data, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/lansweeper/patch-status")
@app.get("/infraportal/api/lansweeper/patch-status")
async def api_lansweeper_patch_status(refresh: bool = False):
    try:
        from routers.lansweeper import get_patch_status
        from routers.cache import cache
        data, timestamp = get_patch_status(force_refresh=refresh)
        age = cache.age_string(timestamp)
        return {"status": "ok", "data": data, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/lansweeper/assets")
@app.get("/infraportal/api/lansweeper/assets")
async def api_lansweeper_assets(refresh: bool = False):
    try:
        from routers.cache import cache
        from routers.database import load_lansweeper_assets, save_lansweeper_assets

        # 1. In-memory cache (fastest — set by scheduler every 6h)
        cached, ts = cache.get("ls_asset_list")
        if cached and not refresh:
            age = cache.age_string(ts)
            return {"status": "ok", "count": len(cached), "assets": cached, "timestamp": age}

        # 2. DB fallback — survives app pool recycles, avoids hitting Lansweeper on every cold start
        if not refresh:
            db_assets, db_ts = load_lansweeper_assets()
            if db_assets is not None:
                cache.set("ls_asset_list", db_assets)  # warm in-memory cache for subsequent requests
                age = cache.age_string(db_ts)
                return {"status": "ok", "count": len(db_assets), "assets": db_assets, "timestamp": age}

        # 3. Live fetch — only when refresh=True or both caches are empty (first ever run)
        from routers.lansweeper import _fetch_all_assets, _basic, _os_group, _days_ago, _EOL_OS
        fields = [
            "assetBasicInfo.name", "assetBasicInfo.type", "assetBasicInfo.typeGroup",
            "assetBasicInfo.ipAddress", "assetBasicInfo.domain",
            "assetBasicInfo.firstSeen", "assetBasicInfo.lastSeen",
            "assetBasicInfo.scannerTypes",
            "operatingSystem.name",
            "assetCustom.stateName", "assetCustom.manufacturer", "assetCustom.model",
        ]
        raw = _fetch_all_assets(fields)
        assets = []
        for item in raw:
            b = _basic(item)
            os_name = (item.get("operatingSystem") or {}).get("name", "")
            og = _os_group(b.get("type"), os_name)
            assets.append({
                "name":            b.get("name"),
                "ip":              b.get("ipAddress"),
                "type":            b.get("typeGroup") or b.get("type"),
                "os":              og,
                "domain":          b.get("domain"),
                "state":           (item.get("assetCustom") or {}).get("stateName"),
                "manufacturer":    (item.get("assetCustom") or {}).get("manufacturer"),
                "model":           (item.get("assetCustom") or {}).get("model"),
                "last_seen":       b.get("lastSeen"),
                "first_seen":      b.get("firstSeen"),
                "days_since_seen": _days_ago(b.get("lastSeen")),
                "scanner_types":   b.get("scannerTypes") or [],
                "is_eol":          og in _EOL_OS,
            })
        cache.set("ls_asset_list", assets)
        save_lansweeper_assets(assets)
        _, ts = cache.get("ls_asset_list")
        age = cache.age_string(ts)
        return {"status": "ok", "count": len(assets), "assets": assets, "timestamp": age}
    except Exception as e:
        return {"status": "error", "message": str(e)}

# ── Unified Asset Detail ──────────────────────────────────────────────────────

@app.get("/api/asset/{name}")
@app.get("/infraportal/api/asset/{name}")
async def api_asset_detail(name: str):
    import re as _re
    result = {"name": name, "sources": {}}

    try:
        from routers.lansweeper import get_asset_detail
        ls_data, _ = get_asset_detail(name)
        result["sources"]["lansweeper"] = ls_data
        if ls_data:
            result["ip"]              = ls_data.get("ip")
            result["os"]              = ls_data.get("os")
            result["os_version"]      = ls_data.get("os_version")
            result["is_eol"]          = ls_data.get("is_eol")
            result["type"]            = ls_data.get("type")
            result["manufacturer"]    = ls_data.get("manufacturer")
            result["model"]           = ls_data.get("model")
            result["cpu"]             = ls_data.get("cpu")
            result["cpu_cores"]       = ls_data.get("cpu_cores")
            result["total_disk_gb"]   = ls_data.get("total_disk_gb")
            result["free_disk_gb"]    = ls_data.get("free_disk_gb")
            result["used_disk_pct"]   = ls_data.get("used_disk_pct")
            result["last_seen"]       = ls_data.get("last_seen")
            result["days_since_seen"] = ls_data.get("days_since_seen")
            result["state"]           = ls_data.get("state")
    except Exception as e:
        result["sources"]["lansweeper"] = {"error": str(e)}

    try:
        from routers.cache import cache
        vm_data, _ = cache.get("detailed_vms")
        if vm_data:
            vms = vm_data.get("vms", [])
            vm = next((v for v in vms if (v.get("name") or "").upper() == name.upper()), None)
            if vm:
                result["sources"]["vmware"] = vm
                result["power_state"]  = vm.get("power_state")
                result["environment"]  = vm.get("environment")
                result["cpu_count"]    = vm.get("cpu_count")
                result["ram_mb"]       = vm.get("memory_size_mb")
                result["vmware_tools"] = vm.get("tools_status")
                result["snapshots"]    = vm.get("snapshot_count", 0)
                result["host"]         = vm.get("host")
                result["cluster"]      = vm.get("cluster")
            else:
                result["sources"]["vmware"] = None
    except Exception as e:
        result["sources"]["vmware"] = {"error": str(e)}

    try:
        from routers.citrix import get_citrix_machine_name_set
        ctx_machines = get_citrix_machine_name_set()
        ctx = ctx_machines.get(name.lower())
        result["sources"]["citrix"] = ctx or None
        if ctx:
            result["citrix_catalog"]   = ctx.get("catalog_name")
            result["citrix_dg"]        = ctx.get("delivery_group_name")
            result["citrix_reg_state"] = ctx.get("registration_state")
            result["citrix_sessions"]  = ctx.get("session_count", 0)
    except Exception as e:
        result["sources"]["citrix"] = {"error": str(e)}

    try:
        from routers.cache import cache
        stale_data, _ = cache.get("ad_stale_computers")
        if stale_data:
            ad_match = next((c for c in stale_data if (c.get("name") or "").upper() == name.upper()), None)
            result["sources"]["active_directory"] = ad_match or None
            if ad_match:
                result["ad_stale"]         = True
                result["ad_days_inactive"] = ad_match.get("days_inactive")
                result["ad_last_login"]    = ad_match.get("last_login")
    except Exception as e:
        result["sources"]["active_directory"] = {"error": str(e)}

    try:
        from routers.cache import cache
        alarms_data, _ = cache.get("opm_alarms")
        active_alerts = []
        if alarms_data:
            for alarm in alarms_data:
                dev      = alarm.get("device_name", "")
                short    = _re.split(r'\.', dev)[0].upper()
                ip       = alarm.get("ip_address")
                asset_ip = result.get("ip")
                if short == name.upper() or (asset_ip and ip and ip == asset_ip):
                    active_alerts.append(alarm)
        result["sources"]["opmanager"]   = active_alerts
        result["active_alert_count"]     = len(active_alerts)
        result["critical_alert_count"]   = sum(1 for a in active_alerts if a.get("severity") == "Critical")
    except Exception as e:
        result["sources"]["opmanager"] = {"error": str(e)}

    recommendations = []
    if result.get("is_eol"):
        recommendations.append({
            "priority": "critical", "category": "Patching",
            "text": f"OS is end-of-life ({result.get('os')}). Schedule upgrade immediately.",
        })
    if (result.get("days_since_seen") or 0) > 30:
        recommendations.append({
            "priority": "warning", "category": "Availability",
            "text": f"Asset not seen in {result['days_since_seen']} days. Verify it is still active.",
        })
    if (result.get("used_disk_pct") or 0) > 85:
        recommendations.append({
            "priority": "warning", "category": "Storage",
            "text": f"Disk usage at {result['used_disk_pct']}%. Consider cleanup or expansion.",
        })
    if result.get("snapshots") and result["snapshots"] > 3:
        recommendations.append({
            "priority": "warning", "category": "VMware",
            "text": f"{result['snapshots']} snapshots detected. Remove stale snapshots to reclaim disk.",
        })
    if result.get("ad_stale"):
        recommendations.append({
            "priority": "info", "category": "Active Directory",
            "text": f"Computer account stale in AD ({result.get('ad_days_inactive')} days). Verify machine is active or disable the account.",
        })
    if result.get("critical_alert_count", 0) > 0:
        recommendations.append({
            "priority": "critical", "category": "Alerts",
            "text": f"{result['critical_alert_count']} active critical alert(s) from OpManager. Immediate investigation required.",
        })
    if result.get("citrix_reg_state") == "Unregistered":
        recommendations.append({
            "priority": "warning", "category": "Citrix",
            "text": "Machine is unregistered in Citrix. Check VDA service and connectivity to Cloud Connector.",
        })

    result["recommendations"] = recommendations

    try:
        from routers.database import get_asset_override
        result["override"] = get_asset_override(name)
    except Exception:
        result["override"] = None

    return {"status": "ok", "asset": result}

# ── Asset Override API ────────────────────────────────────────────────────────

@app.post("/api/asset/{name}/override")
@app.post("/infraportal/api/asset/{name}/override")
async def api_set_asset_override(name: str, request: Request):
    try:
        from routers.database import save_asset_override
        body   = await request.json()
        flag   = body.get("flag", "unimportant")
        reason = body.get("reason", "")
        set_by = body.get("set_by", "portal")
        save_asset_override(name, flag=flag, reason=reason, set_by=set_by)
        return {"status": "ok", "device_name": name, "flag": flag}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.delete("/api/asset/{name}/override")
@app.delete("/infraportal/api/asset/{name}/override")
async def api_delete_asset_override(name: str):
    try:
        from routers.database import delete_asset_override
        delete_asset_override(name)
        return {"status": "ok", "device_name": name}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/api/asset-overrides")
@app.get("/infraportal/api/asset-overrides")
async def api_list_asset_overrides():
    try:
        from routers.database import list_asset_overrides
        return {"status": "ok", "overrides": list_asset_overrides()}
    except Exception as e:
        return {"status": "error", "message": str(e)}

# ── Alerts API ────────────────────────────────────────────────────────────────

def _refresh_alarm_caches_after_action():
    """After acknowledging/clearing alarms in OpManager, our own caches still hold
    the pre-action state — opm_alarms refreshes every 5min on its own schedule, and
    opm_named_groups (per-group severity, used by the NOC topology map) only rebuilds
    hourly. Force both to recompute now so the UI reflects the action immediately
    instead of looking unchanged for up to an hour."""
    try:
        from routers.opmanager import get_alarms
        from routers.cache import cache as _cache
        get_alarms(force_refresh=True)
        _cache.invalidate("opm_named_groups")
    except Exception:
        pass  # best-effort — the action itself already succeeded/failed independently


def _current_actor(request: Request) -> str:
    try:
        from routers.auth import get_current_user
        user = get_current_user(request) or {}
        return user.get("email") or user.get("name") or "unknown"
    except Exception:
        return "unknown"


def _log_alarm_actions(alarm_ids: list, action: str, actor: str, alarms_meta: list | None = None):
    """Best-effort: snapshot each alarm's metadata and record the action, so the
    Recurring Alerts report can later see what the NOC keeps having to touch.
    Never raises — logging must not block the actual ack/clear from succeeding."""
    try:
        from routers.database import log_alert_action
        meta_by_id = {str(m.get("alarm_id")): m for m in (alarms_meta or [])}
        if not meta_by_id:
            from routers.opmanager import get_alarms
            alarms, _ = get_alarms(force_refresh=False)
            meta_by_id = {str(a.get("alarm_id")): a for a in (alarms or [])}
        for aid in alarm_ids:
            m = meta_by_id.get(str(aid), {})
            log_alert_action(
                alarm_id=aid, device_name=m.get("device_name"), category=m.get("category"),
                event_type=m.get("event_type"), severity=m.get("severity"),
                message=m.get("message"), action=action, actor=actor,
            )
    except Exception:
        import logging
        logging.getLogger(__name__).warning("Failed to log alarm actions (non-fatal)", exc_info=True)


@app.post("/api/alerts/acknowledge")
@app.post("/infraportal/api/alerts/acknowledge")
async def api_acknowledge_alerts(request: Request):
    try:
        from routers.opmanager import acknowledge_alarms_bulk
        body = await request.json()
        alarm_ids = body.get("alarm_ids", [])
        if not alarm_ids:
            return {"status": "error", "message": "No alarm IDs provided"}
        success, failed, errors = acknowledge_alarms_bulk(alarm_ids)
        failed_ids = {e["alarm_id"] for e in errors}
        _log_alarm_actions([a for a in alarm_ids if a not in failed_ids], "acknowledge",
                            _current_actor(request), body.get("alarms"))
        _refresh_alarm_caches_after_action()
        return {"status": "ok", "success": success, "failed": failed, "errors": errors}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.post("/api/alerts/clear")
@app.post("/infraportal/api/alerts/clear")
async def api_clear_alerts(request: Request):
    try:
        from routers.opmanager import clear_alarms_bulk
        body = await request.json()
        alarm_ids = body.get("alarm_ids", [])
        if not alarm_ids:
            return {"status": "error", "message": "No alarm IDs provided"}
        success, failed, errors = clear_alarms_bulk(alarm_ids)
        failed_ids = {e["alarm_id"] for e in errors}
        _log_alarm_actions([a for a in alarm_ids if a not in failed_ids], "clear",
                            _current_actor(request), body.get("alarms"))
        _refresh_alarm_caches_after_action()
        return {"status": "ok", "success": success, "failed": failed, "errors": errors}
    except Exception as e:
        return {"status": "error", "message": str(e)}

# ── Bulk alarm actions (background job — safe for thousands of alarms) ───────
# The Reports > Stale Alerts "Select All" flow can select thousands of alarms.
# Clearing/acking them one HTTP call at a time to OpManager can take many minutes —
# far past any reasonable request timeout — so this runs in a background thread and
# the frontend polls /status, following the same fire-and-forget pattern used for
# AI analysis and the VDI report.

_bulk_action_state: dict = {"status": "idle", "total": 0, "completed": 0, "success": 0, "failed": 0, "errors": []}

def _run_bulk_alarm_action_background(action: str, alarm_ids: list, alarms_meta: list, actor: str):
    global _bulk_action_state
    from routers.opmanager import acknowledge_alarm, clear_alarm
    from routers.database import log_alert_action

    fn = acknowledge_alarm if action == "acknowledge" else clear_alarm
    meta_by_id = {str(m.get("alarm_id")): m for m in (alarms_meta or [])}
    for aid in alarm_ids:
        ok, msg = fn(str(aid))
        if ok:
            _bulk_action_state["success"] += 1
            m = meta_by_id.get(str(aid), {})
            try:
                log_alert_action(
                    alarm_id=aid, device_name=m.get("device_name"), category=m.get("category"),
                    event_type=m.get("event_type"), severity=m.get("severity"),
                    message=m.get("message"), action=action, actor=actor,
                )
            except Exception:
                pass
        else:
            _bulk_action_state["failed"] += 1
            _bulk_action_state["errors"].append({"alarm_id": aid, "error": str(msg)})
        _bulk_action_state["completed"] += 1
    _bulk_action_state["status"] = "done"
    _refresh_alarm_caches_after_action()


@app.post("/api/alerts/bulk-action")
@app.post("/infraportal/api/alerts/bulk-action")
async def api_bulk_alarm_action(request: Request):
    global _bulk_action_state
    try:
        body = await request.json()
        action = body.get("action")
        alarm_ids = body.get("alarm_ids") or []
        if action not in ("acknowledge", "clear"):
            return {"status": "error", "message": "action must be 'acknowledge' or 'clear'"}
        if not alarm_ids:
            return {"status": "error", "message": "No alarm IDs provided"}
        if _bulk_action_state["status"] == "running":
            return {"status": "error", "message": "A bulk action is already running — wait for it to finish"}
        # Set state synchronously here, not inside the thread — otherwise the
        # frontend's first poll can race the thread startup and read stale
        # state left over from a previous run.
        _bulk_action_state.update(status="running", total=len(alarm_ids), completed=0,
                                   success=0, failed=0, errors=[])
        t = threading.Thread(
            target=_run_bulk_alarm_action_background,
            args=(action, alarm_ids, body.get("alarms") or [], _current_actor(request)),
            daemon=True,
        )
        t.start()
        return {"status": "running", "total": len(alarm_ids)}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.get("/api/alerts/bulk-action/status")
@app.get("/infraportal/api/alerts/bulk-action/status")
async def api_bulk_alarm_action_status():
    return {"status": _bulk_action_state["status"], **{k: v for k, v in _bulk_action_state.items() if k != "status"}}


# ── Alert hygiene reports (NOC cleanup) ───────────────────────────────────────

@app.get("/api/alerts/reports/{report_type}")
@app.get("/infraportal/api/alerts/reports/{report_type}")
async def api_alert_reports(report_type: str, days: int = 30):
    try:
        from routers.alert_reports import (
            build_stale_alerts_report, build_critical_vs_device_up_report,
            build_noisy_alert_types_report, build_chronic_red_devices_report,
            build_recurring_alerts_report,
        )
        if report_type == "stale":
            data = build_stale_alerts_report(min_age_days=days)
        elif report_type == "critical-vs-device":
            data = build_critical_vs_device_up_report()
        elif report_type == "noisy-types":
            data = build_noisy_alert_types_report(days=days)
        elif report_type == "chronic-red":
            data = build_chronic_red_devices_report(days=days)
        elif report_type == "recurring":
            data = build_recurring_alerts_report(days=days)
        else:
            return {"status": "error", "message": f"Unknown report type: {report_type}"}
        return {"status": "ok", "report": report_type, "count": len(data), "data": data}
    except Exception as e:
        return {"status": "error", "message": str(e)}


# ── Recurring Alerts AI analysis ──────────────────────────────────────────────

_recurring_ai_state: dict = {"status": "idle", "result": None, "error": None}

def _run_recurring_ai_analysis_background():
    global _recurring_ai_state
    try:
        from routers.cache import cache
        from routers.analysis import _call_claude
        from routers.alert_reports import build_recurring_alerts_report

        groups = build_recurring_alerts_report(days=90, min_occurrences=3)
        if not groups:
            result = {"summary": "No alert has been acknowledged/cleared 3+ times in the last 90 days.", "groups": [], "total_groups": 0}
            cache.set("recurring_alerts_ai_analysis", result)
            _recurring_ai_state["result"] = result
            _recurring_ai_state["status"] = "done"
            return

        # Cap the prompt regardless of how many groups exist — worst offenders
        # (highest action_count) matter most and keep token cost/latency bounded.
        top_groups = groups[:40]
        lines = []
        for i, g in enumerate(top_groups):
            evidence = []
            if "orphaned_device" in g["tags"]:
                evidence.append("device no longer appears in OpManager's device list")
            if "likely_decommissioned" in g["tags"]:
                evidence.append("Lansweeper hasn't seen this host recently or marked it decommissioned")
            if "flapping_threshold" in g["tags"]:
                evidence.append("fires as many separate short-lived alarm instances rather than one long outage")
            if g.get("device_status"):
                evidence.append(f"device's own OpManager status is currently {g['device_status']}")
            sample = (g.get("sample_messages") or [""])[0][:200]
            lines.append(
                f"[{i}] Device: {g['device_name']} | Event: {g['event_type']} | Category: {g.get('category') or 'Unknown'} | "
                f"Actioned {g['action_count']}x over {g['span_days']}d (acked {g.get('acknowledge_count', 0)}x, cleared {g.get('clear_count', 0)}x) | "
                f"Evidence: {'; '.join(evidence) or 'none found'} | Sample message: {sample}"
            )

        prompt = f"""You are a NOC alert-hygiene assistant for Zinnia's infrastructure monitoring (ManageEngine OpManager).

Below are alert groups (a specific device + specific check) that the NOC has had to acknowledge or clear
repeatedly over the last 90 days. A healthy alert fires once, gets fixed, and stays clear — repeated action on
the exact same check usually means one of: the monitored resource is decommissioned/offline and the alarm was
never removed from OpManager, the threshold is set too tight for normal variance (flapping), or it's a real
recurring issue that needs an actual fix rather than repeated dismissal.

GROUPS:
{chr(10).join(lines)}

For EACH group above (referenced by its [n] index), give your best-guess root cause and a concrete recommendation.
Respond with ONLY a valid JSON object, no prose, no markdown fences:
{{
  "summary": "2-3 sentence overview of what's driving repeat alert work for the NOC",
  "findings": [
    {{"index": 0,
      "likely_cause": "decommissioned_device|threshold_misconfigured|flapping_flaky_check|real_recurring_issue|unclear",
      "explanation": "1-2 sentences justifying the cause given the evidence",
      "recommendation": "specific action — e.g. remove this monitor, raise threshold to X, escalate to the owning team",
      "confidence": "high|medium|low"}}
  ]
}}"""

        response = _call_claude(model="claude-sonnet-4-6", max_tokens=4000, messages=[{"role": "user", "content": prompt}])

        import json
        text = response.strip()
        if text.startswith("```"):
            text = re.sub(r"^```[a-z]*\n?", "", text)
            text = re.sub(r"\n?```$", "", text)
        parsed = json.loads(text)

        findings_by_index = {f.get("index"): f for f in parsed.get("findings", [])}
        merged = []
        for i, g in enumerate(top_groups):
            f = findings_by_index.get(i, {})
            merged.append({
                **g,
                "ai_likely_cause":   f.get("likely_cause"),
                "ai_explanation":    f.get("explanation"),
                "ai_recommendation": f.get("recommendation"),
                "ai_confidence":     f.get("confidence"),
            })

        result = {"summary": parsed.get("summary", ""), "groups": merged, "total_groups": len(groups)}
        cache.set("recurring_alerts_ai_analysis", result)
        _recurring_ai_state["result"] = result
        _recurring_ai_state["status"] = "done"

    except Exception as e:
        _recurring_ai_state["status"] = "error"
        _recurring_ai_state["error"] = str(e)


@app.post("/api/alerts/reports/recurring-ai-analysis")
@app.post("/infraportal/api/alerts/reports/recurring-ai-analysis")
async def api_recurring_ai_analysis(request: Request):
    global _recurring_ai_state
    from routers.cache import cache

    body = {}
    try:    body = await request.json()
    except: pass
    refresh = body.get("refresh", False)

    if not refresh:
        cached, ts = cache.get("recurring_alerts_ai_analysis")
        if cached:
            return {"status": "ok", "result": cached, "timestamp": cache.age_string(ts), "cached": True}

    if _recurring_ai_state["status"] != "running":
        _recurring_ai_state["status"] = "running"
        _recurring_ai_state["result"] = None
        _recurring_ai_state["error"]  = None
        t = threading.Thread(target=_run_recurring_ai_analysis_background, daemon=True)
        t.start()

    return {"status": "running"}


@app.get("/api/alerts/reports/recurring-ai-analysis/status")
@app.get("/infraportal/api/alerts/reports/recurring-ai-analysis/status")
async def api_recurring_ai_analysis_status():
    global _recurring_ai_state
    if _recurring_ai_state["status"] == "running":
        return {"status": "running"}
    if _recurring_ai_state["status"] == "error":
        return {"status": "error", "message": _recurring_ai_state["error"]}
    if _recurring_ai_state["status"] == "done" and _recurring_ai_state["result"]:
        r = _recurring_ai_state["result"]
        _recurring_ai_state["status"] = "idle"
        from routers.cache import cache
        _, ts = cache.get("recurring_alerts_ai_analysis")
        return {"status": "ok", "result": r, "timestamp": cache.age_string(ts) if ts else "", "cached": False}
    from routers.cache import cache
    cached, ts = cache.get("recurring_alerts_ai_analysis")
    if cached:
        return {"status": "ok", "result": cached, "timestamp": cache.age_string(ts), "cached": True}
    return {"status": "idle"}


# ── Alerts AI Lens ────────────────────────────────────────────────────────────

_alerts_analysis_state: dict = {"status": "idle", "result": None, "error": None}

def _run_alerts_analysis_background():
    """Run focused alert AI analysis in a background thread."""
    global _alerts_analysis_state
    try:
        from routers.cache import cache
        from routers.analysis import _call_claude
        from routers.opmanager import get_alarms

        alarms, _ = get_alarms(force_refresh=False)
        active = [a for a in alarms if a.get("severity") != "Clear"]

        if not active:
            _alerts_analysis_state["result"] = {"summary": "No active alerts.", "top_devices": [], "patterns": [], "tier_breakdown": {}, "recommendations": [], "grade": "A", "grade_note": "No active alerts."}
            _alerts_analysis_state["status"] = "done"
            return

        # Build compact alert payload for Claude
        sev_counts = {}
        device_scores = {}
        evt_counts = {}
        cat_counts = {}
        sev_weights = {"Critical": 3, "Trouble": 2, "Attention": 1, "Service Down": 2}

        for a in active:
            sev  = a.get("severity", "Unknown")
            dev  = (a.get("device_name") or a.get("ip_address") or "Unknown").split(".")[0].upper()
            evt  = a.get("event_type") or "Unknown"
            cat  = a.get("category") or "Unknown"
            sev_counts[sev]   = sev_counts.get(sev, 0) + 1
            device_scores[dev] = device_scores.get(dev, 0) + sev_weights.get(sev, 1)
            evt_counts[evt]    = evt_counts.get(evt, 0) + 1
            cat_counts[cat]    = cat_counts.get(cat, 0) + 1

        top_devices_raw = sorted(device_scores.items(), key=lambda x: x[1], reverse=True)[:20]

        # Tier breakdown from naming convention
        tier_counts = {"Production (P)": 0, "QA (Q)": 0, "Dev/UAT/Test": 0, "Unknown": 0}
        for a in active:
            dev = (a.get("device_name") or "").upper()
            if "TOP" in dev or dev.endswith("P") or "P0" in dev or "P1" in dev or "P2" in dev:
                tier_counts["Production (P)"] += 1
            elif dev.endswith("Q") or "Q0" in dev or "Q1" in dev:
                tier_counts["QA (Q)"] += 1
            elif any(x in dev for x in ["DEV","UAT","TEST","XDMI","CTXMI"]):
                tier_counts["Dev/UAT/Test"] += 1
            else:
                tier_counts["Unknown"] += 1

        prompt = f"""You are analyzing OpManager alerts for Zinnia's Windows/VMware infrastructure.

ACTIVE ALERTS: {len(active)} across {len(device_scores)} devices

SEVERITY BREAKDOWN:
{chr(10).join(f"  {k}: {v}" for k, v in sorted(sev_counts.items(), key=lambda x: sev_weights.get(x[0], 0), reverse=True))}

TOP DEVICES BY SEVERITY SCORE (Critical=3, Trouble=2, Attention=1):
{chr(10).join(f"  {d}: score {s}" for d, s in top_devices_raw)}

TOP EVENT TYPES:
{chr(10).join(f"  {k}: {v}" for k, v in sorted(evt_counts.items(), key=lambda x: x[1], reverse=True)[:15])}

TOP CATEGORIES:
{chr(10).join(f"  {k}: {v}" for k, v in sorted(cat_counts.items(), key=lambda x: x[1], reverse=True)[:10])}

SERVER NAMING: TOP=Topeka, P=Production, Q=QA, D/U/T=Dev/UAT/Test. CTX=Citrix, DB=Database, APP=Application.

Respond with ONLY a valid JSON object, no prose, no markdown fences:
{{
  "summary": "2-3 sentence executive summary of current alert health",
  "grade": "A|B|C|D|F",
  "grade_note": "one line justification",
  "top_devices": [{{"device": "NAME", "score": N, "primary_issue": "what is firing"}}],
  "patterns": [{{"title": "pattern name", "detail": "explanation", "severity": "critical|warning|info"}}],
  "tier_breakdown": {{"Production": N, "QA": N, "Dev/UAT/Test": N}},
  "recommendations": [{{"priority": "high|medium|low", "action": "specific action", "devices": "device names if applicable"}}]
}}"""

        response = _call_claude(
            model      = "claude-sonnet-4-6",
            max_tokens = 1500,
            messages   = [{"role": "user", "content": prompt}],
        )

        # Parse JSON from response
        import json
        text = response.strip()
        if text.startswith("```"):
            text = re.sub(r"^```[a-z]*\n?", "", text)
            text = re.sub(r"\n?```$", "", text)
        result = json.loads(text)

        # Attach chart-ready data
        result["chart_severity"]  = sev_counts
        result["chart_evt_types"] = dict(sorted(evt_counts.items(), key=lambda x: x[1], reverse=True)[:10])
        result["chart_top_devs"]  = dict(top_devices_raw[:10])
        result["chart_tier"]      = tier_counts
        result["total_active"]    = len(active)
        result["total_devices"]   = len(device_scores)

        from routers.cache import cache
        cache.set("alerts_ai_analysis", result)

        _alerts_analysis_state["result"] = result
        _alerts_analysis_state["status"] = "done"

    except Exception as e:
        _alerts_analysis_state["status"] = "error"
        _alerts_analysis_state["error"]  = str(e)


@app.post("/api/alerts/ai-analysis")
@app.post("/infraportal/api/alerts/ai-analysis")
async def api_alerts_ai_analysis(request: Request):
    """Trigger or serve the focused alert AI analysis."""
    global _alerts_analysis_state
    from routers.cache import cache

    body    = {}
    try:    body = await request.json()
    except: pass
    refresh = body.get("refresh", False)

    if not refresh:
        cached, ts = cache.get("alerts_ai_analysis")
        if cached:
            age = cache.age_string(ts)
            return {"status": "ok", "result": cached, "timestamp": age, "cached": True}

    if _alerts_analysis_state["status"] != "running":
        _alerts_analysis_state["status"] = "running"
        _alerts_analysis_state["result"] = None
        _alerts_analysis_state["error"]  = None
        t = threading.Thread(target=_run_alerts_analysis_background, daemon=True)
        t.start()

    return {"status": "running"}


@app.get("/api/alerts/ai-analysis/status")
@app.get("/infraportal/api/alerts/ai-analysis/status")
async def api_alerts_ai_analysis_status():
    """Poll for alerts AI analysis completion."""
    global _alerts_analysis_state
    if _alerts_analysis_state["status"] == "running":
        return {"status": "running"}
    if _alerts_analysis_state["status"] == "error":
        return {"status": "error", "message": _alerts_analysis_state["error"]}
    if _alerts_analysis_state["status"] == "done" and _alerts_analysis_state["result"]:
        r = _alerts_analysis_state["result"]
        _alerts_analysis_state["status"] = "idle"
        from routers.cache import cache
        _, ts = cache.get("alerts_ai_analysis")
        age = cache.age_string(ts) if ts else ""
        return {"status": "ok", "result": r, "timestamp": age, "cached": False}
    from routers.cache import cache
    cached, ts = cache.get("alerts_ai_analysis")
    if cached:
        age = cache.age_string(ts)
        return {"status": "ok", "result": cached, "timestamp": age, "cached": True}
    return {"status": "idle"}


# ── Chat API ──────────────────────────────────────────────────────────────────

@app.post("/api/chat")
@app.post("/infraportal/api/chat")
async def api_chat(request: Request):
    try:
        from routers.analysis import _call_claude
        body     = await request.json()
        messages = body.get("messages", [])
        if not messages:
            return {"status": "error", "message": "No messages provided"}
        # Bug #4 fix: use shared _call_claude() retry helper (retries on 500/529)
        # instead of a bare client.messages.create() that silently fails.
        response = _call_claude(
            model      = "claude-sonnet-4-6",
            max_tokens = 1500,
            messages   = messages,
        )
        return {"status": "ok", "response": response}
    except Exception as e:
        return {"status": "error", "message": str(e)}

# ── VMware AI Lens ────────────────────────────────────────────────────────────

_vmware_analysis_state: dict = {"status": "idle", "result": None, "error": None}

def _run_vmware_analysis_background():
    global _vmware_analysis_state
    try:
        from routers.cache import cache
        from routers.analysis import _call_claude
        vms_data,     _ = cache.get("detailed_vms")
        storage_data, _ = cache.get("vm_storage")
        snapshot_data,_ = cache.get("vm_snapshots")
        ds_data,      _ = cache.get("vm_datastores")
        vms       = vms_data["vms"]          if vms_data      else []
        storage   = storage_data["vms"]      if storage_data  else []
        snapshots = snapshot_data["snapshots"] if snapshot_data else []
        datastores= ds_data["datastores"]    if ds_data       else []
        if not vms:
            _vmware_analysis_state["result"] = {"summary": "No VM data cached.", "grade": "?", "grade_note": "No data", "patterns": [], "recommendations": []}
            _vmware_analysis_state["status"] = "done"
            return
        total_vms  = len(vms)
        powered_on = sum(1 for v in vms if v.get("power_state") == "POWERED_ON")
        eol_vms    = [v for v in vms if v.get("eol")]
        tools_bad  = [v for v in vms if v.get("tools_upgrade_needed")]
        no_ip      = [v for v in vms if v.get("power_state") == "POWERED_ON" and not v.get("ip_address")]
        envs = {}
        for v in vms:
            e = v.get("environment","Unknown"); envs[e] = envs.get(e,0)+1
        low_disk  = [v for v in storage if v.get("has_low_disk")]
        crit_disk = []
        for v in storage:
            for d in v.get("drives",[]):
                if d.get("pct_free",100) <= 10:
                    crit_disk.append(v["name"]+" "+d.get("letter","")+" "+str(d.get("pct_free",""))+"%")
        ds_crit = [d for d in datastores if d.get("pct_free") is not None and d.get("pct_free") <= 10]
        ds_warn = [d for d in datastores if d.get("pct_free") is not None and 10 < d.get("pct_free") <= 20]
        snap_old7 = [s for s in snapshots if (s.get("age_days") or 0) >= 7]
        snap_prod = [s for s in snapshots if s.get("tier") == "Production"]
        snap_vms  = len(set(s.get("vm_name") for s in snapshots))
        tier_counts = {"Production": 0, "QA": 0, "Dev/UAT/Test": 0}
        for v in vms:
            n = v.get("name","").upper()
            if n.endswith("P") or "P0" in n or "P1" in n: tier_counts["Production"] += 1
            elif n.endswith("Q") or "Q0" in n: tier_counts["QA"] += 1
            else: tier_counts["Dev/UAT/Test"] += 1
        eol_names   = ", ".join(v.get("name","")+" ("+( v.get("eol") or {}).get("eol","")+")" for v in eol_vms[:10])
        tools_names = ", ".join(v.get("name","") for v in tools_bad[:10])
        low_names   = ", ".join(v.get("name","") for v in low_disk[:10])
        ds_crit_names = ", ".join(d.get("name","")+" ("+str(d.get("pct_free",""))+"% free)" for d in ds_crit[:5])
        snap_old_names = ", ".join(set(s.get("vm_name","") for s in snap_old7[:10]))
        env_lines  = "\n".join("  "+k+": "+str(v) for k,v in envs.items())
        tier_lines = "\n".join("  "+k+": "+str(v) for k,v in tier_counts.items())
        prompt = (
            "You are analyzing VMware infrastructure for Zinnia.\n\n"
            "VM SUMMARY:\n  Total: "+str(total_vms)+" | Powered on: "+str(powered_on)+" | Powered off: "+str(total_vms-powered_on)+"\n\n"
            "BY ENVIRONMENT:\n"+env_lines+"\n\n"
            "BY TIER:\n"+tier_lines+"\n\n"
            "EOL OS ("+str(len(eol_vms))+" VMs): "+(eol_names or "None")+"\n"
            "VMWARE TOOLS OUTDATED ("+str(len(tools_bad))+" VMs): "+(tools_names or "None")+"\n"
            "POWERED ON WITH NO IP ("+str(len(no_ip))+"): "+", ".join(v.get("name","") for v in no_ip[:10]) or "None"+"\n\n"
            "DISK SPACE:\n  VMs with low disk (<15%): "+str(len(low_disk))+"\n"
            "  Critical drives (<10%): "+(", ".join(crit_disk[:8]) or "None")+"\n\n"
            "DATASTORES:\n  Total: "+str(len(datastores))+" | Critical (<10%): "+str(len(ds_crit))+" | Warning (<20%): "+str(len(ds_warn))+"\n"
            "  Critical datastores: "+(ds_crit_names or "None")+"\n\n"
            "SNAPSHOTS:\n  Total: "+str(len(snapshots))+" across "+str(snap_vms)+" VMs\n"
            "  Older than 7 days: "+str(len(snap_old7))+" | On prod VMs: "+str(len(snap_prod))+"\n"
            "  VMs with old snapshots: "+(snap_old_names or "None")+"\n\n"
            "SERVER NAMING: P=Production, Q=QA, D/U/T=Dev/UAT/Test.\n\n"
            "Respond with ONLY valid JSON, no prose:\n"
            '{"summary":"2-3 sentence summary","grade":"A|B|C|D|F","grade_note":"one line",'
            '"patterns":[{"title":"","detail":"","severity":"critical|warning|info"}],'
            '"recommendations":[{"priority":"high|medium|low","action":"","devices":""}],'
            '"chart_health":{"Healthy":N,"EOL OS":N,"Tools Outdated":N,"Low Disk":N,"No IP":N},'
            '"chart_tiers":{"Production":N,"QA":N,"Dev/UAT/Test":N},'
            '"chart_envs":{"Topeka":N,"VMC on AWS":N,"Candor India":N}}'
        )
        response = _call_claude(model="claude-sonnet-4-6", max_tokens=1500, messages=[{"role":"user","content":prompt}])
        import json as _json, re as _re
        text = response.strip()
        if text.startswith("```"):
            text = _re.sub(r"^```[a-z]*\n?","",text); text = _re.sub(r"\n?```$","",text)
        result = _json.loads(text)
        result.update({"total_vms":total_vms,"powered_on":powered_on,"eol_count":len(eol_vms),
            "tools_count":len(tools_bad),"low_disk_count":len(low_disk),
            "ds_crit_count":len(ds_crit),"snap_old_count":len(snap_old7),"snap_prod_count":len(snap_prod)})
        cache.set("vmware_ai_analysis", result)
        _vmware_analysis_state["result"] = result
        _vmware_analysis_state["status"] = "done"
    except Exception as e:
        _vmware_analysis_state["status"] = "error"
        _vmware_analysis_state["error"]  = str(e)


@app.post("/api/vmware/ai-analysis")
@app.post("/infraportal/api/vmware/ai-analysis")
async def api_vmware_ai_analysis(request: Request):
    global _vmware_analysis_state
    from routers.cache import cache
    body = {}
    try: body = await request.json()
    except: pass
    if not body.get("refresh"):
        cached, ts = cache.get("vmware_ai_analysis")
        if cached: return {"status":"ok","result":cached,"timestamp":cache.age_string(ts),"cached":True}
    if _vmware_analysis_state["status"] != "running":
        _vmware_analysis_state.update({"status":"running","result":None,"error":None})
        threading.Thread(target=_run_vmware_analysis_background, daemon=True).start()
    return {"status":"running"}


@app.get("/api/vmware/ai-analysis/status")
@app.get("/infraportal/api/vmware/ai-analysis/status")
async def api_vmware_ai_analysis_status():
    global _vmware_analysis_state
    from routers.cache import cache
    if _vmware_analysis_state["status"] == "running": return {"status":"running"}
    if _vmware_analysis_state["status"] == "error":   return {"status":"error","message":_vmware_analysis_state["error"]}
    if _vmware_analysis_state["status"] == "done" and _vmware_analysis_state["result"]:
        r = _vmware_analysis_state["result"]; _vmware_analysis_state["status"] = "idle"
        _, ts = cache.get("vmware_ai_analysis")
        return {"status":"ok","result":r,"timestamp":cache.age_string(ts) if ts else "","cached":False}
    cached, ts = cache.get("vmware_ai_analysis")
    if cached: return {"status":"ok","result":cached,"timestamp":cache.age_string(ts),"cached":True}
    return {"status":"idle"}


# ── Citrix AI Lens ────────────────────────────────────────────────────────────

_citrix_analysis_state: dict = {"status": "idle", "result": None, "error": None}

def _run_citrix_analysis_background():
    global _citrix_analysis_state
    try:
        from routers.cache import cache
        from routers.analysis import _call_claude
        summary_data, _  = cache.get("citrix_summary")
        machines_data, _ = cache.get("citrix_machines")
        sessions_data, _ = cache.get("citrix_sessions")
        catalogs_data, _ = cache.get("citrix_catalogs")
        dgs_data, _      = cache.get("citrix_delivery_groups")
        d        = summary_data      if summary_data      else {}
        machines = machines_data     if machines_data     else []
        sessions = sessions_data     if sessions_data     else []
        catalogs = catalogs_data     if catalogs_data     else []
        dgs      = dgs_data          if dgs_data          else []
        if not d and not machines:
            _citrix_analysis_state["result"] = {"summary":"No Citrix data cached.","grade":"?","grade_note":"No data","patterns":[],"recommendations":[]}
            _citrix_analysis_state["status"] = "done"
            return
        total_m   = d.get("total_machines",len(machines))
        reg       = d.get("registered",0)
        unreg     = d.get("unregistered",0)
        maint     = d.get("in_maintenance",0)
        faults    = d.get("with_faults",0)
        active_s  = d.get("active_sessions",0)
        discn_s   = d.get("disconnected_sessions",0)
        total_s   = d.get("total_sessions",0)
        img_stale = d.get("image_out_of_date",0)
        upgr      = d.get("upgrade_available",0)
        total_dgs = d.get("total_delivery_groups",len(dgs))
        total_cats= d.get("total_catalogs",len(catalogs))
        agent_ver = d.get("agent_versions",{})
        broken_cats = [c for c in catalogs if c.get("is_broken")]
        warn_cats   = [c for c in catalogs if (c.get("warnings") or c.get("errors")) and not c.get("is_broken")]
        unreg_machines = [m for m in machines if m.get("registration_state")=="Unregistered"]
        maint_machines = [m for m in machines if m.get("in_maintenance_mode")]
        zones = {}
        for m in machines:
            z = m.get("zone","Unknown"); zones[z] = zones.get(z,0)+1
        chart_machines = {"Registered":reg,"Unregistered":unreg,"In Maintenance":maint,"Faults":faults}
        chart_sessions = {"Active":active_s,"Disconnected":discn_s}
        chart_catalogs = {"OK":total_cats-len(broken_cats)-len(warn_cats),"Warnings":len(warn_cats),"Broken":len(broken_cats),"Upgrade Available":upgr}
        agent_top = ", ".join(k+"x"+str(v) for k,v in sorted(agent_ver.items(),key=lambda x:x[1],reverse=True)[:5])
        unreg_names = ", ".join(m.get("name","") for m in unreg_machines[:10])
        broken_names = ", ".join(c.get("name","") for c in broken_cats[:5])
        zone_lines = "\n".join("  "+k+": "+str(v) for k,v in sorted(zones.items(),key=lambda x:x[1],reverse=True)[:8])
        prompt = (
            "You are analyzing Citrix Cloud infrastructure for Zinnia.\n\n"
            "MACHINES:\n  Total: "+str(total_m)+" | Registered: "+str(reg)+" | Unregistered: "+str(unreg)+" | Maintenance: "+str(maint)+" | Faults: "+str(faults)+"\n"
            "  Unregistered machines: "+(unreg_names or "None")+"\n\n"
            "SESSIONS:\n  Active: "+str(active_s)+" | Disconnected: "+str(discn_s)+" | Total: "+str(total_s)+"\n\n"
            "DELIVERY GROUPS: "+str(total_dgs)+"\n"
            "MACHINE CATALOGS: "+str(total_cats)+" | Broken: "+str(len(broken_cats))+" | Warnings: "+str(len(warn_cats))+" | Upgrades: "+str(upgr)+"\n"
            "  Broken catalogs: "+(broken_names or "None")+"\n\n"
            "STALE MASTER IMAGES: "+str(img_stale)+" machines\n"
            "VDA AGENT VERSIONS: "+(agent_top or "Unknown")+"\n\n"
            "MACHINES BY ZONE:\n"+zone_lines+"\n\n"
            "Respond with ONLY valid JSON, no prose:\n"
            '{"summary":"2-3 sentence summary","grade":"A|B|C|D|F","grade_note":"one line",'
            '"patterns":[{"title":"","detail":"","severity":"critical|warning|info"}],'
            '"recommendations":[{"priority":"high|medium|low","action":"","devices":""}],'
            '"chart_machines":{"Registered":N,"Unregistered":N,"Maintenance":N,"Faults":N},'
            '"chart_sessions":{"Active":N,"Disconnected":N},'
            '"chart_catalogs":{"OK":N,"Warnings":N,"Broken":N,"Upgrade Available":N},'
            '"chart_zones":{"Zone1":N}}'
        )
        response = _call_claude(model="claude-sonnet-4-6", max_tokens=1500, messages=[{"role":"user","content":prompt}])
        import json as _json, re as _re
        text = response.strip()
        if text.startswith("```"):
            text = _re.sub(r"^```[a-z]*\n?","",text); text = _re.sub(r"\n?```$","",text)
        result = _json.loads(text)
        result["chart_zones"] = zones
        cache.set("citrix_ai_analysis", result)
        _citrix_analysis_state["result"] = result
        _citrix_analysis_state["status"] = "done"
    except Exception as e:
        _citrix_analysis_state["status"] = "error"
        _citrix_analysis_state["error"]  = str(e)


@app.post("/api/citrix/ai-analysis")
@app.post("/infraportal/api/citrix/ai-analysis")
async def api_citrix_ai_analysis(request: Request):
    global _citrix_analysis_state
    from routers.cache import cache
    body = {}
    try: body = await request.json()
    except: pass
    if not body.get("refresh"):
        cached, ts = cache.get("citrix_ai_analysis")
        if cached: return {"status":"ok","result":cached,"timestamp":cache.age_string(ts),"cached":True}
    if _citrix_analysis_state["status"] != "running":
        _citrix_analysis_state.update({"status":"running","result":None,"error":None})
        threading.Thread(target=_run_citrix_analysis_background, daemon=True).start()
    return {"status":"running"}


@app.get("/api/citrix/ai-analysis/status")
@app.get("/infraportal/api/citrix/ai-analysis/status")
async def api_citrix_ai_analysis_status():
    global _citrix_analysis_state
    from routers.cache import cache
    if _citrix_analysis_state["status"] == "running": return {"status":"running"}
    if _citrix_analysis_state["status"] == "error":   return {"status":"error","message":_citrix_analysis_state["error"]}
    if _citrix_analysis_state["status"] == "done" and _citrix_analysis_state["result"]:
        r = _citrix_analysis_state["result"]; _citrix_analysis_state["status"] = "idle"
        _, ts = cache.get("citrix_ai_analysis")
        return {"status":"ok","result":r,"timestamp":cache.age_string(ts) if ts else "","cached":False}
    cached, ts = cache.get("citrix_ai_analysis")
    if cached: return {"status":"ok","result":cached,"timestamp":cache.age_string(ts),"cached":True}
    return {"status":"idle"}


# ── Active Directory AI Lens ─────────────────────────────────────────────────

_ad_analysis_state: dict = {"status": "idle", "result": None, "error": None}

def _run_ad_analysis_background():
    global _ad_analysis_state
    try:
        from routers.cache import cache
        from routers.analysis import _call_claude
        from routers.active_directory import get_ad_analysis_for_ai
        import json as _json, re as _re

        d = get_ad_analysis_for_ai()
        if not d or d.get("error"):
            _ad_analysis_state["result"] = {"summary": "No AD data available.", "grade": "?", "grade_note": "No data", "patterns": [], "recommendations": []}
            _ad_analysis_state["status"] = "done"
            return

        s   = d.get("summary", {})
        gpo = d.get("gpo_health", {})
        sc  = d.get("stale_computers", {})
        pp  = d.get("password_policy_issues", [])
        priv = d.get("privileged_accounts", {})

        total_users    = s.get("total_users", 0)
        active_users   = s.get("active_users", 0)
        stale_users    = s.get("stale_users", 0)
        pwd_no_exp     = s.get("pwd_never_expires", 0)
        domain_admins  = s.get("domain_admin_count", 0)
        ent_admins     = s.get("enterprise_admin_count", 0)
        stale_comps    = s.get("stale_computers", 0)
        total_groups   = s.get("total_groups", 0)
        empty_groups   = s.get("empty_groups", 0)

        gpo_total      = gpo.get("total_gpos", 0)
        gpo_orphaned   = gpo.get("orphaned", 0)
        gpo_disabled   = gpo.get("disabled", 0)
        gpo_stale      = gpo.get("stale", 0)
        gpo_linked_dis = gpo.get("linked_disabled", 0)

        zombies        = len(sc.get("zombie_detail", []))
        citrix_stale   = len([c for c in sc.get("top_stale_regular", []) if c.get("citrix_managed")])

        pp_lines = "\n".join(str(p) for p in pp[:5]) if pp else "No issues detected"

        prompt = (
            "You are analyzing Active Directory health for the Zinnia environment.\n\n"
            "USER ACCOUNTS:\n"
            "  Total: " + str(total_users) + " | Active: " + str(active_users) +
            " | Stale (90d+): " + str(stale_users) + " | Pwd Never Expires: " + str(pwd_no_exp) + "\n\n"
            "PRIVILEGED ACCESS:\n"
            "  Domain Admins: " + str(domain_admins) + " | Enterprise Admins: " + str(ent_admins) + "\n"
            "  (>5 domain admins or >2 enterprise admins is a concern)\n\n"
            "GROUPS:\n  Total: " + str(total_groups) + " | Empty: " + str(empty_groups) + "\n\n"
            "STALE COMPUTERS:\n"
            "  Total stale (90d+): " + str(stale_comps) + " | Zombie machines: " + str(zombies) +
            " | Citrix-managed stale: " + str(citrix_stale) + "\n"
            "  (Zombies = stale AD + Citrix unregistered + still powered on — highest priority)\n\n"
            "GROUP POLICY:\n"
            "  Total GPOs: " + str(gpo_total) + " | Orphaned: " + str(gpo_orphaned) +
            " | Stale 2yr+: " + str(gpo_stale) + " | Linked+Disabled: " + str(gpo_linked_dis) +
            " | Disabled: " + str(gpo_disabled) + "\n\n"
            "PASSWORD POLICY ISSUES:\n" + pp_lines + "\n\n"
            "Respond with ONLY valid JSON, no prose:\n"
            '{"summary":"2-3 sentence summary of AD health",'
            '"grade":"A|B|C|D|F","grade_note":"one line reason for grade",'
            '"patterns":[{"title":"","detail":"","severity":"critical|warning|info"}],'
            '"recommendations":[{"priority":"high|medium|low","action":"","devices":""}],'
            '"chart_users":{"Active":' + str(active_users) + ',"Stale":' + str(stale_users) + ',"Pwd Never Expires":' + str(pwd_no_exp) + '},'
            '"chart_privileged":{"Domain Admins":' + str(domain_admins) + ',"Enterprise Admins":' + str(ent_admins) + '},'
            '"chart_gpo":{"Total":' + str(gpo_total) + ',"Orphaned":' + str(gpo_orphaned) + ',"Stale 2yr+":' + str(gpo_stale) + ',"Linked+Disabled":' + str(gpo_linked_dis) + '}}'
        )

        response = _call_claude(model="claude-sonnet-4-6", max_tokens=1500, messages=[{"role": "user", "content": prompt}])
        text = response.strip()
        if text.startswith("```"):
            text = _re.sub(r"^```[a-z]*\n?", "", text)
            text = _re.sub(r"\n?```$", "", text)
        result = _json.loads(text)
        result.update({
            "total_users": total_users, "active_users": active_users,
            "stale_users": stale_users, "pwd_no_exp": pwd_no_exp,
            "domain_admins": domain_admins, "ent_admins": ent_admins,
            "stale_comps": stale_comps, "zombies": zombies,
            "gpo_total": gpo_total, "gpo_orphaned": gpo_orphaned,
        })
        cache.set("ad_ai_analysis", result)
        _ad_analysis_state["result"] = result
        _ad_analysis_state["status"] = "done"
    except Exception as e:
        _ad_analysis_state["status"] = "error"
        _ad_analysis_state["error"]  = str(e)


@app.post("/api/ad/ai-analysis")
@app.post("/infraportal/api/ad/ai-analysis")
async def api_ad_ai_analysis(request: Request):
    global _ad_analysis_state
    from routers.cache import cache
    body = {}
    try: body = await request.json()
    except: pass
    if not body.get("refresh"):
        cached, ts = cache.get("ad_ai_analysis")
        if cached: return {"status": "ok", "result": cached, "timestamp": cache.age_string(ts), "cached": True}
    if _ad_analysis_state["status"] != "running":
        _ad_analysis_state.update({"status": "running", "result": None, "error": None})
        threading.Thread(target=_run_ad_analysis_background, daemon=True).start()
    return {"status": "running"}


@app.get("/api/ad/ai-analysis/status")
@app.get("/infraportal/api/ad/ai-analysis/status")
async def api_ad_ai_analysis_status():
    global _ad_analysis_state
    from routers.cache import cache
    if _ad_analysis_state["status"] == "running": return {"status": "running"}
    if _ad_analysis_state["status"] == "error":   return {"status": "error", "message": _ad_analysis_state["error"]}
    if _ad_analysis_state["status"] == "done" and _ad_analysis_state["result"]:
        r = _ad_analysis_state["result"]; _ad_analysis_state["status"] = "idle"
        _, ts = cache.get("ad_ai_analysis")
        return {"status": "ok", "result": r, "timestamp": cache.age_string(ts) if ts else "", "cached": False}
    cached, ts = cache.get("ad_ai_analysis")
    if cached: return {"status": "ok", "result": cached, "timestamp": cache.age_string(ts), "cached": True}
    return {"status": "idle"}


# ── Assets AI Lens ────────────────────────────────────────────────────────────

_assets_analysis_state: dict = {"status": "idle", "result": None, "error": None}

def _run_assets_analysis_background():
    global _assets_analysis_state
    try:
        from routers.cache import cache
        from routers.analysis import _call_claude
        from routers.lansweeper import get_lansweeper_analysis_for_ai
        import json as _json, re as _re

        d = get_lansweeper_analysis_for_ai()
        if not d or d.get("error"):
            _assets_analysis_state["result"] = {"summary": "No asset data cached.", "grade": "?", "grade_note": "No data", "patterns": [], "recommendations": []}
            _assets_analysis_state["status"] = "done"
            return

        ac           = d.get("asset_counts", {})
        ph           = d.get("patch_health", {})
        total        = ac.get("total", 0) or 0
        eol_count    = ac.get("eol_count", 0) or 0
        not_seen_30  = ac.get("not_seen_30d", 0) or 0
        not_seen_90  = ac.get("not_seen_90d", 0) or 0
        active       = max(0, total - not_seen_30)

        patch_pct    = ph.get("patch_pct", 0) or 0
        patched      = ph.get("patched", 0) or 0
        unpatched    = ph.get("unpatched", 0) or 0
        unknown_p    = ph.get("unknown", 0) or 0

        type_breakdown = ac.get("by_type", {}) or {}
        os_breakdown   = d.get("os_breakdown", {}) or {}
        top_eol        = d.get("top_eol", [])[:8]
        top_unpatched  = d.get("top_unpatched", [])[:8]

        type_lines = "\n".join("  " + k + ": " + str(v) for k, v in list(type_breakdown.items())[:8]) or "  No data"
        os_lines   = "\n".join("  " + k + ": " + str(v) for k, v in list(os_breakdown.items())[:8])   or "  No data"
        eol_names  = ", ".join(str(a.get("name", "")) for a in top_eol)       or "None"
        unp_names  = ", ".join(str(a.get("name", "")) for a in top_unpatched) or "None"

        prompt = (
            "You are analyzing Lansweeper asset inventory health for the Zinnia environment.\n\n"
            "ASSET OVERVIEW:\n"
            "  Total assets: " + str(total) + " | Active (<30d): " + str(active) +
            " | Stale (30-90d): " + str(max(0, not_seen_30 - not_seen_90)) +
            " | Dead (90d+): " + str(not_seen_90) + "\n"
            "  EOL OS count: " + str(eol_count) + "\n\n"
            "PATCH COMPLIANCE:\n"
            "  Patch %: " + str(patch_pct) + "% | Patched: " + str(patched) +
            " | Unpatched: " + str(unpatched) + " | Unknown: " + str(unknown_p) + "\n"
            "  Top unpatched: " + unp_names + "\n\n"
            "BY TYPE:\n" + type_lines + "\n\n"
            "BY OS (top 8):\n" + os_lines + "\n\n"
            "EOL ASSETS (top 8): " + eol_names + "\n\n"
            "GRADING NOTES: Below 70% patch compliance = D or lower. "
            "High EOL count on production-named assets = serious risk. "
            "Many dead assets (90d+) = inventory hygiene issue.\n\n"
            "Respond with ONLY valid JSON, no prose:\n"
            '{"summary":"2-3 sentence summary of asset health",'
            '"grade":"A|B|C|D|F","grade_note":"one line reason for grade",'
            '"patterns":[{"title":"","detail":"","severity":"critical|warning|info"}],'
            '"recommendations":[{"priority":"high|medium|low","action":"","devices":""}],'
            '"chart_visibility":{"Active <30d":' + str(active) + ',"Stale 30-90d":' + str(max(0, not_seen_30 - not_seen_90)) + ',"Dead 90d+":' + str(not_seen_90) + '},'
            '"chart_patch":{"Patched":' + str(patched) + ',"Unpatched":' + str(unpatched) + ',"Unknown":' + str(unknown_p) + '},'
            '"chart_eol":{"Supported":' + str(max(0, total - eol_count)) + ',"EOL OS":' + str(eol_count) + '}}'
        )

        response = _call_claude(model="claude-sonnet-4-6", max_tokens=1500, messages=[{"role": "user", "content": prompt}])
        text = response.strip()
        if text.startswith("```"):
            text = _re.sub(r"^```[a-z]*\n?", "", text)
            text = _re.sub(r"\n?```$", "", text)
        result = _json.loads(text)
        result.update({
            "total": total, "eol_count": eol_count,
            "not_seen_30": not_seen_30, "not_seen_90": not_seen_90,
            "patch_pct": patch_pct,
        })
        cache.set("assets_ai_analysis", result)
        _assets_analysis_state["result"] = result
        _assets_analysis_state["status"] = "done"
    except Exception as e:
        _assets_analysis_state["status"] = "error"
        _assets_analysis_state["error"]  = str(e)


@app.post("/api/assets/ai-analysis")
@app.post("/infraportal/api/assets/ai-analysis")
async def api_assets_ai_analysis(request: Request):
    global _assets_analysis_state
    from routers.cache import cache
    body = {}
    try: body = await request.json()
    except: pass
    if not body.get("refresh"):
        cached, ts = cache.get("assets_ai_analysis")
        if cached: return {"status": "ok", "result": cached, "timestamp": cache.age_string(ts), "cached": True}
    if _assets_analysis_state["status"] != "running":
        _assets_analysis_state.update({"status": "running", "result": None, "error": None})
        threading.Thread(target=_run_assets_analysis_background, daemon=True).start()
    return {"status": "running"}


@app.get("/api/assets/ai-analysis/status")
@app.get("/infraportal/api/assets/ai-analysis/status")
async def api_assets_ai_analysis_status():
    global _assets_analysis_state
    from routers.cache import cache
    if _assets_analysis_state["status"] == "running": return {"status": "running"}
    if _assets_analysis_state["status"] == "error":   return {"status": "error", "message": _assets_analysis_state["error"]}
    if _assets_analysis_state["status"] == "done" and _assets_analysis_state["result"]:
        r = _assets_analysis_state["result"]; _assets_analysis_state["status"] = "idle"
        _, ts = cache.get("assets_ai_analysis")
        return {"status": "ok", "result": r, "timestamp": cache.age_string(ts) if ts else "", "cached": False}
    cached, ts = cache.get("assets_ai_analysis")
    if cached: return {"status": "ok", "result": cached, "timestamp": cache.age_string(ts), "cached": True}
    return {"status": "idle"}


# ── Certs AI Lens ─────────────────────────────────────────────────────────────

_certs_analysis_state: dict = {"status": "idle", "result": None, "error": None}

_CERT_ORG_CONTEXT = """
KNOWN CRITICAL SERVERS — certs on these servers ALWAYS require manual renewal:
  Domain Controllers : CANDC01P, CANDC02P, SETOPDC01P, SETOPDC02P,
                       AWSEASTDC1, AWSEASTDC2, AWSWESTDC1
  Certificate Authority : SETOPCA02P
  ADFS                  : SETOPADFS01P, SETOPADFS02P
  Citrix FAS (Topeka)   : TOPCTXCCFAS01P, TOPCTXCCFAS02P
  Citrix FAS (VMC)      : VMCECTXFAS01P, VMCECTXFAS02P
  Citrix FAS (AWS)      : AWSCTXFAS01P, AWSCTXFAS02P

NOISE — classify as noise and exclude from watch_list / cleanup:
  - Template contains: Machine, Computer, Workstation Authentication, DomainController
  - Subject or requester contains: "ADFS Agent", "PolicyKeyService"
  - Expired more than 1 year ago AND not on a critical server
  - Auto-enrolled user/machine certs renewed regularly by PKI

MANUAL RENEWAL INDICATORS — these go on the watch_list:
  - KDC Authentication / Kerberos Authentication on any DC
  - Token Signing / Token Decryption on ADFS servers
  - SSL/TLS binding on Citrix FAS (FAS requires manual cert binding to the FAS service)
  - The CA cert itself (SETOPCA02P)
  - Any web server or service cert on a named critical server
  - Self-signed certs on known servers

CLEANUP INDICATORS — expired and likely safe to remove:
  - Expired certs on critical servers that have been superseded (newer cert for same CN exists)
  - Expired certs where days_remaining < -30 and classification is not ManualLikely
  - Old ADFS Agent certs (auto-replaced by ADFS, the expired ones are orphans)
"""

def _run_certs_analysis_background():
    global _certs_analysis_state
    try:
        from routers.cache import cache
        from routers.analysis import _call_claude
        from routers.ca_analysis import get_cert_watchlist_for_ai
        import json as _json, re as _re

        candidates, dc_certs, summary = get_cert_watchlist_for_ai()

        if not candidates and not dc_certs:
            _certs_analysis_state["result"] = {
                "watch_list": [], "cleanup": [], "noise_count": 0,
                "summary": "No certificate data available from CA.",
                "error": "No data"
            }
            _certs_analysis_state["status"] = "done"
            return

        # Build cert lines for prompt
        cert_lines = []
        for c in candidates:
            days = c.get("days_remaining")
            days_str = str(days) if days is not None else "unknown"
            cert_lines.append(
                f"  [{c.get('classification','?')}] server={c.get('critical_server') or 'unknown'}"
                f" role={c.get('server_role') or '—'}"
                f" subject={c.get('subject','?')}"
                f" template={c.get('template') or '(none)'}"
                f" expiry={c.get('expiry','?')} days={days_str}"
                f" requester={c.get('requester','?')}"
            )

        dc_lines = []
        for d in dc_certs:
            dc_lines.append(
                f"  dc={d.get('dc_name')} reachable={d.get('reachable')}"
                f" subject={d.get('subject','—')} expiry={d.get('expiry','—')}"
                f" days={d.get('days_remaining','?')} eku={d.get('eku','—')}"
                f" error={d.get('error') or 'none'}"
            )

        total_issued  = (summary or {}).get("total_issued", "?")
        auto_enroll   = (summary or {}).get("auto_enroll_likely", "?")
        machine_certs = (summary or {}).get("machine_certs", 0) or 0
        user_certs    = (summary or {}).get("user_certs", 0) or 0
        ldap_noise    = machine_certs + user_certs

        prompt = (
            "You are a Windows PKI expert reviewing the Zinnia org certificate inventory.\n\n"
            + _CERT_ORG_CONTEXT +
            f"\nCA STATS: {total_issued} total issued, ~{auto_enroll} auto-enrolled (noise floor).\n\n"
            "LDAP-SOURCED CERTS (ALL NOISE — do not include in watch_list or cleanup):\n"
            f"  Machine certs from AD computer objects: {machine_certs}\n"
            f"  User certs from AD user objects:        {user_certs}\n"
            f"  Total LDAP noise:                       {ldap_noise}\n"
            "These are auto-enrolled via PKI autoenrollment, renewed automatically, "
            "require no manual intervention, and should ALL be counted in noise_count.\n\n"
            "CA-ISSUED CANDIDATES (pre-filtered — AutoEnrollLikely with no critical server already excluded):\n"
            + "\n".join(cert_lines or ["  (none)"]) +
            "\n\nDC KERBEROS CERT SCAN RESULTS:\n"
            + "\n".join(dc_lines or ["  (none)"]) +
            "\n\nUsing the org context and noise rules above, classify each cert into exactly one bucket:\n"
            "- watch_list : Needs manual renewal, admin must act before expiry\n"
            "- cleanup    : Expired/orphaned, safe to delete, reduces attack surface\n"
            "- noise      : Auto-managed, ignore\n\n"
            f"Set noise_count to at least {ldap_noise} (the LDAP auto-enrolled certs above) "
            "plus any additional CA-issued certs you classify as noise.\n\n"
            "Respond with ONLY valid JSON — no prose, no markdown:\n"
            '{"summary":"2-3 sentence overall cert posture summary",'
            '"watch_list":[{"server":"","role":"","subject":"","cert_type":"e.g. DC Kerberos / ADFS Token Signing / Citrix FAS SSL / CA Cert",'
            '"expiry":"YYYY-MM-DD","days_remaining":0,"urgency":"ok|notice|warning|critical|expired",'
            '"renewal_notes":"one line on how/when to renew"}],'
            '"cleanup":[{"server":"","subject":"","expiry":"","days_remaining":0,'
            '"reason":"why safe to delete","action":"e.g. Delete from LocalMachine\\\\My on SERVER"}],'
            '"noise_count":0}'
        )

        response = _call_claude(
            model="claude-sonnet-4-6", max_tokens=3000,
            messages=[{"role": "user", "content": prompt}]
        )
        text = response.strip()
        if text.startswith("```"):
            text = _re.sub(r"^```[a-z]*\n?", "", text)
            text = _re.sub(r"\n?```$", "", text)
        result = _json.loads(text)

        # Sort watch list by urgency
        urgency_order = {"expired": 0, "critical": 1, "warning": 2, "notice": 3, "ok": 4}
        result["watch_list"] = sorted(
            result.get("watch_list", []),
            key=lambda x: (urgency_order.get(x.get("urgency", "ok"), 5),
                           x.get("days_remaining") if x.get("days_remaining") is not None else 9999)
        )

        cache.set("certs_ai_analysis", result)
        _certs_analysis_state["result"] = result
        _certs_analysis_state["status"] = "done"

    except Exception as e:
        _certs_analysis_state["status"] = "error"
        _certs_analysis_state["error"]  = str(e)


@app.post("/api/certs/ai-analysis")
@app.post("/infraportal/api/certs/ai-analysis")
async def api_certs_ai_analysis(request: Request):
    global _certs_analysis_state
    from routers.cache import cache
    body = {}
    try: body = await request.json()
    except: pass
    if not body.get("refresh"):
        cached, ts = cache.get("certs_ai_analysis")
        if cached: return {"status": "ok", "result": cached, "timestamp": cache.age_string(ts), "cached": True}
    if _certs_analysis_state["status"] != "running":
        _certs_analysis_state.update({"status": "running", "result": None, "error": None})
        threading.Thread(target=_run_certs_analysis_background, daemon=True).start()
    return {"status": "running"}


@app.get("/api/certs/ai-analysis/status")
@app.get("/infraportal/api/certs/ai-analysis/status")
async def api_certs_ai_analysis_status():
    global _certs_analysis_state
    from routers.cache import cache
    if _certs_analysis_state["status"] == "running": return {"status": "running"}
    if _certs_analysis_state["status"] == "error":   return {"status": "error", "message": _certs_analysis_state["error"]}
    if _certs_analysis_state["status"] == "done" and _certs_analysis_state["result"]:
        r = _certs_analysis_state["result"]; _certs_analysis_state["status"] = "idle"
        _, ts = cache.get("certs_ai_analysis")
        return {"status": "ok", "result": r, "timestamp": cache.age_string(ts) if ts else "", "cached": False}
    cached, ts = cache.get("certs_ai_analysis")
    if cached: return {"status": "ok", "result": cached, "timestamp": cache.age_string(ts), "cached": True}
    return {"status": "idle"}


# ── Jira API ──────────────────────────────────────────────────────────────────

@app.get("/infraportal/api/jira/search")
def api_jira_search(device: str, force_live: bool = False):
    """Search open ITSD tickets matching a device name.
    force_live=true bypasses all caches and goes direct to Jira API.
    """
    tickets = jira.search_tickets_by_device(device, force_live=force_live)
    return {"device": device, "count": len(tickets), "tickets": tickets}

@app.get("/infraportal/api/jira/teams/search")
async def api_jira_teams_search(q: str = ""):
    """Search Atlassian teams by name. Used by the create ticket modal."""
    try:
        teams = jira.search_teams(q)
        return {"status": "ok", "teams": teams}
    except Exception as e:
        return {"status": "error", "message": str(e)}

# ── Jira team config (DB-stored team list) ────────────────────────────────────

@app.get("/infraportal/api/jira/teams")
async def api_jira_teams_list():
    """Return all DB-stored Jira teams."""
    try:
        from routers.database import list_jira_teams
        return {"status": "ok", "teams": list_jira_teams()}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.post("/infraportal/api/jira/teams")
async def api_jira_teams_save(request: Request):
    """Add or update a team. Body: { id, name, sort_order? }"""
    try:
        from routers.database import save_jira_team
        body = await request.json()
        team_id = body.get("id", "").strip()
        name    = body.get("name", "").strip()
        order   = int(body.get("sort_order", 0))
        if not team_id or not name:
            return {"status": "error", "message": "id and name are required"}
        ok = save_jira_team(team_id, name, order)
        return {"status": "ok" if ok else "error"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.delete("/infraportal/api/jira/teams/{team_id:path}")
async def api_jira_teams_delete(team_id: str):
    """Delete a team by ID."""
    try:
        from routers.database import delete_jira_team
        ok = delete_jira_team(team_id)
        return {"status": "ok" if ok else "error"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.post("/infraportal/api/jira/teams/reorder")
async def api_jira_teams_reorder(request: Request):
    """Update team display order. Body: { ids: [...] }"""
    try:
        from routers.database import reorder_jira_teams
        body = await request.json()
        ok = reorder_jira_teams(body.get("ids", []))
        return {"status": "ok" if ok else "error"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.post("/infraportal/api/jira/create")
async def api_jira_create(request: Request):
    """Create an ITSD ticket. Body: { summary, description, priority, device_name, alarm_id, team_id }"""
    try:
        body   = await request.json()
        result = jira.create_ticket(
            summary     = body.get("summary", ""),
            description = body.get("description", ""),
            priority    = body.get("priority", "Medium"),
            device_name = body.get("device_name", ""),
            alarm_id    = body.get("alarm_id", ""),
            team_id     = body.get("team_id", ""),
        )
        if result:
            return {"status": "ok", "key": result["key"], "url": result["url"]}
        return {"status": "error", "message": "Ticket creation failed — check logs"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.get("/infraportal/api/jira/diagnose")
async def api_jira_diagnose():
    """
    Investigate customfield_11600 (Team) to find allowed values and IDs.
    Hit once: /infraportal/api/jira/diagnose
    """
    import requests
    from requests.auth import HTTPBasicAuth
    from routers.jira import VERIFY_SSL, _post, _get, _cfg
    cfg = _cfg()
    auth = HTTPBasicAuth(os.getenv("JIRA_EMAIL",""), os.getenv("JIRA_API_TOKEN",""))
    headers = {"Accept": "application/json", "Content-Type": "application/json"}
    base = cfg["base_url"]
    result = {}

    # ── 1. Field contexts for customfield_11600 ───────────────────────────────
    r1 = requests.get(
        f"{base}/rest/api/3/field/customfield_11600/context",
        auth=auth, headers=headers, timeout=10, verify=VERIFY_SSL
    )
    result["field_contexts"] = {
        "status": r1.status_code,
        "body": r1.json() if r1.ok else r1.text[:500],
    }

    # ── 2. Options/allowed values for each context ────────────────────────────
    context_options = []
    if r1.ok:
        for ctx in r1.json().get("values", []):
            ctx_id = ctx.get("id")
            r2 = requests.get(
                f"{base}/rest/api/3/field/customfield_11600/context/{ctx_id}/option",
                auth=auth, headers=headers, timeout=10, verify=VERIFY_SSL,
                params={"maxResults": 100}
            )
            context_options.append({
                "context_id":   ctx_id,
                "context_name": ctx.get("name"),
                "status":       r2.status_code,
                "options":      r2.json() if r2.ok else r2.text[:500],
            })
    result["field_options"] = context_options

    # ── 3. createmeta for ITSD — check what customfield_11600 expects ─────────
    r3 = requests.get(
        f"{base}/rest/api/3/issue/createmeta/{cfg['helpdesk']}/issuetypes",
        auth=auth, headers=headers, timeout=10, verify=VERIFY_SSL
    )
    result["itsd_issuetypes"] = {
        "status": r3.status_code,
        "body": r3.json() if r3.ok else r3.text[:300],
    }

    # Get fields for Incident type specifically
    incident_id = None
    if r3.ok:
        for it in r3.json().get("issueTypes", []):
            if it.get("name") == "Incident":
                incident_id = it.get("id")
                break
    if incident_id:
        r4 = requests.get(
            f"{base}/rest/api/3/issue/createmeta/{cfg['helpdesk']}/issuetypes/{incident_id}",
            auth=auth, headers=headers, timeout=10, verify=VERIFY_SSL,
            params={"maxResults": 100}
        )
        fields = {}
        if r4.ok:
            raw = r4.json()
            # createmeta returns either {"fields": [...]} or a list directly
            field_list = raw.get("fields", raw) if isinstance(raw, dict) else raw
            if isinstance(field_list, dict):
                field_list = list(field_list.values())
            for f in field_list:
                if not isinstance(f, dict):
                    continue
                fname = f.get("name","")
                fkey  = f.get("key","") or f.get("fieldId","")
                if "11600" in fkey or "team" in fname.lower() or "workgroup" in fname.lower():
                    fields[fkey] = {
                        "name":           fname,
                        "required":       f.get("required"),
                        "schema":         f.get("schema"),
                        "allowedValues":  f.get("allowedValues", [])[:50],
                        "autoCompleteUrl":f.get("autoCompleteUrl"),
                    }
        result["incident_team_fields"] = {"status": r4.status_code, "fields": fields}

    # ── 4. Try fetching a real ITSD ticket that has a team set ────────────────
    r5 = _post("/search/jql", payload={
        "jql":        f'project = {cfg["helpdesk"]} AND "Team" is not EMPTY ORDER BY updated DESC',
        "startAt":    0,
        "maxResults": 3,
        "fields":     ["summary", "customfield_11600", "customfield_15997", "assignee", "status"],
    })
    ticket_samples = []
    if r5:
        for issue in r5.get("issues", []):
            f = issue.get("fields", {})
            ticket_samples.append({
                "key":                issue["key"],
                "summary":            f.get("summary"),
                "customfield_11600":  f.get("customfield_11600"),
                "customfield_15997":  f.get("customfield_15997"),
            })
    result["tickets_with_team"] = ticket_samples

    # ── 5. Same for ITO ───────────────────────────────────────────────────────
    r6 = _post("/search/jql", payload={
        "jql":        f'project = {cfg["ito"]} AND "Team" is not EMPTY ORDER BY updated DESC',
        "startAt":    0,
        "maxResults": 3,
        "fields":     ["summary", "customfield_11600", "customfield_15997", "assignee", "status"],
    })
    ito_samples = []
    if r6:
        for issue in r6.get("issues", []):
            f = issue.get("fields", {})
            ito_samples.append({
                "key":               issue["key"],
                "summary":           f.get("summary"),
                "customfield_11600": f.get("customfield_11600"),
                "customfield_15997": f.get("customfield_15997"),
            })
    result["ito_tickets_with_team"] = ito_samples

    return {"status": "ok", "data": result}


# ── Criticality Registry API ──────────────────────────────────────────────────

@app.get("/api/criticality")
@app.get("/infraportal/api/criticality")
async def api_criticality_list(request: Request):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import list_criticality
        return {"status": "ok", "entries": list_criticality()}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/api/criticality")
@app.post("/infraportal/api/criticality")
async def api_criticality_upsert(request: Request):
    user, _ = require_auth_check(request, "admin")
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import upsert_criticality
        body = await request.json()
        device_name = (body.get("device_name") or "").strip()
        tier        = (body.get("tier") or "P3").strip().upper()
        if not device_name:
            return {"status": "error", "error": "device_name is required"}
        if tier not in ("P1", "P2", "P3", "INFO"):
            return {"status": "error", "error": "tier must be P1, P2, P3, or INFO"}
        ok = upsert_criticality(
            device_name         = device_name,
            tier                = tier,
            service_description = body.get("service_description"),
            blast_radius        = body.get("blast_radius"),
            owner_team          = body.get("owner_team"),
            escalation_slack    = body.get("escalation_slack"),
            escalation_email    = body.get("escalation_email"),
            dependencies        = body.get("dependencies"),
            is_singleton        = bool(body.get("is_singleton", False)),
            notes               = body.get("notes"),
            set_by              = (user or {}).get("email", ""),
        )
        return {"status": "ok" if ok else "error"}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.delete("/api/criticality/{device_name:path}")
@app.delete("/infraportal/api/criticality/{device_name:path}")
async def api_criticality_delete(device_name: str, request: Request):
    user, _ = require_auth_check(request, "admin")
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import delete_criticality
        ok = delete_criticality(device_name)
        return {"status": "ok" if ok else "error"}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/api/criticality/bulk-delete")
@app.post("/infraportal/api/criticality/bulk-delete")
async def api_criticality_bulk_delete(request: Request):
    user, _ = require_auth_check(request, "admin")
    if not user:
        return {"status": "unauthorized"}
    try:
        body = await request.json()
        device_names = body.get("device_names", [])
        if not isinstance(device_names, list) or not device_names:
            return {"status": "error", "error": "device_names list is required"}
        from routers.database import bulk_delete_criticality
        count = bulk_delete_criticality(device_names)
        return {"status": "ok", "deleted": count}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/api/criticality/ai-suggest")
@app.post("/infraportal/api/criticality/ai-suggest")
async def api_criticality_ai_suggest(request: Request):
    user, _ = require_auth_check(request, "admin")
    if not user:
        return {"status": "unauthorized"}
    try:
        body = await request.json()
        device_name = (body.get("device_name") or "").strip()
        if not device_name:
            return {"status": "error", "error": "device_name is required"}

        from routers.cache import cache as _cache

        # Lansweeper enrichment
        ls_data = None
        try:
            assets, _ = _cache.get("lansweeper_assets")
            if assets:
                for a in assets:
                    if (a.get("name") or "").upper() == device_name.upper():
                        ls_data = a
                        break
        except Exception:
            pass

        # VMware enrichment
        vm_data = None
        try:
            detailed, _ = _cache.get("detailed_vms")
            if detailed:
                for vm in detailed.get("vms", []):
                    if (vm.get("name") or "").upper() == device_name.upper():
                        vm_data = vm
                        break
        except Exception:
            pass

        from routers.criticality import ai_suggest_criticality
        suggestion = ai_suggest_criticality(device_name, lansweeper_data=ls_data, vmware_data=vm_data)
        return {"status": "ok", "suggestion": suggestion}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.get("/api/criticality/groups")
@app.get("/infraportal/api/criticality/groups")
async def api_criticality_groups_list(request: Request):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import list_criticality_groups, get_member_counts, compute_all_blast_radii
        groups = list_criticality_groups()
        counts = get_member_counts()
        blast_radii = compute_all_blast_radii()
        for g in groups:
            g["member_count"] = counts.get(g["id"], 0)
            g["computed_blast_radius"] = blast_radii.get(g["group_name"], [])
        return {"status": "ok", "groups": groups}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/api/criticality/groups")
@app.post("/infraportal/api/criticality/groups")
async def api_criticality_groups_save(request: Request):
    user, _ = require_auth_check(request, "admin")
    if not user:
        return {"status": "unauthorized"}
    try:
        body = await request.json()
        if not body.get("group_name"):
            return {"status": "error", "error": "group_name is required"}
        # match_value is optional when an opm_group_name is provided
        if not body.get("match_value") and not body.get("opm_group_name"):
            return {"status": "error", "error": "Either match_value or opm_group_name is required"}
        from routers.database import upsert_criticality_group, get_member_counts
        gid = upsert_criticality_group(body)
        counts = get_member_counts()
        return {"status": "ok", "id": gid, "member_count": counts.get(gid, 0)}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/api/criticality/groups/{gid:int}/delete")
@app.post("/infraportal/api/criticality/groups/{gid:int}/delete")
async def api_criticality_groups_delete(request: Request, gid: int):
    user, _ = require_auth_check(request, "admin")
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import delete_criticality_group
        ok = delete_criticality_group(gid)
        return {"status": "ok" if ok else "error"}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.get("/api/criticality/groups/{gid:int}/members")
@app.get("/infraportal/api/criticality/groups/{gid:int}/members")
async def api_criticality_group_members(request: Request, gid: int):
    """Return all inventory devices that match a group's pattern, with registered/excluded status."""
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.cache import cache as _cache
        from routers.database import list_criticality_groups, get_criticality_map
        groups = list_criticality_groups()
        grp = next((g for g in groups if g["id"] == gid), None)
        if not grp:
            return {"status": "error", "error": "Group not found"}
        match_value = (grp.get("match_value") or "").strip()
        match_type  = grp.get("match_type", "prefix")
        if not match_value:
            return {"status": "ok", "members": [], "total": 0}
        crit_map     = get_criticality_map()
        excluded_set = {n.upper() for n in (grp.get("excluded_devices") or [])}

        def _matches(name):
            n = name.upper(); v = match_value.upper()
            if match_type == "prefix":   return n.startswith(v)
            if match_type == "suffix":   return n.endswith(v)
            if match_type == "contains": return v in n
            if match_type == "exact":    return n == v
            return False

        members = []; seen = set()

        def _add(nm, source, **extra):
            if nm and _matches(nm) and nm.upper() not in seen:
                members.append({"device_name": nm, "source": source,
                                 "registered": nm.upper() in crit_map,
                                 "excluded":   nm.upper() in excluded_set, **extra})
                seen.add(nm.upper())

        try:
            assets, _ = _cache.get("lansweeper_assets")
            if assets:
                for a in assets: _add((a.get("name") or "").strip(), "lansweeper",
                                      type=a.get("type",""), environment="")
        except Exception: pass
        try:
            det, _ = _cache.get("detailed_vms")
            if det:
                for vm in det.get("vms",[]): _add((vm.get("name") or "").strip(), "vmware",
                                                   type="VM", environment=vm.get("environment",""))
        except Exception: pass
        try:
            mach, _ = _cache.get("citrix_machines")
            if mach:
                for m in mach: _add((m.get("name") or "").strip(), "citrix",
                                    type=m.get("os_type",""), environment=m.get("delivery_group_name",""))
        except Exception: pass
        try:
            from routers import meraki as _mk
            for d in _mk._cache.get("devices",{}).get("data",[]):
                _add((d.get("name") or "").strip(), "meraki",
                     type=d.get("model",""), environment=d.get("network_name",""))
        except Exception: pass
        try:
            opm, _ = _cache.get("opm_devices")
            if opm:
                for d in opm: _add((d.get("display_name") or "").strip(), "opmanager",
                                   type=d.get("type",""), environment="")
        except Exception: pass

        members.sort(key=lambda x: (x["registered"], x["excluded"], x["device_name"].upper()))
        return {"status": "ok", "members": members, "total": len(members)}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/api/criticality/groups/{gid:int}/bulk-register")
@app.post("/infraportal/api/criticality/groups/{gid:int}/bulk-register")
async def api_criticality_groups_bulk_register(request: Request, gid: int):
    """Register all unclassified devices matching the group's pattern in one shot."""
    user, _ = require_auth_check(request, "admin")
    if not user:
        return {"status": "unauthorized"}
    try:
        body = await request.json()
        device_names = body.get("device_names", [])  # caller sends the list to register
        if not device_names:
            return {"status": "error", "error": "No device names provided"}
        from routers.database import list_criticality_groups, upsert_criticality
        groups = list_criticality_groups()
        grp = next((g for g in groups if g["id"] == gid), None)
        if not grp:
            return {"status": "error", "error": "Group not found"}
        # skip_exclusions=True when user manually assigns from Unclassified tab
        if not body.get("skip_exclusions"):
            excluded = {n.upper() for n in (grp.get("excluded_devices") or [])}
            if excluded:
                device_names = [n for n in device_names if n.upper() not in excluded]
        registered = []
        import time as _time
        for name in device_names:
            ok = upsert_criticality(
                device_name=name,
                tier=grp["default_tier"],
                service_description=grp.get("service_description") or "",
                blast_radius=grp.get("blast_radius") or "",
                owner_team=grp.get("owner_team") or "",
                escalation_slack=grp.get("escalation_slack") or "",
                escalation_email=grp.get("escalation_email") or "",
                is_singleton=bool(grp.get("is_singleton")),
                notes=grp.get("notes") or f"Auto-registered via group: {grp['group_name']}",
                set_by=(user or {}).get("email", ""),
            )
            if ok:
                registered.append(name)
        return {"status": "ok", "registered": registered, "count": len(registered)}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.get("/api/criticality/groups/{gid:int}/devices")
@app.get("/infraportal/api/criticality/groups/{gid:int}/devices")
async def api_criticality_group_devices_list(request: Request, gid: int):
    """Return explicit device_group_members for a group."""
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import list_group_members, list_criticality_groups, get_criticality_map
        groups = list_criticality_groups()
        grp = next((g for g in groups if g["id"] == gid), None)
        if not grp:
            return {"status": "error", "error": "Group not found"}
        members = list_group_members(gid)
        crit_map = get_criticality_map()
        for m in members:
            m["registered"] = m["device_name"].upper() in crit_map
        return {"status": "ok", "members": members, "total": len(members),
                "opm_group_name": grp.get("opm_group_name", "")}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/api/criticality/groups/{gid:int}/devices/add")
@app.post("/infraportal/api/criticality/groups/{gid:int}/devices/add")
async def api_criticality_group_devices_add(request: Request, gid: int):
    user, _ = require_auth_check(request, "admin")
    if not user:
        return {"status": "unauthorized"}
    try:
        body = await request.json()
        names = body.get("device_names", [])
        source = body.get("source", "manual")
        if not names:
            return {"status": "error", "error": "No device names provided"}
        from routers.database import add_group_members
        added = add_group_members(gid, names, source)
        return {"status": "ok", "added": added}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/api/criticality/groups/{gid:int}/devices/remove")
@app.post("/infraportal/api/criticality/groups/{gid:int}/devices/remove")
async def api_criticality_group_devices_remove(request: Request, gid: int):
    user, _ = require_auth_check(request, "admin")
    if not user:
        return {"status": "unauthorized"}
    try:
        body = await request.json()
        name = body.get("device_name", "").strip()
        if not name:
            return {"status": "error", "error": "device_name required"}
        from routers.database import remove_group_member
        remove_group_member(gid, name)
        return {"status": "ok"}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/api/criticality/groups/{gid:int}/sync-opm")
@app.post("/infraportal/api/criticality/groups/{gid:int}/sync-opm")
async def api_criticality_group_sync_opm(request: Request, gid: int):
    """Fetch all OPM devices in the linked OPM group and populate device_group_members."""
    user, _ = require_auth_check(request, "admin")
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import (list_criticality_groups, clear_opm_members,
                                       add_group_members, list_group_members,
                                       set_opm_sync_status)
        from routers.opmanager import _get_opm_group_member_names
        groups = list_criticality_groups()
        grp = next((g for g in groups if g["id"] == gid), None)
        if not grp:
            return {"status": "error", "error": "Group not found"}
        opm_group = (grp.get("opm_group_name") or "").strip()
        if not opm_group:
            return {"status": "error", "error": "No OPM group linked to this device group"}

        # Resolve members via getDeviceTree → listDevices?groupId=<numeric_id>
        device_names, sync_diag = _get_opm_group_member_names(opm_group)
        found_anywhere = any(e.get("group_found") for e in sync_diag)
        set_opm_sync_status(gid, found_anywhere)

        if not found_anywhere:
            # OPM is the source of truth for group membership — but if it no longer
            # reports this group under this name anywhere, that's more likely a stale
            # link than "the group is now empty." Leave existing membership alone and
            # just flag it, rather than silently wiping known-good data.
            total = len(list_group_members(gid))
            return {"status": "ok", "opm_group": opm_group, "matched": 0,
                    "removed_old": 0, "added": 0, "total": total,
                    "source_missing": True, "diag": sync_diag}

        removed = clear_opm_members(gid)
        added = add_group_members(gid, device_names, "opm")
        total = len(list_group_members(gid))
        return {"status": "ok", "opm_group": opm_group, "matched": len(device_names),
                "removed_old": removed, "added": added, "total": total,
                "source_missing": False, "diag": sync_diag}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/api/criticality/groups/sync-all-from-opm")
@app.post("/infraportal/api/criticality/groups/sync-all-from-opm")
async def api_criticality_groups_sync_all_opm(request: Request):
    """Re-sync device membership for every group that has an opm_group_name set."""
    user, _ = require_auth_check(request, "admin")
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.opmanager import _get_opm_group_member_names, invalidate_opm_group_cache
        from routers.database import (list_criticality_groups, clear_opm_members,
                                       add_group_members, get_member_counts,
                                       set_opm_sync_status)

        # Clear cached group maps so bulk sync always uses fresh data
        invalidate_opm_group_cache()

        groups = list_criticality_groups()
        results = []
        missing_sources = []
        for g in groups:
            opm_name = (g.get("opm_group_name") or "").strip()
            if not opm_name:
                continue
            device_names, diag = _get_opm_group_member_names(opm_name)
            found_anywhere = any(e.get("group_found") for e in diag)
            set_opm_sync_status(g["id"], found_anywhere)

            if not found_anywhere:
                # Don't wipe known-good membership just because OPM didn't report the
                # group this time — flag it for review instead (see single-group sync).
                missing_sources.append(g["group_name"])
                results.append({"group": g["group_name"], "opm_group": opm_name,
                                "added": 0, "removed": 0, "source_missing": True})
                continue

            removed = clear_opm_members(g["id"])
            added = add_group_members(g["id"], device_names, "opm")
            results.append({"group": g["group_name"], "opm_group": opm_name,
                            "added": added, "removed": removed, "source_missing": False})

        counts = get_member_counts()
        return {"status": "ok", "synced": len(results), "results": results,
                "missing_sources": missing_sources,
                "total_devices": sum(counts.values())}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.get("/api/criticality/opm-group-devices")
@app.get("/infraportal/api/criticality/opm-group-devices")
async def api_criticality_opm_group_devices(request: Request, name: str = ""):
    """Diagnostic: raw OPM device list for a group name, plus the same per-server
    match diagnostics _get_opm_group_member_names() uses for Sync — lets you see
    directly why a sync produced fewer devices than OpManager actually reports
    (group not found under this exact name, OPM's own count showing 0, etc.)."""
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    if not name:
        return {"status": "error", "error": "Missing group name"}
    try:
        from routers.opmanager import _get_opm_group_member_names
        devices, diag = _get_opm_group_member_names(name)
        return {"status": "ok", "devices": devices, "diag": diag}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.get("/api/criticality/opm-probe-diag")
@app.get("/infraportal/api/criticality/opm-probe-diag")
async def api_opm_probe_diag(request: Request):
    """Diagnostic: show raw probe device/group data to debug missing groups."""
    user, _ = require_auth_check(request, "admin")
    if not user:
        return {"status": "unauthorized"}
    try:
        import requests as _req
        from routers.opmanager import BASE_URL, PROBE_URLS, OPM_API_KEY, PROBE_API_KEY, PROBE_KEY_MAP, _get
        from routers.cache import cache as _cache

        result = {"central_url": BASE_URL, "probe_urls": PROBE_URLS,
                  "probe_keys": {url: key[:8] + "..." for url, key in PROBE_KEY_MAP.items()},
                  "probe_api_key_set": bool(PROBE_API_KEY and PROBE_API_KEY != OPM_API_KEY),
                  "central": {}, "probes": []}

        def probe_endpoint(base_url, ep, api_key=None, params=None):
            p = {"apiKey": api_key or OPM_API_KEY}
            if params:
                p.update(params)
            r = _req.get(base_url + ep, params=p, verify=False, timeout=15)
            r.raise_for_status()
            raw = r.json()
            out: dict = {"http_status": r.status_code}
            if isinstance(raw, dict):
                out["top_keys"] = list(raw.keys())
                for k, v in raw.items():
                    if isinstance(v, list):
                        out[f"key_{k}_count"] = len(v)
                        if v and isinstance(v[0], dict):
                            out[f"key_{k}_sample_fields"] = list(v[0].keys())
                out["raw_snippet"] = str(raw)[:600]
            elif isinstance(raw, list):
                out["top_type"] = "list"
                out["count"] = len(raw)
                if raw and isinstance(raw[0], dict):
                    out["sample_fields"] = list(raw[0].keys())
                out["raw_snippet"] = str(raw[:2])[:600]
            return out

        # Central: device sample — check mapName, probeName and all candidate group fields
        try:
            cen_raw = _get("/device/listDevices")
            cen_items = cen_raw if isinstance(cen_raw, list) else (cen_raw.get("data") or [])
            result["central"]["device_count"] = len(cen_items)
            if cen_items:
                result["central"]["sample_device_keys"] = list(cen_items[0].keys())
                # Show every field that could encode a group: map, group, category, probe, type
                result["central"]["sample_candidate_fields"] = {
                    k: cen_items[0].get(k) for k in cen_items[0].keys()
                    if any(kw in k.lower() for kw in ("group", "category", "type", "class", "map", "probe"))
                }
                # Distribution of mapName across ALL devices — key question: does it contain app group names?
                map_dist: dict[str, int] = {}
                probe_dist: dict[str, int] = {}
                for d in cen_items:
                    mn = (d.get("mapName") or "").strip() or "(none)"
                    pn = (d.get("probeName") or d.get("probeDisplayName") or "").strip() or "(none)"
                    map_dist[mn]   = map_dist.get(mn, 0) + 1
                    probe_dist[pn] = probe_dist.get(pn, 0) + 1
                # Show top 40 by count
                result["central"]["mapName_distribution"] = dict(
                    sorted(map_dist.items(), key=lambda x: -x[1])[:40])
                result["central"]["probe_distribution"] = dict(
                    sorted(probe_dist.items(), key=lambda x: -x[1])[:20])
        except Exception as e:
            result["central"]["device_error"] = str(e)

        # Central: try group/business-service endpoints on regular REST API
        result["central"]["group_endpoints"] = {}
        for ep in [
            "/group/listGroups", "/group/listGroup", "/group/getAllGroups",
            "/device/listDeviceGroups", "/group/getGroupDetails",
            # device tree — never tested on regular REST path
            "/device/getDeviceTree",
            # business views / application maps
            "/bv/listBusinessViews", "/bv/getBusinessViews",
            "/businessservice/listAll", "/application/listApplications",
            "/monitor/listMonitors",
        ]:
            try:
                params = {"groupBy": "group", "pageName": "groups"} if "getDeviceTree" in ep else None
                result["central"]["group_endpoints"][ep] = probe_endpoint(BASE_URL, ep, params=params)
            except Exception as e:
                result["central"]["group_endpoints"][ep] = {"error": str(e)}

        # Probe: test both /api/json/ and /client/api/json/ base paths
        # Browser DevTools revealed the real internal API is at /client/api/json/
        OPM_PORT_STR = os.getenv("OPMANAGER_PORT", "8060")
        for probe_url in PROBE_URLS:
            host = probe_url.split("/api/json")[0]   # e.g. https://awsopmanprb01p:8060
            client_base = host + "/client/api/json"  # the internal Ember API path
            _pkey = PROBE_KEY_MAP.get(probe_url, PROBE_API_KEY)
            entry: dict = {"url": probe_url, "client_base": client_base,
                           "api_key_used": (_pkey or "")[:8] + "...",
                           "old_path_endpoints": {}, "client_path_endpoints": {}}

            # Regular /api/json/ REST path — try getDeviceTree + group endpoints
            for ep, ep_params in [
                ("/device/listDevices", None),
                ("/device/getDeviceTree", {"groupBy": "group", "pageName": "groups"}),
                ("/group/listGroups", None),
                ("/bv/listBusinessViews", None),
                ("/businessservice/listAll", None),
                # Test if listAllLogicalGroups exists on the REST path (not just client path)
                ("/admin/listAllLogicalGroups", {"isFluidic": "true", "isGroupPage": "true", "page": "0"}),
                ("/admin/listAllLogicalGroups", {"isGroupPage": "true", "_search": "false", "rows": "100", "page": "1", "sortByColumn": "groupDisplayName", "sortByType": "asc"}),
            ]:
                try:
                    entry["old_path_endpoints"][ep] = probe_endpoint(
                        probe_url, ep, api_key=_pkey, params=ep_params)
                except Exception as e:
                    entry["old_path_endpoints"][ep] = {"error": str(e)}

            # /client/api/json/ — try session login first, then use cookie for group endpoints
            entry["client_path_endpoints"] = {}
            session_cookie = None
            ntlm_auth = None
            try:
                from requests_ntlm import HttpNtlmAuth
                AD_USER = os.getenv("AD_USER", "")
                AD_PASSWORD = os.getenv("AD_PASSWORD", "")
                AD_DOMAIN = os.getenv("AD_DOMAIN", "sbl")
                ntlm_auth = HttpNtlmAuth(f"{AD_DOMAIN}\\{AD_USER}", AD_PASSWORD)
                entry["client_path_endpoints"]["ntlm_available"] = True
            except ImportError:
                entry["client_path_endpoints"]["ntlm_available"] = False
            try:
                # Try exchanging the API key for a session cookie
                login_resp = _req.get(
                    client_base + "/user/loginFromApiKey",
                    params={"apiKey": _pkey},
                    verify=False, timeout=10)
                entry["client_path_endpoints"]["login_attempt"] = {
                    "url": str(login_resp.url),
                    "status": login_resp.status_code,
                    "cookies": dict(login_resp.cookies),
                    "snippet": login_resp.text[:400],
                }
                if login_resp.cookies:
                    session_cookie = dict(login_resp.cookies)
            except Exception as e:
                entry["client_path_endpoints"]["login_attempt"] = {"error": str(e)}

            # Group endpoints — three auth strategies for each
            for ep, extra_params in [
                ("/admin/listAllLogicalGroups", {"isGroupPage": "true", "_search": "false", "rows": "100", "page": "1", "sortByColumn": "groupDisplayName", "sortByType": "asc"}),
                ("/device/getDeviceTree", {"groupBy": "group", "pageName": "groups"}),
            ]:
                ep_label = ep.split("/")[-1]
                # Strategy A: session cookie (if login succeeded)
                if session_cookie:
                    label = f"{ep_label}_A_cookie"
                    try:
                        r = _req.get(client_base + ep, params=extra_params,
                                     cookies=session_cookie, verify=False, timeout=15)
                        raw = r.json()
                        out: dict = {"http_status": r.status_code, "auth": "cookie"}
                        if isinstance(raw, dict):
                            out["top_keys"] = list(raw.keys())
                            out["raw_snippet"] = str(raw)[:1200]
                        elif isinstance(raw, list):
                            out["count"] = len(raw)
                            out["raw_snippet"] = str(raw[:3])[:1200]
                        entry["client_path_endpoints"][label] = out
                    except Exception as e:
                        entry["client_path_endpoints"][label] = {"error": str(e)}

                # Strategy B: no auth at all — bare request
                label_b = f"{ep_label}_B_noauth"
                try:
                    r = _req.get(client_base + ep, params=extra_params,
                                 verify=False, timeout=15)
                    raw = r.json()
                    out_b: dict = {"http_status": r.status_code, "auth": "none"}
                    if isinstance(raw, dict):
                        out_b["top_keys"] = list(raw.keys())
                        out_b["raw_snippet"] = str(raw)[:1200]
                    elif isinstance(raw, list):
                        out_b["count"] = len(raw)
                        out_b["raw_snippet"] = str(raw[:3])[:1200]
                    entry["client_path_endpoints"][label_b] = out_b
                except Exception as e:
                    entry["client_path_endpoints"][label_b] = {"error": str(e)}

                # Strategy C: apiKey param (our existing approach — documents why it fails)
                label_c = f"{ep_label}_C_apikey"
                try:
                    r = _req.get(client_base + ep,
                                 params={"apiKey": _pkey, **extra_params},
                                 verify=False, timeout=15)
                    raw = r.json()
                    out_c: dict = {"http_status": r.status_code, "auth": "apikey"}
                    if isinstance(raw, dict):
                        out_c["top_keys"] = list(raw.keys())
                        out_c["raw_snippet"] = str(raw)[:600]
                    elif isinstance(raw, list):
                        out_c["count"] = len(raw)
                        out_c["raw_snippet"] = str(raw[:2])[:600]
                    entry["client_path_endpoints"][label_c] = out_c
                except Exception as e:
                    entry["client_path_endpoints"][label_c] = {"error": str(e)}

                # Strategy D: NTLM Windows auth (AD-integrated environment)
                if ntlm_auth:
                    label_d = f"{ep_label}_D_ntlm"
                    try:
                        r = _req.get(client_base + ep, params=extra_params,
                                     auth=ntlm_auth, verify=False, timeout=20)
                        raw = r.json()
                        out_d: dict = {"http_status": r.status_code, "auth": "ntlm"}
                        if isinstance(raw, dict):
                            out_d["top_keys"] = list(raw.keys())
                            out_d["raw_snippet"] = str(raw)[:1200]
                        elif isinstance(raw, list):
                            out_d["count"] = len(raw)
                            out_d["raw_snippet"] = str(raw[:3])[:1200]
                        entry["client_path_endpoints"][label_d] = out_d
                    except Exception as e:
                        entry["client_path_endpoints"][label_d] = {"error": str(e)}

            result["probes"].append(entry)

        return {"status": "ok", "diag": result}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/api/criticality/groups/auto-create-from-opm")
@app.post("/infraportal/api/criticality/groups/auto-create-from-opm")
async def api_criticality_groups_auto_create_opm(request: Request):
    """Create one device group per OPM group (skips any that already exist by opm_group_name)."""
    user, _ = require_auth_check(request, "admin")
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers import opmanager as _opm
        from routers.database import (list_criticality_groups, upsert_criticality_group,
                                       add_group_members, get_member_counts)
        from routers.cache import cache as _cache

        opm_groups = _opm.get_named_groups()
        existing_groups = list_criticality_groups()
        existing_opm_names = {(g.get("opm_group_name") or "").lower() for g in existing_groups}

        opm_devices, _ = _cache.get("opm_devices")
        if not opm_devices:
            opm_devices, _ = _opm.get_devices(force_refresh=True)

        # Build device index by group name
        devices_by_group: dict[str, list[str]] = {}
        for d in (opm_devices or []):
            gn = (d.get("group_name") or "").strip()
            dn = (d.get("display_name") or "").strip()
            if gn and dn:
                devices_by_group.setdefault(gn.lower(), []).append(dn)

        created = []
        skipped = []
        for og in opm_groups:
            name = (og.get("name") or "").strip()
            if not name:
                continue
            if name.lower() in existing_opm_names:
                skipped.append(name)
                continue
            new_gid = upsert_criticality_group({
                "group_name":    name,
                "opm_group_name": name,
                "match_type":    "prefix",
                "match_value":   "",
                "default_tier":  "P3",
                "owner_team":    "",
                "is_singleton":  False,
                "service_description": f"Imported from OPM group: {name}",
                "blast_radius":  "",
                "notes":         "Auto-created from OPM group — use AI Draft to enrich metadata",
            })
            device_names = devices_by_group.get(name.lower(), [])
            if device_names:
                add_group_members(new_gid, device_names, "opm")
            created.append({"name": name, "id": new_gid, "devices": len(device_names)})

        return {"status": "ok", "created": created, "skipped": skipped,
                "created_count": len(created), "skipped_count": len(skipped)}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/api/criticality/groups/ai-suggest")
@app.post("/infraportal/api/criticality/groups/ai-suggest")
async def api_criticality_groups_ai_suggest(request: Request):
    user, _ = require_auth_check(request, "admin")
    if not user:
        return {"status": "unauthorized"}
    try:
        body = await request.json()
        group_name  = (body.get("group_name") or "").strip()
        match_type  = (body.get("match_type") or "prefix").strip()
        match_value = (body.get("match_value") or "").strip()
        group_id    = body.get("group_id")
        if not group_name and not match_value:
            return {"status": "error", "error": "group_name or match_value is required"}

        # ── Deterministic environment inference from group name ─────────────────
        def _infer_env(n: str) -> str | None:
            n = n.lower()
            if any(x in n for x in ["non-prod", "nonprod", "non prod"]):
                return "QA"
            if re.search(r'[-_ ]prod$|[-_ ]prod[-_ ]|^prod[-_ ]| production', n):
                return "Prod"
            if re.search(r'[-_ ]qa$|[-_ ]qa[-_ ]|^qa[-_ ]', n) or n.endswith(" qa"):
                return "QA"
            if re.search(r'[-_ ]dev$|[-_ ]dev[-_ ]', n):
                return "Dev"
            if re.search(r'[-_ ]test$|[-_ ]test[-_ ]', n):
                return "Test"
            if re.search(r'[-_ ]uat$|[-_ ]uat[-_ ]', n):
                return "UAT"
            return None
        suggested_env = _infer_env(group_name)

        # ── Load existing OPM members when editing ──────────────────────────────
        existing_member_names: list[str] = []
        if group_id:
            try:
                from routers.database import list_group_members
                existing_member_names = [m["device_name"] for m in list_group_members(group_id)]
            except Exception:
                pass

        # ── Derive match pattern from existing member names ─────────────────────
        # Strip domain suffix, extract alphabetic prefix, find common prefix.
        def _strip_domain(name: str) -> str:
            return name.split(".")[0].upper()

        def _alpha_prefix(s: str) -> str:
            return re.match(r'^[A-Z]+', s).group() if re.match(r'^[A-Z]+', s) else ""

        if existing_member_names and not match_value:
            clean = [_strip_domain(n) for n in existing_member_names if n]
            prefixes = [_alpha_prefix(c) for c in clean if _alpha_prefix(c)]
            if prefixes:
                # Find longest common prefix among all member prefixes (min 4 chars)
                def _common_prefix(strs):
                    if not strs: return ""
                    s = min(strs, key=len)
                    for i, c in enumerate(s):
                        if any(x[i] != c for x in strs):
                            return s[:i]
                    return s
                cp = _common_prefix(prefixes)
                if len(cp) >= 4:
                    match_type  = "prefix"
                    match_value = cp

        # Detect which integrations are most relevant based on group name/value keywords
        kw = (group_name + " " + match_value).lower()
        use_meraki = any(k in kw for k in ["cisco", "meraki", "network", "switch", "nexus", "ap ", "wifi", "wireless", "firewall", " mx", " ms ", "catalyst"])
        use_citrix = any(k in kw for k in ["citrix", "ctx", "vdi", "xen", "wem", "xdmi", "vda"])
        use_opm    = any(k in kw for k in ["monitor", "opm", "opmanager", "scom", "probe"])

        from routers.cache import cache as _cache
        from routers.database import get_criticality_map

        crit_map = get_criticality_map()
        pattern_matches = []  # devices matching the name pattern (all sources)
        broad_context   = []  # integration-specific devices even without name match
        seen = set()
        # Existing members should not appear as new suggestions
        existing_upper = {_strip_domain(n) for n in existing_member_names}

        def _matches(name):
            if not match_value:
                return False  # empty pattern matches everything — never match
            n = name.upper()
            v = match_value.upper()
            if match_type == "prefix":   return n.startswith(v)
            if match_type == "suffix":   return n.endswith(v)
            if match_type == "contains": return v in n
            if match_type == "exact":    return n == v
            return False

        def _try_add_pattern(nm, source, **extra):
            short = _strip_domain(nm) if nm else ""
            if nm and nm.upper() not in crit_map and nm.upper() not in seen and short not in existing_upper and _matches(nm):
                pattern_matches.append({"device_name": nm, "source": source, **extra})
                seen.add(nm.upper())

        # --- Lansweeper (always) ---
        try:
            assets, _ = _cache.get("lansweeper_assets")
            if assets:
                for a in assets:
                    nm = (a.get("name") or "").strip()
                    _try_add_pattern(nm, "lansweeper",
                                     os=a.get("os") or a.get("os_name", ""),
                                     type=a.get("type", ""))
        except Exception:
            pass

        # --- VMware (always) ---
        try:
            detailed, _ = _cache.get("detailed_vms")
            if detailed:
                for vm in detailed.get("vms", []):
                    nm = (vm.get("name") or "").strip()
                    _try_add_pattern(nm, "vmware",
                                     os=vm.get("os_name", ""),
                                     type="VM",
                                     environment=vm.get("environment", ""))
        except Exception:
            pass

        # --- Meraki (pattern match always; broad context when Cisco/network keywords detected) ---
        try:
            from routers import meraki as _meraki_mod
            meraki_devices = _meraki_mod._cache.get("devices", {}).get("data", [])
            meraki_broad: list = []
            for d in meraki_devices:
                nm = (d.get("name") or "").strip()
                extra = {
                    "type": d.get("model", ""),
                    "os": "/".join(d.get("product_types", [])),
                    "environment": d.get("network_name", ""),
                }
                if not nm or nm.upper() in crit_map or nm.upper() in seen:
                    continue
                if _matches(nm):
                    pattern_matches.append({"device_name": nm, "source": "meraki", **extra})
                    seen.add(nm.upper())
                elif use_meraki and len(meraki_broad) < 20:
                    meraki_broad.append({"device_name": nm, "source": "meraki", **extra})
                    seen.add(nm.upper())
            broad_context.extend(meraki_broad)
        except Exception:
            pass

        # --- Citrix (pattern match always; broad context when Citrix keywords detected) ---
        try:
            machines, _ = _cache.get("citrix_machines")
            if machines:
                citrix_broad: list = []
                for m in machines:
                    nm = (m.get("name") or "").strip()
                    extra = {
                        "type": m.get("os_type", ""),
                        "os": m.get("os_type", ""),
                        "environment": m.get("delivery_group_name", ""),
                    }
                    if not nm or nm.upper() in crit_map or nm.upper() in seen:
                        continue
                    if _matches(nm):
                        pattern_matches.append({"device_name": nm, "source": "citrix", **extra})
                        seen.add(nm.upper())
                    elif use_citrix and len(citrix_broad) < 20:
                        citrix_broad.append({"device_name": nm, "source": "citrix", **extra})
                        seen.add(nm.upper())
                broad_context.extend(citrix_broad)
        except Exception:
            pass

        # --- OpManager devices (pattern match always; broad context when monitoring keywords) ---
        try:
            opm_devs, _ = _cache.get("opm_devices")
            if opm_devs:
                opm_broad: list = []
                for d in opm_devs:
                    nm = (d.get("display_name") or "").strip()
                    extra = {
                        "type": d.get("type", ""),
                        "os": d.get("vendor", ""),
                        "environment": d.get("category", ""),
                    }
                    if not nm or nm.upper() in crit_map or nm.upper() in seen:
                        continue
                    if _matches(nm):
                        pattern_matches.append({"device_name": nm, "source": "opmanager", **extra})
                        seen.add(nm.upper())
                    elif use_opm and len(opm_broad) < 20:
                        opm_broad.append({"device_name": nm, "source": "opmanager", **extra})
                        seen.add(nm.upper())
                broad_context.extend(opm_broad)
        except Exception:
            pass

        # Pattern matches first, then integration-specific broad context
        sample_devices = pattern_matches + broad_context

        from routers.criticality import ai_suggest_group
        suggestion = ai_suggest_group(group_name, match_type, match_value, sample_devices,
                                      existing_members=existing_member_names)
        # Include deterministic environment suggestion in suggestion dict
        if suggested_env and not suggestion.get("environment"):
            suggestion["environment"] = suggested_env
        # Return up to 50 sample devices so the frontend can show a preview
        # (these are ADDITIONAL candidates not yet in the group)
        preview = [{"device_name": d["device_name"], "source": d.get("source",""),
                    "type": d.get("type",""), "environment": d.get("environment","")}
                   for d in sample_devices[:50]]
        return {"status": "ok", "suggestion": suggestion,
                "sample_count": len(sample_devices), "samples": preview,
                "existing_count": len(existing_member_names),
                "derived_match_value": match_value if existing_member_names else None}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.get("/api/criticality/inventory")
@app.get("/infraportal/api/criticality/inventory")
async def api_criticality_inventory(request: Request):
    """Return all Lansweeper + VMware devices not yet in the criticality registry."""
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.cache import cache as _cache
        from routers.database import get_criticality_map
        crit_map = get_criticality_map()
        devices = []
        seen = set()

        try:
            assets, _ = _cache.get("lansweeper_assets")
            if assets:
                for a in assets:
                    nm = (a.get("name") or "").strip()
                    if not nm or nm.upper() in crit_map or nm.upper() in seen:
                        continue
                    devices.append({
                        "device_name": nm,
                        "source":      "lansweeper",
                        "os":          a.get("os") or a.get("os_name", ""),
                        "type":        a.get("type", ""),
                        "last_seen":   a.get("last_seen", ""),
                        "ip":          a.get("ip") or a.get("ip_address", ""),
                    })
                    seen.add(nm.upper())
        except Exception:
            pass

        try:
            detailed, _ = _cache.get("detailed_vms")
            if detailed:
                for vm in detailed.get("vms", []):
                    nm = (vm.get("name") or "").strip()
                    if not nm or nm.upper() in crit_map or nm.upper() in seen:
                        continue
                    devices.append({
                        "device_name":  nm,
                        "source":       "vmware",
                        "os":           vm.get("os_name", ""),
                        "type":         "VM",
                        "power_state":  vm.get("power_state", ""),
                        "environment":  vm.get("environment", ""),
                        "ip":           vm.get("ip_address", ""),
                    })
                    seen.add(nm.upper())
        except Exception:
            pass

        return {"status": "ok", "devices": devices, "count": len(devices)}
    except Exception as e:
        return {"status": "error", "error": str(e)}


# ── Device IP Enrichment ─────────────────────────────────────────────────────

import re as _re
import threading as _threading

_IP_RE = _re.compile(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$')
_enrich_cache:   dict = {}   # name → {resolved_name, source, os, type} | None
_enrich_pending: set  = set()  # IPs currently being DNS-resolved in background


def _build_inventory_maps(cache):
    """Return (by_ip, by_name) dicts built from OPM/Lansweeper/VMware caches."""
    by_ip:   dict = {}
    by_name: dict = {}

    def _add(resolved_name, source, os_, type_, ip=""):
        entry = {"resolved_name": resolved_name, "source": source, "os": os_, "type": type_}
        if ip:
            existing = by_ip.get(ip)
            if not existing or (not existing.get("resolved_name") and resolved_name):
                by_ip[ip] = entry
        if resolved_name:
            by_name.setdefault(resolved_name.upper(), entry)

    try:
        opm_devs, _ = cache.get("opm_devices")
        for d in (opm_devs or []):
            ip = (d.get("ip_address") or "").strip()
            nm = (d.get("display_name") or "").strip()
            if nm and not _IP_RE.match(nm):
                _add(nm, "OpManager", d.get("type", ""), d.get("category", ""), ip)
            elif ip:
                by_ip.setdefault(ip, {"resolved_name": "", "source": "OpManager",
                                      "os": d.get("type", ""), "type": d.get("category", "")})
    except Exception:
        pass

    try:
        ls_assets, _ = cache.get("lansweeper_assets")
        for a in (ls_assets or []):
            ip = (a.get("ip") or a.get("ip_address") or "").strip()
            nm = (a.get("name") or "").strip()
            if nm:
                _add(nm, "Lansweeper",
                     a.get("os") or a.get("os_name", ""),
                     a.get("type", ""), ip)
    except Exception:
        pass

    try:
        detail, _ = cache.get("detailed_vms")
        for vm in (detail or {}).get("vms", []):
            ip = (vm.get("ip_address") or "").strip()
            nm = (vm.get("name") or "").strip()
            if nm:
                _add(nm, "VMware", vm.get("os_name", ""), "VM", ip)
    except Exception:
        pass

    return by_ip, by_name


def _dns_resolve_batch(ips: list) -> None:
    """Spawn one thread per IP (max 20 concurrent) for parallel PTR lookups."""
    import socket, threading as _t
    socket.setdefaulttimeout(3)   # set once, process-wide; safe for DNS-only use
    sem = _t.Semaphore(20)

    def _one(ip):
        with sem:
            existing = _enrich_cache.get(ip) or {}
            if existing.get("resolved_name"):
                _enrich_pending.discard(ip)
                return
            try:
                hostname, _, _ = socket.gethostbyaddr(ip)
                _enrich_cache[ip] = {
                    "resolved_name": hostname,
                    "source": existing.get("source") or "DNS",
                    "os": existing.get("os", ""),
                    "type": existing.get("type", ""),
                }
            except Exception:
                if not existing:
                    _enrich_cache[ip] = None
            finally:
                _enrich_pending.discard(ip)

    for ip in ips:
        _t.Thread(target=_one, args=(ip,), daemon=True).start()


@app.post("/api/criticality/enrich")
@app.post("/infraportal/api/criticality/enrich")
async def api_criticality_enrich(request: Request):
    """
    Phase 1 (instant): cross-reference names/IPs against OPM/Lansweeper/VMware caches.
    Phase 2 (async):   for IP-only names not found in inventory, fire DNS PTR lookups
                       in a background thread; frontend polls once for those results.
    Returns: {enrichment, dns_pending: bool}
    """
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        body  = await request.json()
        names = [n.strip() for n in (body.get("names") or []) if n.strip()]
        if not names:
            return {"status": "ok", "enrichment": {}, "dns_pending": False}

        from routers.cache import cache as _cache
        by_ip, by_name = _build_inventory_maps(_cache)

        result:    dict = {}
        need_dns:  list = []

        for name in names:
            # Already cached (inventory hit or previous DNS run)
            if name in _enrich_cache:
                r = _enrich_cache[name]
                if r and (r.get("resolved_name") or r.get("os") or r.get("type")):
                    result[name] = r
                # If it's a pending DNS IP, it's still in-flight — don't re-queue
                continue

            if _IP_RE.match(name):
                # Try inventory first
                inv = by_ip.get(name)
                if inv:
                    result[name] = inv
                    if inv.get("resolved_name"):
                        _enrich_cache[name] = inv
                    else:
                        # Inventory has OS/type but no hostname — show what we have
                        # and still queue DNS so the name can be resolved
                        result[name] = inv
                        if name not in _enrich_pending:
                            need_dns.append(name)
                else:
                    # Not in any inventory — queue for DNS
                    if name not in _enrich_pending:
                        need_dns.append(name)
            else:
                # Hostname — look up by name for OS/type enrichment
                inv = by_name.get(name.upper())
                _enrich_cache[name] = inv
                if inv and (inv.get("os") or inv.get("type")):
                    result[name] = inv

        # Fire DNS batch in background for unresolved IPs
        if need_dns:
            _enrich_pending.update(need_dns)
            _threading.Thread(target=_dns_resolve_batch, args=(need_dns,),
                              daemon=True).start()

        still_pending = bool(_enrich_pending.intersection(names))
        return {"status": "ok", "enrichment": result, "dns_pending": still_pending}
    except Exception as e:
        return {"status": "error", "error": str(e)}


# ── Criticality Topology (new) ───────────────────────────────────────────────

@app.get("/api/criticality/topology")
@app.get("/infraportal/api/criticality/topology")
async def api_criticality_topology(request: Request):
    """
    Returns groups enriched with alert status, location, device lists, plus
    dependency edges and a geographic summary — all in one payload.
    """
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.opmanager import get_named_groups
        from routers.database import list_group_dependencies

        force = request.query_params.get("refresh") == "1"
        group_nodes = get_named_groups(force_refresh=force)

        deps = list_group_dependencies()

        # Also include any group names referenced in dependencies that aren't already in the list
        known_names = {g["name"] for g in group_nodes}
        for d in deps:
            for gname in (d["from_group"], d["to_group"]):
                if gname and gname not in known_names:
                    group_nodes.append({
                        "id": gname, "name": gname, "description": "", "device_count": None,
                        "alert_count": 0, "worst_sev": 5, "status": "clear",
                        "location": "Unknown", "source": "dependency",
                    })
                    known_names.add(gname)

        # Root-cause scoring: which groups, if failed, would cascade the most?
        dep_graph: dict = {}  # to_group → [from_group, ...]
        for d in deps:
            dep_graph.setdefault(d["to_group"], []).append(d["from_group"])

        alerting_groups = {n["name"] for n in group_nodes if n["status"] in ("critical","warning","minor")}
        root_candidates = []
        for gname_rc, dependents in dep_graph.items():
            if gname_rc not in alerting_groups:
                continue
            cascade_score = sum(1 for d in dependents if d in alerting_groups)
            if cascade_score:
                root_candidates.append({"group": gname_rc, "cascade_score": cascade_score,
                                        "dependents_in_alarm": [d for d in dependents if d in alerting_groups]})
        root_candidates.sort(key=lambda x: -x["cascade_score"])

        return {
            "status":          "ok",
            "groups":          sorted(group_nodes, key=lambda g: g["name"].lower()),
            "dependencies":    deps,
            "root_candidates": root_candidates[:5],
        }
    except Exception as e:
        import traceback
        return {"status": "error", "error": str(e), "trace": traceback.format_exc()[-800:]}


@app.get("/api/criticality/dependencies")
@app.get("/infraportal/api/criticality/dependencies")
async def api_list_dependencies(request: Request):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import list_group_dependencies
        return {"status": "ok", "dependencies": list_group_dependencies()}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/api/criticality/dependencies")
@app.post("/infraportal/api/criticality/dependencies")
async def api_add_dependency(request: Request):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import add_group_dependency
        body = await request.json()
        dep_id = add_group_dependency(
            from_group   = body.get("from_group",""),
            to_group     = body.get("to_group",""),
            dep_type     = body.get("dep_type","application"),
            confidence   = body.get("confidence","medium"),
            ai_suggested = bool(body.get("ai_suggested")),
            notes        = body.get("notes",""),
        )
        if dep_id is None:
            return {"status": "duplicate", "message": "This dependency already exists."}
        return {"status": "ok", "id": dep_id}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/api/criticality/dependencies/{dep_id:int}/delete")
@app.post("/infraportal/api/criticality/dependencies/{dep_id:int}/delete")
async def api_delete_dependency(request: Request, dep_id: int):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import delete_group_dependency
        ok = delete_group_dependency(dep_id)
        return {"status": "ok" if ok else "error"}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/api/criticality/dependencies/suggest")
@app.post("/infraportal/api/criticality/dependencies/suggest")
async def api_dep_suggest(request: Request):
    """Fire-and-forget AI dependency suggestion across all Device Groups.
    Sources candidates from criticality_groups only (not raw OpManager groups) —
    suggesting edges for OPM's generic device-type buckets ("Server", "Switch",
    etc.) is exactly what produced the phantom dependency edges cleaned up
    alongside this fix."""
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.criticality import run_dep_suggest_background, get_dep_suggest_status
        from routers.database import list_criticality_groups, list_group_members, list_locations

        state = get_dep_suggest_status()
        if state["running"]:
            return {"status": "already_running"}

        groups = list_criticality_groups()
        loc_by_id = {l["id"]: l["name"] for l in list_locations()}

        enriched = []
        for g in groups:
            members = list_group_members(g["id"])
            enriched.append({
                "id":           str(g["id"]),
                "name":         g["group_name"],
                "description":  g.get("service_description", ""),
                "device_count": len(members),
                "location":     loc_by_id.get(g.get("location_id"), "?"),
                "devices":      [m["device_name"] for m in members[:6] if m.get("device_name")],
            })

        run_dep_suggest_background(enriched)
        return {"status": "started"}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.get("/api/criticality/dependencies/suggest/status")
@app.get("/infraportal/api/criticality/dependencies/suggest/status")
async def api_dep_suggest_status(request: Request):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.criticality import get_dep_suggest_status
        return {"status": "ok", **get_dep_suggest_status()}
    except Exception as e:
        return {"status": "error", "error": str(e)}


# ── Topology Locations ────────────────────────────────────────────────────────

@app.get("/api/criticality/locations")
@app.get("/infraportal/api/criticality/locations")
async def api_list_locations(request: Request):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import list_locations
        return {"status": "ok", "locations": list_locations()}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/api/criticality/locations")
@app.post("/infraportal/api/criticality/locations")
async def api_upsert_location(request: Request):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import upsert_location
        body = await request.json()
        if not (body.get("name") or "").strip():
            return {"status": "error", "error": "name is required"}
        lid = upsert_location(body)
        return {"status": "ok", "id": lid}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/api/criticality/locations/{loc_id:int}/delete")
@app.post("/infraportal/api/criticality/locations/{loc_id:int}/delete")
async def api_delete_location(request: Request, loc_id: int):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import delete_location
        ok = delete_location(loc_id)
        return {"status": "ok" if ok else "error"}
    except Exception as e:
        return {"status": "error", "error": str(e)}


# ── OPM Custom Group Names ───────────────────────────────────────────────────

@app.get("/api/criticality/geocode")
@app.get("/infraportal/api/criticality/geocode")
async def api_geocode(request: Request, q: str = ""):
    """Geocode a city/address string to lat/lng via Nominatim (OpenStreetMap). No API key needed."""
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    q = q.strip()
    if not q:
        return {"status": "error", "error": "q is required"}
    try:
        import requests as _req, urllib3 as _u3
        _u3.disable_warnings(_u3.exceptions.InsecureRequestWarning)
        r = _req.get(
            "https://nominatim.openstreetmap.org/search",
            params={"q": q, "format": "json", "limit": 3, "addressdetails": 1},
            headers={"User-Agent": "ZinniaInfraPortal/1.0 (markus.lanio@zinnia.com)"},
            timeout=8, verify=False,
        )
        r.raise_for_status()
        results = r.json()
        if not results:
            return {"status": "error", "error": f"No results found for \"{q}\""}
        hits = []
        for item in results[:3]:
            addr = item.get("address", {})
            label_parts = [
                addr.get("city") or addr.get("town") or addr.get("village") or addr.get("county") or "",
                addr.get("state") or addr.get("region") or "",
                addr.get("country") or "",
            ]
            label = ", ".join(p for p in label_parts if p) or item.get("display_name", "")
            hits.append({
                "lat":   float(item["lat"]),
                "lng":   float(item["lon"]),
                "label": label,
                "display_name": item.get("display_name", ""),
            })
        return {"status": "ok", "results": hits}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/api/criticality/opm-groups/add")
@app.post("/infraportal/api/criticality/opm-groups/add")
async def api_add_opm_group(request: Request):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import add_opm_group_name
        from routers.cache import cache as _cache
        body = await request.json()
        name = (body.get("name") or "").strip()
        if not name:
            return {"status": "error", "error": "name is required"}
        add_opm_group_name(name)
        _cache.invalidate("opm_named_groups")
        return {"status": "ok"}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/api/criticality/opm-groups/delete")
@app.post("/infraportal/api/criticality/opm-groups/delete")
async def api_delete_opm_group(request: Request):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import delete_opm_group_name
        from routers.cache import cache as _cache
        body = await request.json()
        name = (body.get("name") or "").strip()
        if not name:
            return {"status": "error", "error": "name is required"}
        delete_opm_group_name(name)
        _cache.invalidate("opm_named_groups")
        return {"status": "ok"}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.get("/api/criticality/noc-map")
@app.get("/infraportal/api/criticality/noc-map")
async def api_noc_map(request: Request):
    """
    Returns locations enriched with live alert status from OpManager groups.
    Used by the NOC world map.
    """
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import list_locations, list_group_dependencies
        from routers.opmanager import get_named_groups
        from routers.cache import cache as _cache

        # Build group status map from alarm-derived named groups
        named_groups = get_named_groups()
        group_status: dict = {g["name"]: g for g in named_groups}

        # Enrich locations with status
        locations = list_locations()
        enriched = []
        SEV_ORDER = {"critical": 1, "warning": 2, "minor": 3, "clear": 4}
        for loc in locations:
            loc_groups = loc.get("opm_groups") or []
            worst = "clear"
            alert_count = 0
            for gname in loc_groups:
                gs = group_status.get(gname, {})
                alert_count += gs.get("alert_count", 0)
                gs_status = gs.get("status", "clear")
                if SEV_ORDER.get(gs_status, 5) < SEV_ORDER.get(worst, 5):
                    worst = gs_status
            enriched.append({
                **loc,
                "status":      worst,
                "alert_count": alert_count,
                "group_statuses": {
                    gname: group_status.get(gname, {"status": "clear", "alert_count": 0})
                    for gname in loc_groups
                },
            })

        deps = list_group_dependencies()
        return {
            "status":       "ok",
            "locations":    enriched,
            "dependencies": deps,
            "all_groups":   named_groups,
        }
    except Exception as e:
        import traceback
        return {"status": "error", "error": str(e), "trace": traceback.format_exc()[-800:]}


@app.get("/api/noc/location-groups")
@app.get("/infraportal/api/noc/location-groups")
async def api_noc_location_groups(request: Request, name: str = ""):
    """Device groups assigned to a specific location, with live alarm overlay."""
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import get_conn as _db_conn
        from routers.cache import cache as _cache

        conn = _db_conn()
        try:
            loc_row = conn.execute(
                "SELECT id, name, lat, lng, description FROM topology_locations "
                "WHERE name=? COLLATE NOCASE", (name,)
            ).fetchone()
            if not loc_row:
                return {"status": "error", "error": f"Location '{name}' not found"}
            loc = dict(loc_row)

            grp_rows = conn.execute(
                "SELECT cg.id, cg.group_name, cg.opm_group_name, cg.environment, "
                "cg.owner_team, cg.default_tier, cg.service_description, "
                "(SELECT COUNT(*) FROM device_group_members dgm WHERE dgm.group_id = cg.id) AS device_count "
                "FROM criticality_groups cg "
                "WHERE cg.location_id = ?",
                (loc["id"],)
            ).fetchall()
            groups = [dict(r) for r in grp_rows]
        finally:
            conn.close()

        alarms, _ = _cache.get("opm_alarms")
        alarms = alarms or []
        SEV_NORM  = {"Critical": "Critical", "Service Down": "Critical",
                     "Trouble": "Major", "Attention": "Warning"}
        SEV_ORDER = {"Critical": 1, "Major": 2, "Warning": 3, "Clear": 99}

        grp_alarm_cnt: dict = {}
        grp_worst_sev: dict = {}
        for a in alarms:
            gn = (a.get("group_name") or "").strip()
            if not gn:
                continue
            grp_alarm_cnt[gn] = grp_alarm_cnt.get(gn, 0) + 1
            s = SEV_NORM.get(a.get("severity", ""), "Warning")
            if SEV_ORDER.get(s, 99) < SEV_ORDER.get(grp_worst_sev.get(gn, "Clear"), 99):
                grp_worst_sev[gn] = s

        enriched = []
        for g in groups:
            names = [g["group_name"]]
            if g.get("opm_group_name"):
                names.append(g["opm_group_name"])
            ac  = max((grp_alarm_cnt.get(n, 0) for n in names), default=0)
            sev = "Clear"
            for n in names:
                s = grp_worst_sev.get(n, "Clear")
                if SEV_ORDER.get(s, 99) < SEV_ORDER.get(sev, 99):
                    sev = s
            if ac == 0:
                sev = "Clear"
            enriched.append({
                "id":          g["id"],
                "name":        g["group_name"],
                "environment": g["environment"] or "Non-Prod",
                "owner_team":  g["owner_team"] or "",
                "tier":        g["default_tier"] or "",
                "description": g["service_description"] or "",
                "device_count": g["device_count"] or 0,
                "alarm_count": ac,
                "alarm_sev":   sev,
            })

        enriched.sort(key=lambda x: (SEV_ORDER.get(x["alarm_sev"], 99), x["name"]))
        return {"status": "ok", "location": loc, "groups": enriched}
    except Exception as e:
        import traceback
        return {"status": "error", "error": str(e), "trace": traceback.format_exc()[-800:]}


# ── NOC Operations Dashboard ─────────────────────────────────────────────────

@app.get("/noc")
@app.get("/infraportal/noc")
async def noc_page(request: Request):
    user, redirect = require_auth_check(request)
    if redirect: return redirect
    return templates.TemplateResponse(request, "noc.html", _ctx(request, user, "noc"))


@app.get("/api/noc/status")
@app.get("/infraportal/api/noc/status")
async def api_noc_status(request: Request):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.cache import cache as _cache
        from routers.database import get_criticality_map
        from collections import defaultdict

        crit_map = get_criticality_map()
        alarms, ts = _cache.get("opm_alarms")
        if not alarms:
            alarms = []

        def _loc(name: str) -> str:
            n = name.upper()
            if n.startswith("CAN"):  return "Candor"
            if n.startswith("VMCE"): return "VMC"
            if n.startswith("TOP"):  return "Topeka"
            return "Other"

        # Normalize OpManager severity labels to standard NOC labels
        SEV_NORM = {
            "Critical":     "Critical",
            "Trouble":      "Major",
            "Attention":    "Warning",
            "Service Down": "Major",
            "Clear":        "Clear",
        }
        sev_order = {"Critical": 1, "Major": 2, "Warning": 3, "Minor": 4, "Clear": 5}
        tier_order = {"P1": 0, "P2": 1, "P3": 2, "INFO": 3, None: 4}

        # Enrich and deduplicate by device_name — keep worst severity per device
        by_device: dict = {}
        for a in alarms:
            name = (a.get("device_name") or "").strip()
            if not name:
                continue
            norm_sev = SEV_NORM.get(a.get("severity", ""), a.get("severity", "Unknown"))
            crit = crit_map.get(name.upper(), {})
            entry = {
                **a,
                "severity":              norm_sev,
                "alarm_name":            a.get("message") or a.get("event_type") or a.get("category") or "",
                "location":              _loc(name),
                "criticality_tier":      crit.get("tier"),
                "criticality_team":      crit.get("owner_team"),
                "criticality_desc":      crit.get("service_description"),
                "criticality_blast":     crit.get("blast_radius"),
                "criticality_singleton": bool(crit.get("is_singleton")),
                "alarm_count":           1,
            }
            if name not in by_device:
                by_device[name] = entry
            else:
                existing = by_device[name]
                existing["alarm_count"] += 1
                # Keep worst severity
                if sev_order.get(norm_sev, 99) < sev_order.get(existing["severity"], 99):
                    by_device[name] = {**entry, "alarm_count": existing["alarm_count"]}

        enriched = sorted(
            by_device.values(),
            key=lambda x: (
                sev_order.get(x.get("severity", ""), 99),
                tier_order.get(x.get("criticality_tier"), 4),
            )
        )

        alarm_counts = {"Critical": 0, "Major": 0, "Warning": 0, "Minor": 0, "total": len(enriched)}
        for a in enriched:
            s = a.get("severity", "")
            if s in alarm_counts:
                alarm_counts[s] += 1

        locations: dict = {
            "Topeka": {"total": 0, "P1": 0, "P2": 0, "P3": 0, "INFO": 0, "unclassified": 0},
            "Candor": {"total": 0, "P1": 0, "P2": 0, "P3": 0, "INFO": 0, "unclassified": 0},
            "VMC":    {"total": 0, "P1": 0, "P2": 0, "P3": 0, "INFO": 0, "unclassified": 0},
            "Other":  {"total": 0, "P1": 0, "P2": 0, "P3": 0, "INFO": 0, "unclassified": 0},
        }
        for a in enriched:
            loc  = a.get("location", "Other")
            if loc not in locations: loc = "Other"
            locations[loc]["total"] += 1
            tier   = a.get("criticality_tier") or "unclassified"
            bucket = tier if tier in ("P1", "P2", "P3", "INFO") else "unclassified"
            locations[loc][bucket] += 1

        p1_alarms = [a for a in enriched if a.get("criticality_tier") == "P1"]

        # Cascade detection — only on unique devices (already deduplicated)
        critical_devs = [a for a in enriched if a.get("severity") in ("Critical", "Major")]

        by_loc: dict = defaultdict(list)
        for a in critical_devs:
            loc = a.get("location", "Other")
            if loc != "Other":
                by_loc[loc].append(a)

        cascades = []
        for loc, la in by_loc.items():
            if len(la) >= 3:
                teams = {a.get("criticality_team") for a in la if a.get("criticality_team")}
                has_network = "Network" in teams
                cascades.append({
                    "id":             f"loc_{loc.lower()}",
                    "type":           "location",
                    "label":          f"{loc} — {len(la)} devices alerting",
                    "location":       loc,
                    "device_count":   len(la),
                    "alarm_count":    sum(a.get("alarm_count", 1) for a in la),
                    "suspected_cause": "Network/upstream connectivity issue" if has_network
                                       else f"Infrastructure issue at {loc} site",
                    "escalate_to":    "Network" if has_network
                                      else ("multiple teams" if len(teams) > 1
                                            else (next(iter(teams)) if teams else "Unknown")),
                    "confidence":     "high" if has_network else "medium",
                    "devices":        [a["device_name"] for a in la[:10]],
                })

        by_team: dict = defaultdict(list)
        for a in critical_devs:
            t = a.get("criticality_team")
            if t:
                by_team[t].append(a)

        loc_cascade_devs = {d for c in cascades if c["type"] == "location" for d in c["devices"]}
        for team, ta in by_team.items():
            non_loc = [a for a in ta if a["device_name"] not in loc_cascade_devs]
            if len(non_loc) >= 4:
                cascades.append({
                    "id":             f"team_{team.lower().replace(' ', '_').replace('/', '_')}",
                    "type":           "team",
                    "label":          f"{team} — {len(ta)} devices alerting",
                    "team":           team,
                    "device_count":   len(ta),
                    "alarm_count":    sum(a.get("alarm_count", 1) for a in ta),
                    "suspected_cause": f"Widespread {team} infrastructure issue",
                    "escalate_to":    team,
                    "confidence":     "medium",
                    "devices":        [a["device_name"] for a in ta[:10]],
                })

        return {
            "status":       "ok",
            "alarm_counts": alarm_counts,
            "locations":    locations,
            "cascades":     cascades,
            "p1_alarms":    p1_alarms[:20],
            "alarms":       list(enriched[:200]),
            "timestamp":    _cache.age_string(ts) if ts else "unknown",
        }
    except Exception as e:
        import traceback
        return {"status": "error", "error": str(e), "trace": traceback.format_exc()}


_noc_triage_state: dict = {"status": "idle"}


@app.post("/api/noc/triage")
@app.post("/infraportal/api/noc/triage")
async def api_noc_triage_start(request: Request):
    global _noc_triage_state
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    if _noc_triage_state.get("status") == "running":
        return {"status": "running"}
    _noc_triage_state = {"status": "running"}

    import threading
    def _run():
        global _noc_triage_state
        try:
            from routers.cache import cache as _cache
            from routers.database import get_criticality_map
            from routers.criticality import ai_noc_triage
            from collections import defaultdict

            SEV_NORM = {"Critical": "Critical", "Trouble": "Major",
                        "Attention": "Warning", "Service Down": "Major", "Clear": "Clear"}
            sev_order = {"Critical": 1, "Major": 2, "Warning": 3, "Minor": 4, "Clear": 5}

            crit_map = get_criticality_map()
            alarms, _ = _cache.get("opm_alarms")
            if not alarms:
                alarms = []

            def _loc(name: str) -> str:
                n = name.upper()
                if n.startswith("CAN"):  return "Candor"
                if n.startswith("VMCE"): return "VMC"
                if n.startswith("TOP"):  return "Topeka"
                return "Other"

            # Deduplicate by device — keep worst severity
            by_device: dict = {}
            for a in alarms:
                name = (a.get("device_name") or "").strip()
                if not name:
                    continue
                norm_sev = SEV_NORM.get(a.get("severity", ""), a.get("severity", "Unknown"))
                crit = crit_map.get(name.upper(), {})
                entry = {
                    **a,
                    "severity":         norm_sev,
                    "alarm_name":       a.get("message") or a.get("event_type") or "",
                    "location":         _loc(name),
                    "criticality_tier": crit.get("tier"),
                    "criticality_team": crit.get("owner_team"),
                    "criticality_desc": crit.get("service_description"),
                }
                if name not in by_device or sev_order.get(norm_sev, 99) < sev_order.get(by_device[name]["severity"], 99):
                    by_device[name] = entry

            enriched = list(by_device.values())
            critical_devs = [a for a in enriched if a.get("severity") in ("Critical", "Major")]

            by_loc: dict = defaultdict(list)
            for a in critical_devs:
                loc = a.get("location", "Other")
                if loc != "Other":
                    by_loc[loc].append(a)
            cascades = []
            for loc, la in by_loc.items():
                if len(la) >= 3:
                    cascades.append({"location": loc, "device_count": len(la),
                                     "devices": [a["device_name"] for a in la[:10]]})

            result = ai_noc_triage(enriched, crit_map, cascades)
            _noc_triage_state = {"status": "ok", "result": result}
        except Exception as e:
            _noc_triage_state = {"status": "error", "error": str(e)}

    threading.Thread(target=_run, daemon=True).start()
    return {"status": "running"}


@app.get("/api/noc/triage/status")
@app.get("/infraportal/api/noc/triage/status")
async def api_noc_triage_status(request: Request):
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    return _noc_triage_state


@app.get("/api/noc/topology")
@app.get("/infraportal/api/noc/topology")
async def api_noc_topology(request: Request):
    """Group topology for the NOC root-cause page and the Dependency Editor.
    Nodes are Device Groups ONLY (criticality_groups) — no raw OpManager groups.
    OpManager is the source of truth for *populating* Device Groups (via Sync),
    not for deciding what shows up here; mixing in unregistered OPM names is what
    produced duplicate/renamed-group confusion and phantom "Server"/"Switch"-style
    entries earlier. Edges are dependencies between two registered Device Groups."""
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import get_conn as _db_conn
        import json as _json

        conn = _db_conn()
        try:
            grp_rows = conn.execute(
                "SELECT cg.id, cg.group_name, cg.environment, cg.owner_team, "
                "cg.default_tier, cg.service_description, cg.group_category, tl.name AS location_name "
                "FROM criticality_groups cg "
                "LEFT JOIN topology_locations tl ON cg.location_id = tl.id"
            ).fetchall()
            grp_dicts: list = [dict(r) for r in grp_rows]
            crit_names: set = {d["group_name"] for d in grp_dicts}

            dep_rows = conn.execute(
                "SELECT id, from_group, to_group, dep_type, confidence, notes FROM group_dependencies"
            ).fetchall()
            # Only edges between two registered Device Groups — anything else is a
            # leftover from before Device-Groups-only (e.g. AI Suggest once ran
            # against raw OpManager device-type buckets like "Server"/"Switch").
            edges = [dict(r) for r in dep_rows
                     if r["from_group"] in crit_names and r["to_group"] in crit_names]

            loc_rows = conn.execute(
                "SELECT name, lat, lng, description, opm_groups FROM topology_locations"
            ).fetchall()
            locations = [{**dict(lr), "opm_groups": _json.loads(lr["opm_groups"] or "[]")} for lr in loc_rows]

            # Severity from the SAME source the device panel uses: device_group_members
            # + live alarms, so the node and the panel always agree.
            from routers.cache import cache as _alarm_cache
            live_alarms, _ = _alarm_cache.get("opm_alarms")
            live_alarms = live_alarms or []
            SEV_RANK = {"Critical": 1, "Service Down": 2, "Trouble": 3, "Attention": 4, "Clear": 99}
            device_sev: dict = {}
            for a in live_alarms:
                dn = (a.get("device_name") or "").strip().upper()
                if not dn:
                    continue
                sev = a.get("severity", "Clear")
                cur = device_sev.get(dn)
                if not cur or SEV_RANK.get(sev, 50) < SEV_RANK.get(cur, 99):
                    device_sev[dn] = sev

            member_rows = conn.execute("SELECT group_id, device_name FROM device_group_members").fetchall()
            group_alert_count: dict = {}   # criticality_groups.id -> count
            group_worst_sev_id: dict = {}  # criticality_groups.id -> raw severity label
            for r in member_rows:
                gid = r["group_id"]
                dn  = (r["device_name"] or "").strip().upper()
                sev = device_sev.get(dn, "Clear")
                if sev != "Clear":
                    group_alert_count[gid] = group_alert_count.get(gid, 0) + 1
                    cur = group_worst_sev_id.get(gid, "Clear")
                    if SEV_RANK.get(sev, 50) < SEV_RANK.get(cur, 99):
                        group_worst_sev_id[gid] = sev
        finally:
            conn.close()

        SEV_NORM = {"Critical": "Critical", "Service Down": "Critical",
                    "Trouble": "Major", "Attention": "Warning"}
        nodes = []
        for d in grp_dicts:
            alm_c = group_alert_count.get(d["id"], 0)
            raw_sv = group_worst_sev_id.get(d["id"], "Clear")
            norm_sv = SEV_NORM.get(raw_sv, raw_sv)
            nodes.append({
                "id":             d["group_name"],
                "name":           d["group_name"],
                "group_id":       d["id"],
                "environment":    d.get("environment") or "Non-Prod",
                "owner_team":     d.get("owner_team") or "",
                "tier":           d.get("default_tier") or "",
                "description":    d.get("service_description") or "",
                "group_category": d.get("group_category") or "App",
                "location":       d.get("location_name") or "",
                "alert_count":    alm_c,
                "alert_sev":      norm_sv if alm_c > 0 else "Clear",
            })

        return {
            "status":    "ok",
            "nodes":     nodes,
            "edges":     edges,
            "locations": locations,
        }
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.get("/api/noc/group-devices")
@app.get("/infraportal/api/noc/group-devices")
async def api_noc_group_devices(request: Request, name: str = "", group_id: int = 0):
    """Device-level nodes for a group with live alarm overlay — used by NOC drill-down."""
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.database import get_conn as _db_conn
        from routers.cache import cache as _cache

        conn = _db_conn()
        try:
            grp = None
            if group_id:
                grp = conn.execute(
                    "SELECT id FROM criticality_groups WHERE id=?", (group_id,)
                ).fetchone()
            if not grp and name:
                # Try exact group_name match first, then opm_group_name
                grp = conn.execute(
                    "SELECT id FROM criticality_groups WHERE group_name=? COLLATE NOCASE",
                    (name,)
                ).fetchone()
            if not grp and name:
                grp = conn.execute(
                    "SELECT id FROM criticality_groups WHERE opm_group_name=? COLLATE NOCASE",
                    (name,)
                ).fetchone()
            if not grp:
                return {"status": "ok", "devices": [], "group_name": name,
                        "group_found": False}
            members = conn.execute(
                "SELECT m.device_name, r.tier, r.owner_team, r.service_description "
                "FROM device_group_members m "
                "LEFT JOIN asset_criticality r ON r.device_name=m.device_name COLLATE NOCASE "
                "WHERE m.group_id=?",
                (grp["id"],)
            ).fetchall()
        finally:
            conn.close()

        alarms, _ = _cache.get("opm_alarms")
        alarms = alarms or []
        SEV_NORM = {"Critical": "Critical", "Service Down": "Critical",
                    "Trouble": "Major", "Attention": "Warning"}
        SEV_ORD  = {"Critical": 1, "Major": 2, "Warning": 3}
        alarm_map: dict = {}
        for a in alarms:
            dn = (a.get("device_name") or "").strip().upper()
            if not dn:
                continue
            norm  = SEV_NORM.get(a.get("severity", ""), a.get("severity", "Clear"))
            entry = alarm_map.setdefault(dn, {"sev": "Clear", "text": "", "ids": []})
            aid = a.get("alarm_id")
            if aid is not None:
                entry["ids"].append(aid)
            if SEV_ORD.get(norm, 99) < SEV_ORD.get(entry["sev"], 99):
                entry["sev"]  = norm
                entry["text"] = a.get("message") or a.get("alarmMessage") or a.get("event_type") or ""

        devices = []
        for m in members:
            dn  = m["device_name"] or ""
            alm = alarm_map.get(dn.upper(), {})
            devices.append({
                "name":        dn,
                "tier":        m["tier"] or "",
                "team":        m["owner_team"] or "",
                "description": m["service_description"] or "",
                "alarm_sev":   alm.get("sev", "Clear"),
                "alarm_text":  alm.get("text", ""),
                "alarm_ids":   alm.get("ids", []),
            })
        SEV_SORT = {"Critical": 0, "Major": 1, "Warning": 2, "Clear": 3}
        devices.sort(key=lambda d: (SEV_SORT.get(d["alarm_sev"], 3), d["name"]))

        return {"status": "ok", "devices": devices, "group_name": name,
                "group_found": True, "member_count": len(devices)}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.get("/api/criticality/inventory/merged")
@app.get("/infraportal/api/criticality/inventory/merged")
async def api_criticality_inventory_merged(request: Request, refresh: str = ""):
    """Merged inventory from OPM + Lansweeper + VMware + AD.
    Includes OU path and environment hints for classification."""
    user, _ = require_auth_check(request)
    if not user:
        return {"status": "unauthorized"}
    try:
        from routers.cache import cache as _cache
        import json as _json

        force = refresh == "1"
        devices: dict = {}  # upper(name) → record

        def _merge(name, **kwargs):
            key = name.upper()
            if key not in devices:
                devices[key] = {"name": name, "sources": [], **kwargs}
            else:
                devices[key]["sources"] = list(set(devices[key]["sources"] + kwargs.get("sources", [])))
                for k, v in kwargs.items():
                    if k != "sources" and v and not devices[key].get(k):
                        devices[key][k] = v

        # ── OpManager ─────────────────────────────────────────────────────────
        try:
            opm_devices, _ = _cache.get("opm_devices")
            if opm_devices:
                for d in opm_devices:
                    nm = (d.get("display_name") or d.get("name") or "").strip()
                    if not nm:
                        continue
                    _merge(nm, sources=["OpManager"], ip=d.get("ip_address", ""),
                           os="", type=d.get("category", ""), environment="",
                           group=d.get("group_name", ""), ou_path="", last_seen="")
        except Exception:
            pass

        # ── Lansweeper ────────────────────────────────────────────────────────
        try:
            ls_assets, _ = _cache.get("lansweeper_assets")
            if ls_assets:
                for a in ls_assets:
                    nm = (a.get("name") or "").strip()
                    if not nm:
                        continue
                    _merge(nm, sources=["Lansweeper"], ip=a.get("ip", ""),
                           os=a.get("os") or a.get("os_name", ""),
                           type=a.get("type", ""),
                           environment="",
                           group="", ou_path="",
                           last_seen=a.get("last_seen", ""))
        except Exception:
            pass

        # ── VMware ────────────────────────────────────────────────────────────
        try:
            detailed, _ = _cache.get("detailed_vms")
            if detailed:
                for vm in (detailed.get("vms") or []):
                    nm = (vm.get("name") or "").strip()
                    if not nm:
                        continue
                    _merge(nm, sources=["VMware"], ip=vm.get("ip_address", ""),
                           os=vm.get("os_name", ""), type="VM",
                           environment=vm.get("environment", ""),
                           group="", ou_path="", last_seen="")
        except Exception:
            pass

        # ── Active Directory (cache-only — never block on live LDAP here) ────────
        # The scheduler job job_ad_reports warms ad_all_computers_ou every 6 h.
        try:
            ad_computers, _ = _cache.get("ad_all_computers_ou")
            if ad_computers:
                for c in ad_computers:
                    nm = c.get("name", "").strip()
                    if not nm:
                        continue
                    _merge(nm, sources=["AD"],
                           ip="",
                           os=c.get("os", ""),
                           type="Computer",
                           environment=c.get("environment", "Unknown"),
                           group="",
                           ou_path=c.get("ou_path", ""),
                           last_seen=c.get("last_login") or "")
        except Exception:
            pass

        result = sorted(devices.values(), key=lambda x: x["name"].lower())
        return {"status": "ok", "devices": result, "count": len(result)}
    except Exception as e:
        return {"status": "error", "error": str(e)}


# ── VDI Cost Report API ──────────────────────────────────────────────────────
# Fire-and-forget pattern (same as api_analyze) to avoid IIS 504 timeouts.
# The report makes hundreds of sequential Graph API calls and can take 2-5 min.

_vdi_cost_state: dict = {"status": "idle", "result": None, "error": None, "cost": 35.0}

def _run_vdi_cost_background(cost_per_machine: float, force_refresh: bool) -> None:
    global _vdi_cost_state
    try:
        data = get_vdi_cost_report(force_refresh=force_refresh, cost_per_machine=cost_per_machine)
        _vdi_cost_state["result"] = data
        _vdi_cost_state["status"] = "done"
        _vdi_cost_state["cost"]   = cost_per_machine
    except Exception as exc:
        import logging
        logging.getLogger(__name__).exception("vdi_cost background error")
        _vdi_cost_state["status"] = "error"
        _vdi_cost_state["error"]  = str(exc)


@app.get("/api/vdi-cost/report")
@app.get("/infraportal/api/vdi-cost/report")
async def api_vdi_cost_report(cost_per_machine: float = 35.0, refresh: str = ""):
    """
    Start or serve the VDI cost report.
    Cache hit: returns data immediately.
    Cache miss / refresh=1: fires background thread, returns {status:"running"}.
    Poll /api/vdi-cost/report/status for completion.
    """
    global _vdi_cost_state
    force = (refresh == "1")
    if not force:
        try:
            from routers.cache import cache
            cached, _ = cache.get(f"vdi_cost_report_{int(cost_per_machine * 100)}")
            if cached:
                return cached
        except Exception:
            pass
    if _vdi_cost_state["status"] != "running":
        _vdi_cost_state.update({"status": "running", "result": None, "error": None, "cost": cost_per_machine})
        threading.Thread(target=_run_vdi_cost_background, args=(cost_per_machine, force), daemon=True).start()
    return {"status": "running"}


@app.get("/api/vdi-cost/report/status")
@app.get("/infraportal/api/vdi-cost/report/status")
async def api_vdi_cost_report_status(cost_per_machine: float = 35.0):
    """Poll for VDI cost report completion."""
    global _vdi_cost_state
    if _vdi_cost_state["status"] == "running":
        return {"status": "running"}
    if _vdi_cost_state["status"] == "error":
        return {"status": "error", "message": _vdi_cost_state["error"]}
    if _vdi_cost_state["status"] == "done" and _vdi_cost_state["result"]:
        result = _vdi_cost_state["result"]
        _vdi_cost_state["status"] = "idle"
        return result
    try:
        from routers.cache import cache
        cached, _ = cache.get(f"vdi_cost_report_{int(cost_per_machine * 100)}")
        if cached:
            return cached
    except Exception:
        pass
    return {"status": "idle"}


# ── Server Share Audit ─────────────────────────────────────────────────────────

@app.get("/api/ad/share-audit/servers")
@app.get("/infraportal/api/ad/share-audit/servers")
async def api_share_audit_servers(refresh: bool = False):
    try:
        from routers.share_audit import get_servers
        servers, ts = get_servers(force_refresh=refresh)
        env_counts: dict = {}
        for s in servers:
            env_counts[s['env']] = env_counts.get(s['env'], 0) + 1
        return {"status": "ok", "servers": servers, "total": len(servers),
                "env_counts": env_counts, "timestamp": ts}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.post("/api/ad/share-audit/scan")
@app.post("/infraportal/api/ad/share-audit/scan")
async def api_share_audit_scan(request: Request):
    try:
        body        = await request.json()
        env_filter  = body.get("env_filter") or None
        sample_size = int(body["sample_size"]) if body.get("sample_size") else None
        from routers.share_audit import start_scan, get_scan_state
        started = start_scan(env_filter=env_filter, sample_size=sample_size)
        if not started:
            return {"status": "already_running", "state": get_scan_state()}
        return {"status": "started"}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.get("/api/ad/share-audit/status")
@app.get("/infraportal/api/ad/share-audit/status")
async def api_share_audit_status():
    try:
        from routers.share_audit import get_scan_state
        return {"status": "ok", "state": get_scan_state()}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.get("/api/ad/share-audit/results")
@app.get("/infraportal/api/ad/share-audit/results")
async def api_share_audit_results(env: str = ""):
    try:
        from routers.share_audit import get_results
        data = get_results()
        if not data:
            return {"status": "no_data"}
        servers = data.get("servers", [])
        results = data.get("results", {})
        if env and env != "all":
            servers = [s for s in servers if s.get("env") == env]
        return {
            "status":       "ok",
            "servers":      servers,
            "results":      results,
            "completed_at": data.get("completed_at"),
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.get("/api/ad/share-audit/pdf")
@app.get("/infraportal/api/ad/share-audit/pdf")
async def api_share_audit_pdf(env: str = "production"):
    import re
    from datetime import datetime as _dt
    from fastapi.responses import Response as FastResponse
    try:
        from routers.share_audit import generate_pdf
        pdf_bytes = generate_pdf(env_filter=env or "production")
        env_part  = re.sub(r'[^a-z0-9]', '_', (env or 'production').lower())
        filename  = f"ShareAudit_{env_part}_{_dt.now().strftime('%Y%m%d')}.pdf"
        return FastResponse(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except ImportError as e:
        return {"status": "error", "error": str(e), "install_hint": "pip install xhtml2pdf"}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.get("/api/ad/share-audit/excel")
@app.get("/infraportal/api/ad/share-audit/excel")
async def api_share_audit_excel(env: str = "all"):
    import re
    from datetime import datetime as _dt
    from fastapi.responses import Response as FastResponse
    try:
        from routers.share_audit import generate_excel
        xlsx_bytes = generate_excel(env_filter=env or "all")
        env_part   = re.sub(r'[^a-z0-9]', '_', (env or 'all').lower())
        filename   = f"ShareAudit_{env_part}_{_dt.now().strftime('%Y%m%d')}.xlsx"
        return FastResponse(
            content=xlsx_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    except ImportError as e:
        return {"status": "error", "error": str(e), "install_hint": "pip install openpyxl"}
    except Exception as e:
        return {"status": "error", "error": str(e)}


@app.post("/api/ad/share-audit/set-env")
@app.post("/infraportal/api/ad/share-audit/set-env")
async def api_share_audit_set_env(request: Request):
    user, _ = require_auth_check(request, "admin")
    if not user:
        return {"status": "unauthorized"}
    try:
        body        = await request.json()
        server_name = (body.get("server_name") or "").strip()
        environment = (body.get("environment") or "").strip()
        valid_envs  = ("production", "qa", "test", "dev", "other")
        if not server_name or environment not in valid_envs:
            return {"status": "error", "message": "Invalid server_name or environment"}
        from routers.database import set_share_audit_env_override
        from routers.cache import cache
        ok = set_share_audit_env_override(server_name, environment,
                                           set_by=(user or {}).get("email"))
        if ok:
            cache.invalidate("share_audit_servers")
        return {"status": "ok" if ok else "error"}
    except Exception as e:
        return {"status": "error", "message": str(e)}