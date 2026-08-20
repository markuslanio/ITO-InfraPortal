"""
Server Share Audit — enumerate SMB shares and permissions on AD-joined servers.
Used for access reviews and compliance audits.

Connection strategy: pywin32 Win32 API (in-process, no subprocess).
  1. WNetAddConnection2  — authenticates as AD service account over SMB port 445
  2. NetShareEnum(lvl 2) — enumerates shares via Windows API
  3. GetNamedSecurityInfo — reads NTFS DACL via UNC path
  4. WNetCancelConnection2 — cleanup
No WinRM (5985) or DCOM/RPC (135) required. Works from IIS worker processes.
"""
import io
import logging
import os
import threading
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime

logger = logging.getLogger(__name__)

# ── scan state ────────────────────────────────────────────────────────────────

_scan_state: dict = {
    "running":      False,
    "progress":     0,
    "total":        0,
    "servers":      [],
    "results":      {},
    "started_at":   None,
    "completed_at": None,
    "error":        None,
}
_scan_lock = threading.Lock()

# ── environment classification ────────────────────────────────────────────────

_ENV_SUFFIX = {"P": "production", "Q": "qa", "T": "test", "D": "dev"}
ENV_LABELS  = {
    "production": "Production",
    "qa":         "QA",
    "test":       "Test",
    "dev":        "Dev",
    "other":      "Other",
}
ENV_COLORS = {
    "production": ("#ffe0e0", "#cc0000"),
    "qa":         ("#fff3cc", "#996600"),
    "test":       ("#e0f0ff", "#0066cc"),
    "dev":        ("#e0ffe0", "#006600"),
    "other":      ("#f0f0f0", "#666666"),
}
_ENV_SORT = {"production": 0, "qa": 1, "test": 2, "dev": 3, "other": 4}


def _classify_env(name: str) -> str:
    n = (name or "").upper().rstrip("$")
    return _ENV_SUFFIX.get(n[-1] if n else "", "other")


# ── Win32 API scanner ─────────────────────────────────────────────────────────
# Uses pywin32 (win32net / win32wnet / win32security) — runs in-process,
# no PowerShell subprocess, works from IIS worker processes.
#
#   WNetAddConnection2    — authenticate as AD service account (SMB port 445)
#   NetShareEnum (lvl 2)  — enumerate shares
#   GetNamedSecurityInfo  — read NTFS DACL via UNC path
#   WNetCancelConnection2 — cleanup

_STYPE_IPC  = 0x00000003
_STYPE_MASK = 0x000000FF   # low byte = share type


def _mask_to_str(mask: int) -> str:
    if (mask & 0x1F01FF) == 0x1F01FF:    return "FullControl"
    if (mask & 0x0301BF) == 0x0301BF:    return "Modify"
    if (mask & 0x0201A9) == 0x0201A9:    return "ReadAndExecute"
    if (mask & 0x00120089) == 0x00120089: return "Read"
    if (mask & 0x00100116) == 0x00100116: return "Write"
    return f"Special(0x{mask:08X})"


def _scan_server(server_name: str, timeout: int = 60) -> dict:
    """
    Enumerate SMB shares and NTFS ACLs via Win32 API (pywin32).
    In-process — no subprocess, no PowerShell startup, IIS-compatible.
    Returns {"shares": [...]} or {"error": "..."}.
    """
    try:
        import win32net
        import win32wnet
        import win32security
        import pywintypes
    except ImportError:
        return {"error": "pywin32 not installed — run: pip install pywin32"}

    domain   = os.getenv("AD_DOMAIN", "")
    username = os.getenv("AD_USER",   "")
    password = os.getenv("AD_PASSWORD", "")
    ipc_path = f"\\\\{server_name}\\IPC$"
    connected = False

    try:
        # Step 1: authenticate via WNetAddConnection2 (SMB port 445)
        if username and password:
            user_str = f"{domain}\\{username}" if domain else username
            try:
                nr = win32wnet.NETRESOURCE()
                nr.lpRemoteName = ipc_path
                nr.dwType = 0
                win32wnet.WNetAddConnection2(nr, password, user_str, 0)
                connected = True
            except pywintypes.error:
                pass  # continue with machine account

        # Step 2: enumerate shares via NetShareEnum level 2 (includes path/remark)
        try:
            shares_raw, _, _ = win32net.NetShareEnum(server_name, 2)
        except pywintypes.error as e:
            return {"error": f"NetShareEnum: {e.strerror} (err {e.winerror})"}

        results = []
        for s in shares_raw:
            stype = s.get('type', 0)
            if (stype & _STYPE_MASK) == _STYPE_IPC:
                continue
            sname = s['netname']
            if sname == 'IPC$':
                continue

            spath = s.get('path', '')
            sdesc = s.get('remark', '')
            unc   = f"\\\\{server_name}\\{sname}"

            # Step 3: read NTFS DACL via GetNamedSecurityInfo over UNC (SMB port 445)
            ntfs = []
            try:
                sd = win32security.GetNamedSecurityInfo(
                    unc,
                    win32security.SE_FILE_OBJECT,
                    win32security.DACL_SECURITY_INFORMATION,
                )
                dacl = sd.GetSecurityDescriptorDacl()
                if dacl:
                    for i in range(dacl.GetAceCount()):
                        (atype, aflags), mask, sid = dacl.GetAce(i)
                        try:
                            name, dom, _ = win32security.LookupAccountSid(None, sid)
                            identity = f"{dom}\\{name}" if dom else name
                        except Exception:
                            identity = str(win32security.ConvertSidToStringSid(sid))
                        ntfs.append({
                            'identity':  identity,
                            'rights':    _mask_to_str(mask),
                            'type':      'Allow' if atype == 0 else 'Deny',
                            'inherited': bool(aflags & 0x10),
                        })
            except Exception:
                pass  # ACL unreadable — include share with empty ACL list

            results.append({
                'name':        sname,
                'path':        spath,
                'description': sdesc,
                'share_acl':   [],
                'ntfs_acl':    ntfs,
            })

        return {"shares": results}

    except Exception as e:
        return {"error": str(e)[:300]}

    finally:
        # Step 4: release SMB connection
        if connected:
            try:
                win32wnet.WNetCancelConnection2(ipc_path, 0, False)
            except Exception:
                pass


# ── server list from AD ───────────────────────────────────────────────────────

def _fetch_servers_from_ad() -> list:
    try:
        from routers.active_directory import (
            get_conn, paged_search, attr, get_ft_from_entry,
            _is_master_image_by_name,
        )
        conn = get_conn()
        entries = paged_search(
            conn,
            '(&(objectClass=computer)(operatingSystem=*Server*))',
            attributes=['cn', 'operatingSystem', 'lastLogonTimestamp'],
        )
        servers = []
        for e in entries:
            if e.get('type') != 'searchResEntry':
                continue
            name = attr(e, 'cn')
            if not name or _is_master_image_by_name(name):
                continue
            # Skip cloud-hosted AD-joined objects not reachable from on-prem
            nl = name.lower()
            if nl.startswith('amznfsx') or nl.startswith('aws') or nl.startswith('ec2amaz'):
                continue
            dt = get_ft_from_entry(e, 'lastLogonTimestamp')
            servers.append({
                'name':       name,
                'os':         attr(e, 'operatingSystem'),
                'last_login': dt.strftime('%Y-%m-%d') if dt else 'Never',
                'env':        _classify_env(name),
            })
        conn.unbind()
        return sorted(servers, key=lambda s: s['name'].lower())
    except Exception as e:
        logger.error("share_audit: AD fetch failed: %s", e)
        return []


def _apply_overrides(servers: list) -> list:
    try:
        from routers.database import get_share_audit_env_overrides
        overrides = {k.upper(): v for k, v in get_share_audit_env_overrides().items()}
    except Exception:
        overrides = {}
    result = []
    for s in servers:
        s = dict(s)
        ov = overrides.get(s['name'].upper())
        if ov:
            s['env'] = ov
            s['env_override'] = True
        result.append(s)
    return result


def get_servers(force_refresh: bool = False):
    """Return (servers_list, age_string) from cache or AD."""
    from routers.cache import cache
    if not force_refresh:
        data, ts = cache.get("share_audit_servers")
        if data is not None:
            return _apply_overrides(data), cache.age_string(ts)
    servers = _fetch_servers_from_ad()
    cache.set("share_audit_servers", servers)
    _, ts = cache.get("share_audit_servers")
    return _apply_overrides(servers), cache.age_string(ts)


# ── scan orchestration ────────────────────────────────────────────────────────

def get_scan_state() -> dict:
    with _scan_lock:
        return {
            "running":      _scan_state["running"],
            "progress":     _scan_state["progress"],
            "total":        _scan_state["total"],
            "started_at":   _scan_state["started_at"],
            "completed_at": _scan_state["completed_at"],
            "error":        _scan_state["error"],
        }


def start_scan(env_filter: str = None, sample_size: int = None) -> bool:
    """Fire-and-forget scan. Returns False if already running."""
    with _scan_lock:
        if _scan_state["running"]:
            return False
        _scan_state.update({
            "running":      True,
            "progress":     0,
            "total":        0,
            "results":      {},
            "servers":      [],
            "error":        None,
            "started_at":   datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "completed_at": None,
        })
    threading.Thread(target=_scan_bg, args=(env_filter, sample_size), daemon=True).start()
    return True


def _scan_bg(env_filter: str, sample_size: int = None):
    try:
        # Use cached list if available — the /servers endpoint warms it on page load.
        # Force refresh only when the cache is cold (first scan after restart).
        from routers.cache import cache as _cache
        cached, _ = _cache.get("share_audit_servers")
        servers, _ = get_servers(force_refresh=(cached is None))
        if env_filter and env_filter != "all":
            servers = [s for s in servers if s['env'] == env_filter]
        if sample_size and sample_size > 0:
            servers = servers[:sample_size]

        with _scan_lock:
            _scan_state["servers"] = servers
            _scan_state["total"]   = len(servers)

        logger.info("share_audit: scanning %d servers (env=%s, sample=%s)",
                    len(servers), env_filter, sample_size)

        def _do_one(srv):
            res = _scan_server(srv['name'])
            if "error" in res:
                logger.warning("share_audit: %s — %s", srv['name'], res["error"])
            with _scan_lock:
                _scan_state["results"][srv['name'].upper()] = res
                _scan_state["progress"] += 1

        with ThreadPoolExecutor(max_workers=5) as pool:
            list(pool.map(_do_one, servers))

        with _scan_lock:
            _scan_state["running"]      = False
            _scan_state["completed_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        from routers.cache import cache
        cache.set("share_audit_results", {
            "servers":      _scan_state["servers"],
            "results":      dict(_scan_state["results"]),
            "completed_at": _scan_state["completed_at"],
            "env_filter":   env_filter,
        })
        n_err = sum(1 for r in _scan_state["results"].values() if "error" in r)
        logger.info("share_audit: done — %d servers, %d errors", len(servers), n_err)

    except Exception as e:
        logger.error("share_audit: bg error: %s", e, exc_info=True)
        with _scan_lock:
            _scan_state["running"] = False
            _scan_state["error"]   = str(e)


def get_results() -> dict | None:
    """Return last scan results (live state or cache)."""
    with _scan_lock:
        if _scan_state["results"]:
            return {
                "servers":      list(_scan_state["servers"]),
                "results":      dict(_scan_state["results"]),
                "completed_at": _scan_state["completed_at"],
            }
    from routers.cache import cache
    data, _ = cache.get("share_audit_results")
    return data


# ── Excel report ─────────────────────────────────────────────────────────────

def _esc(s) -> str:
    return str(s or '').replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


def generate_html(env_filter: str = "production") -> str:
    data         = get_results() or {"servers": [], "results": {}, "completed_at": ""}
    all_servers  = data.get("servers",  [])
    results      = data.get("results",  {})
    completed_at = data.get("completed_at", "")

    if env_filter and env_filter != "all":
        servers = [s for s in all_servers if s.get("env") == env_filter]
    else:
        servers = list(all_servers)

    servers = sorted(servers,
                     key=lambda s: (_ENV_SORT.get(s.get("env", "other"), 99),
                                    s["name"].lower()))

    env_label = (ENV_LABELS.get(env_filter, env_filter.title())
                 if env_filter and env_filter != "all"
                 else "All Environments")
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    n_ok  = sum(1 for s in servers
                if s["name"].upper() in results
                and "error" not in results[s["name"].upper()])
    n_err = sum(1 for s in servers
                if s["name"].upper() in results
                and "error" in results[s["name"].upper()])

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  body {{ font-family: Arial, sans-serif; font-size: 10pt; color: #222; padding: 20px; }}
  h1   {{ font-size: 17pt; color: #003366; margin: 0 0 4px; }}
  .meta {{ font-size: 9pt; color: #555; margin-bottom: 18px; }}
  .summary-box {{ background:#f0f4ff; border:1px solid #c0d0f0; padding:8px 14px;
                  margin-bottom:18px; border-radius:4px; font-size:9pt; }}
  .summary-box td {{ padding:2px 14px 2px 0; }}
  .server-block  {{ margin-bottom:18px; page-break-inside:avoid; }}
  .srv-header    {{ background:#003366; color:white; padding:5px 10px;
                    font-size:10pt; font-weight:bold; border-radius:3px 3px 0 0; }}
  .srv-body      {{ border:1px solid #cce; border-top:none; padding:8px 12px; background:#fff; }}
  .env-badge     {{ display:inline; padding:1px 7px; border-radius:3px;
                    font-size:8pt; font-weight:bold; margin-left:8px; }}
  .share-block   {{ margin:8px 0 10px 10px; border-left:3px solid #aaccee; padding-left:10px; }}
  .share-name    {{ font-size:10.5pt; font-weight:bold; color:#003366; margin-bottom:2px; }}
  .share-meta    {{ font-size:8pt; color:#666; margin-bottom:6px; }}
  .sec-label     {{ font-size:8pt; font-weight:bold; color:#003366; margin:6px 0 2px; }}
  table {{ width:100%; border-collapse:collapse; margin-bottom:4px; font-size:8.5pt; }}
  th    {{ background:#003366; color:white; padding:3px 8px; text-align:left; font-size:8pt; }}
  td    {{ padding:3px 8px; border-bottom:1px solid #eee; }}
  tr:nth-child(even) {{ background:#f8f8f8; }}
  .inh  {{ color:#aaa; }}
  .error-box {{ background:#fff0f0; border:1px solid #ffcccc; padding:5px 10px;
                color:#cc0000; font-size:9pt; border-radius:3px; }}
  .no-data   {{ color:#888; font-style:italic; font-size:9pt; }}
</style>
</head><body>
<h1>Server Share Audit Report</h1>
<div class="meta">Zinnia Infrastructure Portal &mdash; {_esc(env_label)}</div>
<div class="summary-box"><table>
  <tr><td><b>Generated:</b></td><td>{_esc(now_str)}</td>
      <td><b>Scan Completed:</b></td><td>{_esc(completed_at or 'N/A')}</td></tr>
  <tr><td><b>Environment:</b></td><td>{_esc(env_label)}</td>
      <td><b>Servers in Report:</b></td><td>{len(servers)}</td></tr>
  <tr><td><b>Scanned OK:</b></td><td>{n_ok}</td>
      <td><b>Scan Errors:</b></td><td>{n_err}</td></tr>
</table></div>
"""

    for srv in servers:
        name = srv['name']
        env  = srv.get('env', 'other')
        bg, fg = ENV_COLORS.get(env, ("#f0f0f0", "#666"))
        res  = results.get(name.upper())

        html += f'<div class="server-block">'
        html += f'<div class="srv-header">{_esc(name)}'
        html += (f'<span class="env-badge" style="background:{bg};color:{fg}">'
                 f'{_esc(ENV_LABELS.get(env, env))}</span>')
        if srv.get('os'):
            html += (f'<span style="font-weight:normal;font-size:9pt">'
                     f'&nbsp;&mdash;&nbsp;{_esc(srv["os"])}</span>')
        html += '</div><div class="srv-body">'

        if res is None:
            html += '<p class="no-data">Not included in scan.</p>'
        elif 'error' in res:
            html += f'<div class="error-box">Scan error: {_esc(res["error"])}</div>'
        elif not res.get('shares'):
            html += '<p class="no-data">No shares found (server offline or no accessible shares).</p>'
        else:
            for sh in res['shares']:
                sname    = sh.get('name', '')
                spath    = sh.get('path', '')
                sdesc    = sh.get('description', '')
                sh_acl   = sh.get('share_acl') or []
                ntfs_acl = sh.get('ntfs_acl')  or []
                ntfs_dir = [a for a in ntfs_acl if not a.get('inherited')]
                ntfs_inh = [a for a in ntfs_acl if a.get('inherited')]

                html += '<div class="share-block">'
                html += f'<div class="share-name">{_esc(sname)}</div>'
                parts = []
                if spath: parts.append(f'Path: {_esc(spath)}')
                if sdesc: parts.append(f'Desc: {_esc(sdesc)}')
                if parts:
                    html += f'<div class="share-meta">{" &mdash; ".join(parts)}</div>'

                html += '<div class="sec-label">Share Permissions</div>'
                if sh_acl:
                    html += '<table><tr><th>Account</th><th>Right</th><th>Type</th></tr>'
                    for a in sh_acl:
                        html += (f'<tr><td>{_esc(a.get("account",""))}</td>'
                                 f'<td>{_esc(a.get("right",""))}</td>'
                                 f'<td>{_esc(a.get("type",""))}</td></tr>')
                    html += '</table>'
                else:
                    html += '<p class="no-data">Not collected (requires WinRM — SMB-only scan mode).</p>'

                html += '<div class="sec-label">Folder (NTFS) Permissions — Explicit</div>'
                if ntfs_dir:
                    html += '<table><tr><th>Identity</th><th>Rights</th><th>Type</th></tr>'
                    for a in ntfs_dir:
                        html += (f'<tr><td>{_esc(a.get("identity",""))}</td>'
                                 f'<td>{_esc(a.get("rights",""))}</td>'
                                 f'<td>{_esc(a.get("type",""))}</td></tr>')
                    html += '</table>'
                else:
                    html += '<p class="no-data">No explicit NTFS entries (all inherited or unavailable).</p>'

                if ntfs_inh:
                    html += '<div class="sec-label">Folder (NTFS) Permissions — Inherited</div>'
                    html += '<table><tr><th>Identity</th><th>Rights</th><th>Type</th></tr>'
                    for a in ntfs_inh:
                        html += (f'<tr class="inh"><td>{_esc(a.get("identity",""))}</td>'
                                 f'<td>{_esc(a.get("rights",""))}</td>'
                                 f'<td>{_esc(a.get("type",""))}</td></tr>')
                    html += '</table>'

                html += '</div>'  # share-block

        html += '</div></div>\n'  # srv-body, server-block

    html += '</body></html>'
    return html


def generate_pdf(env_filter: str = "production") -> bytes:
    """Convert share audit HTML to PDF using xhtml2pdf (same as SOC report)."""
    try:
        from xhtml2pdf import pisa
    except ImportError:
        raise ImportError("xhtml2pdf is not installed. Run: pip install xhtml2pdf")
    html = generate_html(env_filter)
    buf  = io.BytesIO()
    result = pisa.CreatePDF(html, dest=buf, encoding="utf-8")
    if result.err:
        raise RuntimeError(f"xhtml2pdf: {result.err} error(s) during PDF generation")
    return buf.getvalue()


def generate_excel(env_filter: str = "production") -> bytes:
    """Generate share audit report as .xlsx using openpyxl."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    data         = get_results() or {"servers": [], "results": {}, "completed_at": ""}
    all_servers  = data.get("servers",  [])
    results      = data.get("results",  {})
    completed_at = data.get("completed_at", "")

    if env_filter and env_filter != "all":
        servers = [s for s in all_servers if s.get("env") == env_filter]
    else:
        servers = list(all_servers)
    servers = sorted(servers, key=lambda s: (_ENV_SORT.get(s.get("env", "other"), 99), s["name"].lower()))

    env_label = ENV_LABELS.get(env_filter, env_filter.title()) if env_filter and env_filter != "all" else "All Environments"
    now_str   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb = openpyxl.Workbook()

    # ── styles ────────────────────────────────────────────────────────────────
    hdr_font    = Font(bold=True, color="FFFFFF", size=10)
    hdr_fill    = PatternFill("solid", fgColor="003366")
    sub_fill    = PatternFill("solid", fgColor="E8EFF8")
    inh_font    = Font(color="999999", size=9)
    wrap        = Alignment(wrap_text=True, vertical="top")
    thin        = Border(bottom=Side(style="thin", color="CCCCCC"))
    center      = Alignment(horizontal="center", vertical="top")

    def _hdr_row(ws, row, values, fill=hdr_fill):
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = hdr_font; c.fill = fill; c.alignment = center

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 36
    ws_sum.column_dimensions["C"].width = 16
    ws_sum.column_dimensions["D"].width = 16
    ws_sum.column_dimensions["E"].width = 22

    ws_sum.merge_cells("A1:E1")
    title_cell = ws_sum["A1"]
    title_cell.value = f"Server Share Audit — {env_label}"
    title_cell.font  = Font(bold=True, size=13, color="003366")
    title_cell.alignment = Alignment(horizontal="left")

    ws_sum.append([])
    ws_sum.append(["Generated", now_str, "", "Scan Completed", completed_at or "N/A"])
    ws_sum.append(["Environment", env_label])
    ws_sum.append([])

    n_ok  = sum(1 for s in servers if s["name"].upper() in results and "error" not in results[s["name"].upper()])
    n_err = sum(1 for s in servers if s["name"].upper() in results and "error" in results[s["name"].upper()])
    ws_sum.append(["Servers in report", len(servers), "", "Scanned OK", n_ok])
    ws_sum.append(["Scan errors", n_err])
    ws_sum.append([])

    _hdr_row(ws_sum, ws_sum.max_row + 1,
             ["Server", "OS", "Environment", "Shares", "Status"])
    for srv in servers:
        name = srv["name"]
        res  = results.get(name.upper())
        if res is None:        status, n_shares = "Not scanned", ""
        elif "error" in res:   status, n_shares = f"Error: {res['error'][:60]}", ""
        else:                  status, n_shares = "OK", len(res.get("shares", []))
        ws_sum.append([name, srv.get("os",""), ENV_LABELS.get(srv.get("env","other"),"Other"), n_shares, status])

    # ── Sheet 2: Permissions (flat — one row per ACE) ─────────────────────────
    ws_perm = wb.create_sheet("Permissions")
    for col, (hdr, width) in enumerate([
        ("Server",18), ("Environment",14), ("Share",18), ("Path",30),
        ("Identity",36), ("Rights",18), ("Type",8), ("Inherited",10)
    ], 1):
        ws_perm.column_dimensions[ws_perm.cell(1, col).column_letter].width = width

    _hdr_row(ws_perm, 1, ["Server","Environment","Share","Path","Identity","Rights","Type","Inherited"])
    row_num = 2
    for srv in servers:
        name = srv["name"]
        env  = ENV_LABELS.get(srv.get("env","other"), "Other")
        res  = results.get(name.upper())
        if not res or "error" in res or not res.get("shares"):
            continue
        for sh in res["shares"]:
            for ace in (sh.get("ntfs_acl") or []):
                is_inh = ace.get("inherited", False)
                r = ws_perm.row_dimensions[row_num]
                for col, val in enumerate([
                    name, env, sh.get("name",""), sh.get("path",""),
                    ace.get("identity",""), ace.get("rights",""),
                    ace.get("type",""), "Yes" if is_inh else "No"
                ], 1):
                    c = ws_perm.cell(row=row_num, column=col, value=val)
                    c.alignment = wrap
                    c.border    = thin
                    if is_inh:
                        c.font = inh_font
                row_num += 1

    ws_perm.auto_filter.ref = f"A1:H{row_num - 1}"
    ws_perm.freeze_panes    = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
